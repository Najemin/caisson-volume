"""케이슨 체적 · 무게중심 산정  Ver.1.0   [Streamlit 판 · 단일 파일]

케이슨의 제원만 입력하면 3D 모델링(CAD) 없이 부재별 체적과 무게중심을 산정한다.

실행 :  streamlit run app.py

※ 이 파일 하나에 프로그램 전체가 들어 있다.
   원본 Tkinter 프로그램(케이슨체적_v1.py)에서
       · 계산부(§0~§4) 는 글자 하나 바꾸지 않고 그대로 옮겼고,
       · 화면부(§5)   만 Streamlit 으로 새로 썼다.
   그러므로 체적·무게중심·계산서·도면 결과는 원본과 완전히 같다.
   원본에서 손댄 곳은 셋뿐이다.
       (1) tkinter import 삭제           - 웹에는 창이 없다
       (2) matplotlib 'TkAgg' → 'Agg'    - 화면 없는 서버용 백엔드
       (3) 'Malgun Gothic' 못박기 → 설치된 한글 글꼴 자동 탐색 (리눅스 대응)


[산정 원리]  부호 있는 기본도형 분해 (signed primitive decomposition)
    케이슨은 직육면체·원기둥·삼각프리즘의 합(+)과 차(−)로 오차 없이 표현된다.
    메시·복셀 근사를 쓰지 않으므로 결과가 해석적으로 정확하고, 각 부재의
    산출식을 그대로 계산서에 적을 수 있다.

        V   = Σ sign_i · V_i
        Cg  = Σ (sign_i · V_i · c_i) / V          sign = +1 추가 / −1 공제

[좌표계]
        x : 케이슨 폭 방향.   전면(해측) 외면 = 0,  후면 외면 = B
        y : 케이슨 길이 방향. 좌측 단부 외면 = 0,  우측 단부 외면 = L
        z : 연직 방향.        저판 저면 = 0,       케이슨 정단 = H
                              (전단키는 z < 0, 상치는 z > H 구간을 차지한다)

    무게중심은 실무 계산서마다 기준이 달라 세 가지를 모두 출력한다.
        · Xg (전면 외면 기준) / Zg (저판 저면 기준)
        · e  = Xg − B/2  (폭 중앙 기준 편심) - 전도검토에 그대로 쓰인다
        · Yg (좌측 기준). 좌우 대칭이면 L/2 가 나오므로 입력 검증 지표가 된다

[지원 형식]
        · 무공 격실형 케이슨 (종격벽 · 횡격벽)
        · 유공(슬릿) 케이슨  (원형 / 사각 유공, 유수실 셀 지정)
        · 풋팅 · 헌치 · 전단키
        · 상치콘크리트 (본체 + 파라펫 + 전면 경사)

[자동 검산]  ※ 이 프로그램의 핵심 안전장치
    (1) 외곽체적 대조 - 부재 분해에서 유도한 공극과 셀 격자에서 유도한 공극을
        서로 다른 두 경로로 계산해 잔차를 본다. 격벽 교차부 중복공제 누락 같은
        오류가 즉시 드러난다. 매 계산마다 항상 수행한다.
    (2) 좌우 대칭 확인 - 좌우 대칭 입력이면 Yg = L/2 여야 한다.
    (3) 몬테카를로 형상 검증 ([형상 검증] 버튼) - 바운딩박스에 난수점을 뿌려
        부호 합으로 체적·도심을 독립 산정하고 해석해와 대조한다. 아울러 모든
        표본점의 부호 합이 0 또는 1 인지 확인해 중복·과공제를 직접 찾아낸다.

[본 프로그램의 범위 외]
    중량 · 부력 · 수중중량 · 흘수 · 경심고 GM · 활동/전도/지지력 검토,
    철근량, 부재 응력, 진수·예항 안정.
    ※ 무게중심 산정에 필요한 최소 입력으로 단위중량을 받지만(§재료 탭),
       이는 재료가 다른 부재군을 합성하기 위한 가중치일 뿐이다.

[적용 가정]
    · 헌치 교차부(코너)는 y방향 헌치를 우선하고 x방향 헌치를 각 단부에서
      a 만큼 절단해 중복을 제거한다. 이때 코너마다 누락되는 체적은 정확히
      a²b/6 이며, [헌치 코너 보정] 을 켜면 정확한 도심과 함께 되살린다.
    · 풋팅 두께는 저판 두께 이하여야 한다.
    · 속채움 상단이 헌치 상단보다 낮으면 헌치 공제가 과대해지므로 경고한다.

[구성]
    0. 공통 예외 · 표시 유틸
    1. 기하 기본도형 (Prim : Box / Cyl / TriPrism / Lump)
    2. 제원 (Spec) 및 케이슨 엔진 (CaissonEngine)
    3. 계산서 생성 (ReportBuilder)
    4. 도해 (Plotter) · 2D 도면 (DxfExporter) · 3D 모델 (Dxf3DExporter)

    5. 화면 (Streamlit)      ← 원본 §5 CaissonApp 을 갈아 끼운 부분

[원본 → Streamlit 대응]
    tk.StringVar             → st.session_state  (값도 원본과 같은 '문자열')
    ttk.Notebook 8 탭        → st.tabs 9 탭 (저장·불러오기 탭을 하나 더 둔다)
    ttk.Entry / Combobox     → st.text_input / st.selectbox
    비활성 (state='disabled')→ 위젯의 disabled 인자  (_update_gates 그대로 이식)
    messagebox               → st.error / st.warning / st.info
    filedialog (폴더 저장)   → st.download_button (항목별 + 전체 ZIP)
    filedialog (불러오기)    → st.file_uploader  (.cais 파일 그대로 호환)
    FigureCanvasTkAgg        → st.pyplot (3D 는 시선각 입력으로 회전한다)
    ScrolledText (고정폭)    → 고정폭 웹글꼴 <pre>  (표 정렬 규칙 _dw 동일)
    [셀 격자 갱신] 버튼      → 매번 자동 동기화 (버튼을 누를 필요가 없어졌다)
    [나가기]                 → 웹에는 없다. [기본 예제로 되돌리기] 로 대신한다.
"""

import copy
import html
import io
import json
import math
import os
import tempfile
import traceback
import unicodedata
import warnings
import zipfile
from dataclasses import dataclass, field
from typing import List, Tuple, Optional

import numpy as np

import matplotlib
# 서버(웹)에서 돌아가므로 화면 없는 백엔드를 쓴다. 원본의 'TkAgg' 자리다.
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle, Circle, Polygon as MplPolygon
from mpl_toolkits.mplot3d.art3d import Poly3DCollection, Line3DCollection

import streamlit as st
import streamlit.components.v1 as components



# ------------------------------------------------------------------
# 도해용 한글 글꼴
#
#   원본은 윈도우 전용으로 'Malgun Gothic' 을 못박아 두었다. Streamlit 은
#   리눅스 서버에서도 돌아가므로 설치된 것 중에서 골라 쓴다. 하나도 없으면
#   글꼴만 기본값으로 두고(네모 글자가 될 수 있다) 계산은 그대로 진행한다.
# ------------------------------------------------------------------
KOREAN_FONT_CANDIDATES = (
    'Malgun Gothic', 'AppleGothic', 'Apple SD Gothic Neo',
    'NanumGothic', 'Nanum Gothic', 'NanumBarunGothic',
    'Noto Sans CJK KR', 'Noto Sans KR', 'Source Han Sans KR',
    'Pretendard', 'D2Coding', 'UnDotum', 'Baekmuk Gulim', 'Gulim', 'Dotum',
)


def _find_hangul_font(font_manager):
    """설치된 글꼴 중 한글 글리프를 실제로 갖고 있는 것을 하나 고른다."""
    try:
        from matplotlib.ft2font import FT2Font
    except Exception:
        return None

    def rank(name):                     # 한글용으로 그럴듯한 순서
        for i, key in enumerate(('Nanum', 'CJK KR', 'Sans KR', 'CJK',
                                 'Gulim', 'Dotum', 'Batang', 'Gothic')):
            if key in name:
                return i
        return 99

    seen = set()
    for f in sorted(font_manager.fontManager.ttflist,
                    key=lambda f: (rank(f.name), f.name)):
        if rank(f.name) == 99 or f.name in seen:
            continue
        seen.add(f.name)
        try:
            face = FT2Font(f.fname)
            if face.get_char_index(0xAC00) and face.get_char_index(0xD7A3):
                return f.name           # '가' 와 '힣' 이 모두 있다
        except Exception:
            continue
    return None


def setup_matplotlib_font():
    """설치된 한글 글꼴을 찾아 matplotlib 에 물린다. 고른 이름을 돌려준다."""
    try:
        from matplotlib import font_manager
        have = {f.name for f in font_manager.fontManager.ttflist}
    except Exception:
        have = set()
    pick = next((f for f in KOREAN_FONT_CANDIDATES if f in have), None)
    if pick is None:
        # 이름표는 달라도 한글 글리프를 갖고 있는 글꼴을 실제로 찾아 쓴다.
        #   예) 서버에 'Noto Sans CJK KR' 은 없고 'Noto Sans CJK JP' 만 깔린 경우.
        #       통합 글꼴이라 한글이 온전히 들어 있는데 이름만 다르다.
        #   이름으로 짐작하지 않고 '가'·'힣' 이 실제로 들어 있는지 본다.
        pick = _find_hangul_font(font_manager)
    if pick:
        plt.rcParams['font.family'] = pick
    plt.rcParams['axes.unicode_minus'] = False      # 원본과 동일
    return pick


KOREAN_FONT = setup_matplotlib_font()


# ==========================================
# 0. 공통 예외 · 표시 유틸
# ==========================================
class InputError(Exception):
    """사용자 입력/모델 정합성 오류. 메시지를 그대로 사용자에게 보여준다."""
    pass


EPS = 1e-9


# 동아시아 '애매폭(Ambiguous)' 문자(· × ³ Σ γ π § → ■ 등)의 표시 폭.
#
#   같은 문자를 한글 고정폭 글꼴(돋움체·굴림체)은 2칸, 서양 글꼴(D2Coding·
#   Consolas)은 1칸으로 그린다. 계산서 표는 이 폭을 기준으로 열을 맞추므로
#   실제로 쓰는 글꼴에 맞춰야 한다. CaissonApp 이 글꼴을 고른 뒤 실측값으로
#   set_ambiguous_width() 를 불러 확정한다.
_AMB_W = 2


def _dw(s):
    """문자열의 표시 폭(한글=2, ASCII=1, 애매폭=_AMB_W)."""
    w = 0
    for ch in str(s):
        e = unicodedata.east_asian_width(ch)
        w += 2 if e in ('W', 'F') else (_AMB_W if e == 'A' else 1)
    return w


def set_ambiguous_width(n):
    """애매폭 규칙을 바꾸고 그 폭에 의존하는 구분선을 다시 만든다."""
    global _AMB_W, LINE3
    _AMB_W = 2 if n >= 2 else 1
    LINE3 = '·' * (W // _AMB_W)     # 애매폭 문자로 만든 유일한 구분선


def _pad(s, width, align='left'):
    """표시 폭 기준 패딩 (고정폭 글꼴에서 표가 어긋나지 않도록)."""
    s = str(s)
    gap = max(0, width - _dw(s))
    if align == 'right':
        return ' ' * gap + s
    if align == 'center':
        L = gap // 2
        return ' ' * L + s + ' ' * (gap - L)
    return s + ' ' * gap


def _cut(s, width):
    """표시 폭 기준으로 잘라 정확히 width 칸을 차지하게 만든다.

    잘린 것은 끝에 '..' 를 붙여 알린다. 표의 모든 행이 같은 폭이 되어야
    열이 어긋나 보이지 않는다.
    """
    s = str(s)
    if _dw(s) <= width:
        return _pad(s, width)
    out, w = '', 0
    for ch in s:
        cw = _dw(ch)
        if w + cw > width - 2:
            break
        out += ch
        w += cw
    return _pad(out + '..', width)


def _wrap(s, width, indent=''):
    """표시 폭 기준 줄바꿈. 경고문처럼 긴 문장을 계산서 폭에 맞춘다."""
    words, lines, cur = str(s).split(), [], ''
    for wd in words:
        nxt = (cur + ' ' + wd) if cur else wd
        if _dw(indent + nxt) > width and cur:
            lines.append(indent + cur)
            cur = wd
        else:
            cur = nxt
    if cur:
        lines.append(indent + cur)
    return lines or [indent]


def _fmt(v, nd=3, inf='-'):
    """무한대/None 을 안전하게 표기."""
    if v is None:
        return inf
    if isinstance(v, float) and (math.isinf(v) or math.isnan(v)):
        return '∞' if v > 0 else inf
    return f"{v:.{nd}f}"


# ==========================================
# 1. 기하 기본도형
#
#    각 도형은 다음 4가지를 제공한다.
#      volume()   해석적 체적
#      centroid() 해석적 도심 (x, y, z)
#      contains(P) 점 포함 판정 - 몬테카를로 검증 전용 (P: (N,3) ndarray)
#      bbox()     바운딩박스 (x0,x1,y0,y1,z0,z1)
#      desc()     계산서에 적을 산출식 문자열
#
#    압출 도형의 축 규약
#      axis='x' → 단면 평면좌표 (u,v) = (y,z), 압출축 w = x
#      axis='y' → 단면 평면좌표 (u,v) = (x,z), 압출축 w = y
#      axis='z' → 단면 평면좌표 (u,v) = (x,y), 압출축 w = z
# ==========================================

_AXIS_UVW = {'x': (1, 2, 0), 'y': (0, 2, 1), 'z': (0, 1, 2)}


class Prim:
    """기하 기본도형 공통 인터페이스."""

    #: contains() 를 지원하지 않는 도형(Lump)은 False 로 둔다.
    samplable = True

    def volume(self):
        raise NotImplementedError

    def centroid(self):
        raise NotImplementedError

    def contains(self, P):
        raise NotImplementedError

    def bbox(self):
        raise NotImplementedError

    def desc(self):
        return ''


class Box(Prim):
    """축에 나란한 직육면체. 저판·벽체·격벽·셀·사각슬릿·전단키에 쓴다."""

    def __init__(self, x0, y0, z0, dx, dy, dz):
        self.x0, self.y0, self.z0 = float(x0), float(y0), float(z0)
        self.dx, self.dy, self.dz = float(dx), float(dy), float(dz)

    def volume(self):
        return self.dx * self.dy * self.dz

    def centroid(self):
        return (self.x0 + self.dx / 2.0,
                self.y0 + self.dy / 2.0,
                self.z0 + self.dz / 2.0)

    def contains(self, P):
        x, y, z = P[:, 0], P[:, 1], P[:, 2]
        return ((x >= self.x0) & (x <= self.x0 + self.dx) &
                (y >= self.y0) & (y <= self.y0 + self.dy) &
                (z >= self.z0) & (z <= self.z0 + self.dz))

    def bbox(self):
        return (self.x0, self.x0 + self.dx,
                self.y0, self.y0 + self.dy,
                self.z0, self.z0 + self.dz)

    def desc(self):
        return f"{self.dx:.3f} × {self.dy:.3f} × {self.dz:.3f}"


class Cyl(Prim):
    """원기둥. 원형 유공(축 = x)에 쓴다.

    axis  : 압출축
    u, v  : 단면 평면상의 원 중심
    w0, h : 압출축 시작 좌표와 길이
    """

    def __init__(self, axis, u, v, w0, r, h):
        if axis not in _AXIS_UVW:
            raise InputError(f"Cyl: 알 수 없는 축 '{axis}'")
        self.axis = axis
        self.u, self.v = float(u), float(v)
        self.w0, self.r, self.h = float(w0), float(r), float(h)

    def volume(self):
        return math.pi * self.r ** 2 * self.h

    def centroid(self):
        iu, iv, iw = _AXIS_UVW[self.axis]
        c = [0.0, 0.0, 0.0]
        c[iu], c[iv], c[iw] = self.u, self.v, self.w0 + self.h / 2.0
        return tuple(c)

    def contains(self, P):
        iu, iv, iw = _AXIS_UVW[self.axis]
        du = P[:, iu] - self.u
        dv = P[:, iv] - self.v
        w = P[:, iw]
        return ((du * du + dv * dv <= self.r ** 2) &
                (w >= self.w0) & (w <= self.w0 + self.h))

    def bbox(self):
        iu, iv, iw = _AXIS_UVW[self.axis]
        lo = [0.0, 0.0, 0.0]
        hi = [0.0, 0.0, 0.0]
        lo[iu], hi[iu] = self.u - self.r, self.u + self.r
        lo[iv], hi[iv] = self.v - self.r, self.v + self.r
        lo[iw], hi[iw] = self.w0, self.w0 + self.h
        return (lo[0], hi[0], lo[1], hi[1], lo[2], hi[2])

    def desc(self):
        return f"π × {self.r:.3f}² × {self.h:.3f}"


class TriPrism(Prim):
    """직각삼각형 단면 압출체. 헌치와 상치 경사면에 쓴다.

    단면은 직각점 (u0, v0) 과 두 다리 (du, dv) 로 정의한다.
    du·dv 는 음수를 허용하며 부호가 곧 다리의 방향이다.
        꼭짓점 : (u0, v0), (u0+du, v0), (u0, v0+dv)
        면적   : |du·dv| / 2
        도심   : (u0 + du/3, v0 + dv/3)
    """

    def __init__(self, axis, u0, v0, du, dv, w0, h):
        if axis not in _AXIS_UVW:
            raise InputError(f"TriPrism: 알 수 없는 축 '{axis}'")
        self.axis = axis
        self.u0, self.v0 = float(u0), float(v0)
        self.du, self.dv = float(du), float(dv)
        self.w0, self.h = float(w0), float(h)

    def volume(self):
        return abs(self.du * self.dv) / 2.0 * self.h

    def centroid(self):
        iu, iv, iw = _AXIS_UVW[self.axis]
        c = [0.0, 0.0, 0.0]
        c[iu] = self.u0 + self.du / 3.0
        c[iv] = self.v0 + self.dv / 3.0
        c[iw] = self.w0 + self.h / 2.0
        return tuple(c)

    def contains(self, P):
        iu, iv, iw = _AXIS_UVW[self.axis]
        if abs(self.du) < EPS or abs(self.dv) < EPS:
            return np.zeros(len(P), dtype=bool)
        s = (P[:, iu] - self.u0) / self.du
        t = (P[:, iv] - self.v0) / self.dv
        w = P[:, iw]
        return ((s >= 0.0) & (t >= 0.0) & (s + t <= 1.0) &
                (w >= self.w0) & (w <= self.w0 + self.h))

    def bbox(self):
        iu, iv, iw = _AXIS_UVW[self.axis]
        lo = [0.0, 0.0, 0.0]
        hi = [0.0, 0.0, 0.0]
        lo[iu], hi[iu] = min(self.u0, self.u0 + self.du), max(self.u0, self.u0 + self.du)
        lo[iv], hi[iv] = min(self.v0, self.v0 + self.dv), max(self.v0, self.v0 + self.dv)
        lo[iw], hi[iw] = self.w0, self.w0 + self.h
        return (lo[0], hi[0], lo[1], hi[1], lo[2], hi[2])

    def desc(self):
        return f"1/2 × {abs(self.du):.3f} × {abs(self.dv):.3f} × {self.h:.3f}"


class Lump(Prim):
    """체적과 도심만 주어진 보정항.

    헌치 코너 보정처럼 체적·도심은 해석적으로 정확히 알지만 단순 도형으로
    표현되지 않는 영역에 쓴다. 점 포함 판정을 할 수 없으므로 몬테카를로
    검증에서는 제외되며, 그 사실을 계산서와 검증 결과에 명시한다.
    """

    samplable = False

    def __init__(self, v, cx, cy, cz, text=''):
        self.v = float(v)
        self.c = (float(cx), float(cy), float(cz))
        self.text = text

    def volume(self):
        return self.v

    def centroid(self):
        return self.c

    def contains(self, P):
        raise NotImplementedError("Lump 는 점 포함 판정을 지원하지 않는다.")

    def bbox(self):
        x, y, z = self.c
        return (x, x, y, y, z, z)

    def desc(self):
        return self.text


# ------------------------------------------------------------------
# 부재 (Part) - 도형 하나에 이름·부재군·부호를 붙인 것
# ------------------------------------------------------------------
GRP_CONC = 'CONC'       # 케이슨 콘크리트
GRP_FILL = 'FILL'       # 속채움
GRP_COPING = 'COPING'   # 상치콘크리트

GRP_LABEL = {GRP_CONC: '케이슨 콘크리트', GRP_FILL: '속채움', GRP_COPING: '상치콘크리트'}


@dataclass
class Part:
    name: str
    group: str
    sign: int
    prim: Prim
    note: str = ''
    #: 외곽 직육면체(B×L×H) 안에 들어가는 부재인지. 검산 대상 판별에 쓴다.
    inner: bool = True
    #: 계산서 요약 표기에서 한 줄로 묶을 소계 이름. 비어 있으면 개별 표기한다.
    roll: str = ''

    @property
    def v(self):
        """부호를 반영한 체적."""
        return self.sign * self.prim.volume()

    @property
    def c(self):
        return self.prim.centroid()


def mass_props(parts):
    """부재 목록의 (체적, Cx, Cy, Cz). 체적이 0 이면 도심은 0 을 돌려준다."""
    V = 0.0
    mx = my = mz = 0.0
    for p in parts:
        v = p.v
        cx, cy, cz = p.c
        V += v
        mx += v * cx
        my += v * cy
        mz += v * cz
    if abs(V) < 1e-12:
        return 0.0, 0.0, 0.0, 0.0
    return V, mx / V, my / V, mz / V


# ==========================================
# 2. 제원 (Spec) 및 케이슨 엔진
# ==========================================

TYPE_SOLID, TYPE_PERF = '무공', '유공'
CELL_FILL, CELL_CHAMBER, CELL_EMPTY = '속채움', '유수실', '공셀'
CELL_TYPES = [CELL_FILL, CELL_CHAMBER, CELL_EMPTY]

MODE_EVEN, MODE_MANUAL = '등간격', '직접입력'
HOLE_CIRCLE, HOLE_RECT = '원형', '사각'
HOLE_GRID = '격자'
COL_CELL, COL_EVEN = '셀중앙', '등간격'
SIDE_FRONT, SIDE_REAR = '전면', '후면'

# 유공을 뚫을 수 있는 벽. 관통축이 다르다.
#   x축 관통 : 전면벽 · 후면벽 · 종격벽   (단면 평면좌표 (u, v) = (y, z))
#   y축 관통 : 좌측벽 · 우측벽 · 횡격벽   (단면 평면좌표 (u, v) = (x, z))
WALL_FRONT = '전면벽'
WALL_REAR = '후면벽'
WALL_LEFT = '좌측벽'
WALL_RIGHT = '우측벽'
WALL_OUTER = (WALL_FRONT, WALL_REAR, WALL_LEFT, WALL_RIGHT)
#: 벽 이름 → 관통축
WALL_AXIS = {WALL_FRONT: 'x', WALL_REAR: 'x', WALL_LEFT: 'y', WALL_RIGHT: 'y'}


def wall_axis(name):
    """벽 이름 → 유공 관통축. 격벽은 이름 앞머리로 가른다."""
    if name in WALL_AXIS:
        return WALL_AXIS[name]
    return 'x' if str(name).startswith('종격벽') else 'y'

# KDS 64 10 10 재료표 (수상 단위중량, kN/m3)
G_RC = 24.00        # 철근콘크리트
G_PC = 22.60        # 무근콘크리트
G_STONE = 18.00     # 사석·잡석


@dataclass
class PartitionSpec:
    """격벽 1매. pos 는 중심 좌표(종격벽=x, 횡격벽=y), h=0 이면 벽체 전高."""
    pos: float = 0.0
    t: float = 0.4
    h: float = 0.0


@dataclass
class HoleSpec:
    """유공 1개의 중심 좌표 (직접입력 모드)."""
    y: float = 0.0
    z: float = 0.0


@dataclass
class KeySpec:
    """전단키 1개. x=전면에서의 시점, w=폭, d=저판 저면 아래 깊이."""
    x: float = 0.0
    w: float = 0.5
    d: float = 0.5


@dataclass
class Spec:
    """케이슨 제원 일체."""
    # --- 기본 ---
    project: str = 'OO항 방파제 축조공사'
    name: str = '표준 케이슨 (유공형)'
    ctype: str = TYPE_PERF
    B: float = 15.0          # 폭 (전후 방향)
    L: float = 20.0          # 길이 (연장 방향)
    H: float = 15.0          # 높이 (저판 저면 → 정단)
    tb: float = 1.0          # 저판 두께
    tf: float = 0.5          # 전면벽 두께 (해측)
    tr: float = 0.5          # 후면벽 두께 (항내측)
    ts_l: float = 0.5        # 좌측벽 두께 (y = 0 쪽)
    ts_r: float = 0.5        # 우측벽 두께 (y = L 쪽)

    # --- 격벽 ---
    lp_mode: str = MODE_EVEN
    lp_n: int = 2
    lp_t: float = 0.40
    lp_h: float = 0.0
    lp_list: List[PartitionSpec] = field(default_factory=list)
    tp_mode: str = MODE_EVEN
    tp_n: int = 3
    tp_t: float = 0.40
    tp_h: float = 0.0
    tp_list: List[PartitionSpec] = field(default_factory=list)

    # --- 셀 (속채움) ---
    #: 기본(일괄) 채움 상단 z. 셀별 지정이 없는 속채움 셀에 적용한다.
    fill_top: float = 15.0
    cell_types: List[str] = field(default_factory=list)
    #: 셀별 채움 상단 z. cell_types 와 같은 순서.
    #:   None  = 미지정 → 구분에 따른 기본값 (속채움 = fill_top, 그 외 = 채우지 않음)
    #:   0 이하 = 채우지 않음.   유수실도 값을 주면 그 높이까지 채운다.
    cell_tops: List[Optional[float]] = field(default_factory=list)
    #: 기본(일괄) 유수실 덮개 두께. 셀별 지정이 없는 유수실에 적용한다.
    cover_t: float = 0.0
    #: 셀별 덮개 두께. cell_types 와 같은 순서.
    #:   None = 미지정 → 유수실이면 cover_t, 그 외 0.   0 이면 덮개 없음.
    cell_covers: List[Optional[float]] = field(default_factory=list)

    # --- 유공 ---
    hole_on: bool = True
    #: 유공을 뚫을 벽 목록. 형식·치수·행 배치는 모든 벽이 공용으로 쓰고,
    #: 열 배치만 벽마다 그 벽의 칸에 맞춰 자동 정렬한다.
    hole_walls: List[str] = field(default_factory=lambda: [WALL_FRONT])
    hole_shape: str = HOLE_CIRCLE
    hole_mode: str = HOLE_GRID
    hole_d: float = 2.00      # 원형 직경
    hole_w: float = 1.50      # 사각 폭 (y방향)
    hole_hh: float = 2.50     # 사각 높이 (z방향)
    row_n: int = 3
    row_z0: float = 5.0
    row_dz: float = 3.0
    col_mode: str = COL_CELL
    #: 셀중앙 배치일 때 격실(칸) 하나에 넣을 유공 개수. 1 이면 칸 중앙에 1 개다.
    col_per_cell: int = 1
    #: 격실 안 유공 중심간 거리. 0 이면 칸 순폭을 개수만큼 등분한 중앙에 놓는다.
    col_gap: float = 0.0
    col_n: int = 4
    col_y0: float = 2.5
    col_dy: float = 5.0
    hole_list: List[HoleSpec] = field(default_factory=list)

    # --- 풋팅 ---
    ft_front: float = 0.50
    ft_rear: float = 0.50
    ft_side: float = 0.00
    ft_t: float = 0.50

    # --- 헌치 ---
    hn_on: bool = True        # 수평 헌치 (저판-벽체 접합부)
    hn_a: float = 0.50        # 수평 다리
    hn_b: float = 0.50        # 연직 다리
    hn_corner: bool = True    # 코너 보정 적용 (수평·수직 공통)
    vh_on: bool = True        # 수직 헌치 (우각부 - 격벽·외벽 연직 모서리)
    vh_c: float = 0.30        # 수직 헌치 다리 (45° 이므로 x·y 공통)

    # --- 전단키 ---
    keys: List[KeySpec] = field(default_factory=list)

    # --- 상치 ---
    cp_on: bool = True
    cp_x0: float = 0.0
    cp_w: float = 15.0
    cp_h: float = 2.00
    cp_z0: float = 0.0        # 0 이면 케이슨 정단 H
    cp_par_on: bool = True
    cp_par_side: str = SIDE_FRONT
    cp_par_w: float = 1.00
    cp_par_h: float = 1.50
    cp_sl_a: float = 0.30     # 전면 상단 경사 (수평)
    cp_sl_b: float = 0.30     # 전면 상단 경사 (연직)

    # --- 재료 (합성 무게중심용) ---
    use_gamma: bool = True
    g_conc: float = G_RC
    g_fill: float = G_STONE
    g_cop: float = G_PC


class CaissonEngine:
    """제원(Spec) → 부재(Part) 분해 → 체적·무게중심."""

    def __init__(self, sp: Spec):
        self.sp = sp
        self.parts: List[Part] = []
        self.cells: List[dict] = []
        self.lps: List[dict] = []
        self.tps: List[dict] = []
        self.holes: List[dict] = []
        self.warns: List[str] = []
        self.q: dict = {}
        self.check: dict = {}

    # ------------------------------------------------------------------
    # 파생 제원
    # ------------------------------------------------------------------
    @property
    def hw(self):
        """벽체 높이 = H − 저판 두께."""
        return self.sp.H - self.sp.tb

    @property
    def x_in0(self):
        """내부 순공간의 전면측 x (전면벽 내면)."""
        return self.sp.tf

    @property
    def x_in1(self):
        return self.sp.B - self.sp.tr

    @property
    def y_in0(self):
        return self.sp.ts_l

    @property
    def y_in1(self):
        return self.sp.L - self.sp.ts_r

    # ------------------------------------------------------------------
    # 입력 검증
    # ------------------------------------------------------------------
    def _validate_basic(self):
        sp = self.sp
        for label, v in (('폭 B', sp.B), ('길이 L', sp.L), ('높이 H', sp.H),
                         ('저판 두께', sp.tb), ('전면벽 두께', sp.tf),
                         ('후면벽 두께', sp.tr), ('좌측벽 두께', sp.ts_l),
                         ('우측벽 두께', sp.ts_r)):
            if v <= 0:
                raise InputError(f"[기본제원] {label} 는 0 보다 커야 합니다. (입력 {v})")
        if sp.tb >= sp.H:
            raise InputError(f"[기본제원] 저판 두께 {sp.tb:.3f}m 가 전체 높이 "
                             f"{sp.H:.3f}m 이상입니다.")
        if sp.tf + sp.tr >= sp.B:
            raise InputError(f"[기본제원] 전면벽+후면벽 두께 "
                             f"({sp.tf:.3f}+{sp.tr:.3f}={sp.tf + sp.tr:.3f}m) 가 "
                             f"폭 B={sp.B:.3f}m 이상입니다.")
        if sp.ts_l + sp.ts_r >= sp.L:
            raise InputError(f"[기본제원] 좌+우측벽 두께 "
                             f"({sp.ts_l:.3f}+{sp.ts_r:.3f}="
                             f"{sp.ts_l + sp.ts_r:.3f}m) 가 "
                             f"길이 L={sp.L:.3f}m 이상입니다.")

    def _build_partitions(self, mode, n, t, h, lst, lo, hi, label):
        """격벽 목록 생성 및 검증. 반환 원소 = dict(pos, t, h, lo, hi)."""
        out = []
        if mode == MODE_EVEN:
            if n > 0:
                if t <= 0:
                    raise InputError(f"[격벽] {label} 두께가 0 이하입니다. (입력 {t})")
                step = (hi - lo) / (n + 1)
                for i in range(n):
                    out.append(dict(pos=lo + step * (i + 1), t=t,
                                    h=(h if h > 0 else self.hw)))
        else:
            for i, p in enumerate(lst):
                if p.t <= 0:
                    raise InputError(f"[격벽] {label} #{i + 1} 두께가 0 이하입니다.")
                out.append(dict(pos=p.pos, t=p.t,
                                h=(p.h if p.h > 0 else self.hw)))
        out.sort(key=lambda d: d['pos'])
        for i, d in enumerate(out):
            d['lo'] = d['pos'] - d['t'] / 2.0
            d['hi'] = d['pos'] + d['t'] / 2.0
            if d['h'] <= 0 or d['h'] > self.hw + EPS:
                raise InputError(f"[격벽] {label} #{i + 1} 높이 {d['h']:.3f}m 가 "
                                 f"벽체 높이 {self.hw:.3f}m 를 벗어납니다.")
            if d['lo'] < lo - EPS or d['hi'] > hi + EPS:
                raise InputError(
                    f"[격벽] {label} #{i + 1} 이 순내부 구간 "
                    f"[{lo:.3f}, {hi:.3f}] 을 벗어납니다. "
                    f"(중심 {d['pos']:.3f}, 두께 {d['t']:.3f} → "
                    f"{d['lo']:.3f}~{d['hi']:.3f})")
        for i in range(len(out) - 1):
            if out[i]['hi'] > out[i + 1]['lo'] - EPS:
                raise InputError(
                    f"[격벽] {label} #{i + 1} 과 #{i + 2} 가 서로 겹치거나 맞닿습니다. "
                    f"({out[i]['hi']:.3f} ≥ {out[i + 1]['lo']:.3f})")
        return out

    # ------------------------------------------------------------------
    # 셀 격자
    # ------------------------------------------------------------------
    def _spans(self, parts, lo, hi):
        """격벽으로 분할된 순내부 구간 목록 [(a, b, ha, hb), ...].

        ha, hb 는 그 구간의 양쪽 경계벽 높이다. 외벽은 항상 전高(hw) 이고
        감격벽(부분 높이 격벽)이면 그 격벽의 높이가 된다. 수직 헌치를 코너를
        이루는 두 벽 중 낮은 쪽까지만 세우기 위해 필요하다.
        """
        out = []
        cur, hcur = lo, self.hw
        for d in parts:
            out.append((cur, d['lo'], hcur, d['h']))
            cur, hcur = d['hi'], d['h']
        out.append((cur, hi, hcur, self.hw))
        return out

    def _default_cell_type(self, ix):
        """셀 구분 기본값 - 유공형식이면 최전열을 유수실로 본다."""
        if self.sp.ctype == TYPE_PERF and ix == 0:
            return CELL_CHAMBER
        return CELL_FILL

    def _default_fill_top(self, ct):
        """셀별 채움 상단이 지정되지 않았을 때의 기본값.

        속채움 셀만 일괄값(fill_top)까지 채운다. 유수실·공셀은 채우지 않는 것이
        기본이며, 유수실은 셀별로 값을 주면 그 높이까지 채워진다.
        """
        return self.sp.fill_top if ct == CELL_FILL else 0.0

    def _make_cells(self):
        sp = self.sp
        xs = self._spans(self.lps, self.x_in0, self.x_in1)
        ys = self._spans(self.tps, self.y_in0, self.y_in1)
        # 유공 배치 등 나머지 코드는 (a, b) 만 쓰므로 형태를 유지한다
        self.xspans = [(a, b) for a, b, _h0, _h1 in xs]
        self.yspans = [(a, b) for a, b, _h0, _h1 in ys]
        types = list(sp.cell_types)
        tops = list(sp.cell_tops)
        covs = list(sp.cell_covers)
        cells = []
        k = 0
        for ix, (x0, x1, hx0, hx1) in enumerate(xs):
            for iy, (y0, y1, hy0, hy1) in enumerate(ys):
                ct = types[k] if k < len(types) else self._default_cell_type(ix)
                if ct not in CELL_TYPES:
                    ct = self._default_cell_type(ix)
                ft = tops[k] if k < len(tops) else None
                if ct == CELL_EMPTY:
                    ft = 0.0                    # 공셀은 정의상 채우지 않는다
                elif ft is None:
                    ft = self._default_fill_top(ct)
                cv = covs[k] if k < len(covs) else None
                if cv is None:                  # 미지정 - 유수실만 덮는다
                    cv = sp.cover_t if ct == CELL_CHAMBER else 0.0
                cells.append(dict(name=f"C{ix + 1}-{iy + 1}", ix=ix, iy=iy,
                                  x0=x0, x1=x1, y0=y0, y1=y1, ctype=ct,
                                  fill_top=float(ft), cover=max(0.0, float(cv)),
                                  hx0=hx0, hx1=hx1, hy0=hy0, hy1=hy1))
                k += 1
        self.cells = cells

    # ------------------------------------------------------------------
    # 유공 배치
    # ------------------------------------------------------------------
    def hole_wall_list(self):
        """유공을 뚫을 벽 이름 목록 (중복 제거, 정의된 순서)."""
        sp = self.sp
        if not sp.hole_on:
            return []
        order = list(WALL_OUTER) + [f"종격벽 #{i + 1}" for i in range(len(self.lps))] \
            + [f"횡격벽 #{j + 1}" for j in range(len(self.tps))]
        want = list(sp.hole_walls)
        out = [w for w in order if w in want]
        bad = [w for w in want if w not in order]
        if bad:
            raise InputError(
                f"[유공] 대상 벽체 {', '.join(bad)} 을(를) 찾을 수 없습니다. "
                f"종격벽 {len(self.lps)} 매, 횡격벽 {len(self.tps)} 매입니다. "
                f"[대상 벽체 목록 갱신] 을 눌러 다시 고르십시오.")
        return out

    def wall_geom(self, name):
        """유공 대상 벽의 기하. 축에 무관한 (u, v, w) 좌표계로 돌려준다.

            axis='x' : w = x, (u, v) = (y, z)      전면·후면·종격벽
            axis='y' : w = y, (u, v) = (x, z)      좌측·우측·횡격벽

        spans 는 그 벽을 따라 늘어선 칸 목록으로, 셀중앙 열 배치에 쓴다.
        """
        sp, e = self.sp, self
        ax = wall_axis(name)
        if name == WALL_FRONT:
            w0, t, u0, u1 = 0.0, sp.tf, 0.0, sp.L
        elif name == WALL_REAR:
            w0, t, u0, u1 = sp.B - sp.tr, sp.tr, 0.0, sp.L
        elif name == WALL_LEFT:
            # 측벽은 전·후면벽 사이 구간만 차지한다 (모서리 중복을 없애려고)
            w0, t, u0, u1 = 0.0, sp.ts_l, e.x_in0, e.x_in1
        elif name == WALL_RIGHT:
            w0, t, u0, u1 = sp.L - sp.ts_r, sp.ts_r, e.x_in0, e.x_in1
        else:
            lst, lab = (self.lps, '종격벽') if ax == 'x' else (self.tps, '횡격벽')
            try:
                d = lst[int(str(name).split('#')[-1]) - 1]
            except Exception:
                raise InputError(f"[유공] 대상 벽체 '{name}' 을 찾을 수 없습니다. "
                                 f"{lab}이 {len(lst)} 매뿐입니다.")
            w0, t = d['lo'], d['t']
            u0, u1 = ((e.y_in0, e.y_in1) if ax == 'x' else (e.x_in0, e.x_in1))
            return dict(name=name, axis=ax, w0=w0, t=t, u0=u0, u1=u1,
                        v0=sp.tb, v1=sp.tb + d['h'],
                        spans=(e.yspans if ax == 'x' else e.xspans),
                        face=(u1 - u0) * d['h'])
        return dict(name=name, axis=ax, w0=w0, t=t, u0=u0, u1=u1,
                    v0=sp.tb, v1=sp.H,
                    spans=(e.yspans if ax == 'x' else e.xspans),
                    face=(u1 - u0) * self.hw)

    def wall_names(self):
        """유공을 뚫을 수 있는 벽 이름 전체 (외벽 4 + 격벽)."""
        return (list(WALL_OUTER)
                + [f"종격벽 #{i + 1}" for i in range(len(self.lps))]
                + [f"횡격벽 #{j + 1}" for j in range(len(self.tps))])

    def _hole_prim(self, g, u, v, hu, hv):
        """유공 도형 하나. 축에 따라 Cyl/Box 인수 순서를 맞춘다."""
        sp = self.sp
        if sp.hole_shape == HOLE_CIRCLE:
            return Cyl(g['axis'], u, v, g['w0'], hu, g['t'])
        if g['axis'] == 'x':
            return Box(g['w0'], u - hu, v - hv, g['t'], 2 * hu, 2 * hv)
        return Box(u - hu, g['w0'], v - hv, 2 * hu, g['t'], 2 * hv)

    def _cell_centered_cols(self, g):
        """셀중앙 배치의 유공 중심 u 목록 (벽 하나에 대해).

        간격 p 를 주면 칸 중앙을 기준으로 대칭 배치하고, 0 이면 칸 순폭을
        개수만큼 등분한 중앙에 놓는다. 칸마다 순폭이 다르므로 벽 전체 균일
        피치(등간격)로는 만들 수 없는 배치다.

            p > 0 : u_k = 칸중앙 + (k − (n−1)/2)·p
            p = 0 : u_k = 칸시작 + 순폭 × (2k+1) / (2n)

        들어가지 않는 치수를 넣으면 겹침 오류 대신 허용 한계를 알려준다.
        """
        sp = self.sp
        n = sp.col_per_cell
        if n <= 0:
            raise InputError(f"[유공] 격실당 개수가 {n} 입니다. 1 이상을 넣으세요.")
        wide = sp.hole_d if sp.hole_shape == HOLE_CIRCLE else sp.hole_w
        label = '직경' if sp.hole_shape == HOLE_CIRCLE else '폭'
        p = sp.col_gap
        us = []
        for j, (a, b) in enumerate(g['spans'], start=1):
            w = b - a
            where = f"[유공] {g['name']} 칸 #{j} ({a:.3f} ~ {b:.3f}, 순폭 {w:.3f}m)"
            if n > 1 and p > EPS:
                need = (n - 1) * p + wide
                if need >= w - EPS:
                    raise InputError(
                        f"{where} 에 {n} 개를 간격 {p:.3f}m 로 넣으면 "
                        f"{n - 1} × {p:.3f} + {wide:.3f} = {need:.3f}m 가 필요해 "
                        f"칸을 벗어납니다.\n"
                        f"간격을 {(w - wide) / (n - 1):.3f}m 미만으로 하거나 "
                        f"유공 {label}을 줄이십시오.")
                if p < wide - EPS:
                    raise InputError(
                        f"{where} - 유공 간격 {p:.3f}m 가 {label} {wide:.3f}m 보다 "
                        f"작아 유공끼리 겹칩니다.")
                c = (a + b) / 2.0
                us += [c + (k - (n - 1) / 2.0) * p for k in range(n)]
                continue
            if n > 1 and wide * n >= w - EPS:
                raise InputError(
                    f"{where} 에 {n} 개를 넣으려면 유공 {label}이 "
                    f"{w / n:.3f}m 미만이어야 합니다. "
                    f"(현재 {wide:.3f}m → {n} × {wide:.3f} = {wide * n:.3f}m)\n"
                    f"유공을 줄이거나, 격실당 개수를 줄이거나, 격벽을 넓히십시오.")
            us += [a + w * (2 * k + 1) / (2 * n) for k in range(n)]
        return us

    def _build_holes(self):
        """대상 벽마다 유공을 배치한다.

        형식·치수·행 배치는 모든 벽이 공용이고, 열 배치만 벽마다 그 벽의 칸에
        맞춰 정렬한다. 겹침 검사는 **같은 벽 안에서만** 한다 - 전면벽과 후면벽의
        유공이 같은 (u, v) 에 있어도 서로 다른 벽이라 겹치지 않는다.
        """
        sp = self.sp
        self.holes = []
        self.wall_geoms = []
        if not sp.hole_on:
            return
        walls = self.hole_wall_list()
        if not walls:
            raise InputError("[유공] 대상 벽체를 하나도 고르지 않았습니다. "
                             "유공을 끄거나 벽체를 고르십시오.")
        circle = (sp.hole_shape == HOLE_CIRCLE)
        if circle and sp.hole_d <= 0:
            raise InputError(f"[유공] 직경 {sp.hole_d:.3f}m 가 0 이하입니다.")
        if not circle and (sp.hole_w <= 0 or sp.hole_hh <= 0):
            raise InputError("[유공] 사각 유공의 폭·높이가 0 이하입니다.")
        hu = sp.hole_d / 2.0 if circle else sp.hole_w / 2.0
        hv = sp.hole_d / 2.0 if circle else sp.hole_hh / 2.0

        idx = 0
        for name in walls:
            g = self.wall_geom(name)
            self.wall_geoms.append(g)
            ulab, vlab = ('y', 'z') if g['axis'] == 'x' else ('x', 'z')

            # --- 중심 좌표 목록 ---
            if sp.hole_mode == MODE_MANUAL:
                centers = [(h.y, h.z) for h in sp.hole_list]
            else:
                if sp.row_n <= 0:
                    raise InputError("[유공] 행 개수가 0 입니다. "
                                     "유공을 끄거나 개수를 넣으세요.")
                vs = [sp.row_z0 + i * sp.row_dz for i in range(sp.row_n)]
                if sp.col_mode == COL_CELL:
                    us = self._cell_centered_cols(g)
                else:
                    if sp.col_n <= 0:
                        raise InputError("[유공] 열 개수가 0 입니다.")
                    us = [sp.col_y0 + j * sp.col_dy for j in range(sp.col_n)]
                centers = [(u, v) for v in vs for u in us]
            if not centers:
                raise InputError(f"[유공] {name} 에 배치된 유공이 하나도 없습니다.")

            first = len(self.holes)
            for u, v in centers:
                idx += 1
                if u - hu < g['u0'] - EPS or u + hu > g['u1'] + EPS:
                    raise InputError(
                        f"[유공] {name} #{idx} ({ulab}={u:.3f}, {vlab}={v:.3f}) 가 "
                        f"그 벽의 {ulab} 범위 [{g['u0']:.3f}, {g['u1']:.3f}] 를 "
                        f"벗어납니다.")
                if v - hv < g['v0'] - EPS or v + hv > g['v1'] + EPS:
                    raise InputError(
                        f"[유공] {name} #{idx} ({ulab}={u:.3f}, {vlab}={v:.3f}) 가 "
                        f"그 벽의 {vlab} 범위 [{g['v0']:.3f}, {g['v1']:.3f}] 를 "
                        f"벗어납니다.")
                self.holes.append(dict(
                    i=idx, wall=name, axis=g['axis'], w0=g['w0'], t=g['t'],
                    u=u, v=v, hu=hu, hv=hv, circle=circle,
                    prim=self._hole_prim(g, u, v, hu, hv)))

            # --- 같은 벽 안에서 겹침 (겹치면 이중 공제가 되어 체적이 틀린다) ---
            grp = self.holes[first:]
            for a in range(len(grp)):
                for b in range(a + 1, len(grp)):
                    A, Bh = grp[a], grp[b]
                    if circle:
                        dist = math.hypot(A['u'] - Bh['u'], A['v'] - Bh['v'])
                        if dist < A['hu'] + Bh['hu'] - EPS:
                            raise InputError(
                                f"[유공] {name} #{A['i']} 와 #{Bh['i']} 가 서로 "
                                f"겹칩니다. 중심거리 {dist:.3f}m < "
                                f"직경 {sp.hole_d:.3f}m")
                    elif (abs(A['u'] - Bh['u']) < A['hu'] + Bh['hu'] - EPS and
                          abs(A['v'] - Bh['v']) < A['hv'] + Bh['hv'] - EPS):
                        raise InputError(f"[유공] {name} #{A['i']} 와 #{Bh['i']} 가 "
                                         f"서로 겹칩니다.")

            # --- 직교하는 격벽과의 간섭 ---
            #   외벽이면 격벽이 유공 뒤를 막을 뿐이라 체적에는 영향이 없다 → 경고.
            #   대상이 격벽이면 유공이 교차부를 관통해 '교차부 공제' 와 겹쳐
            #   이중 공제가 되므로 → 오류.
            cross, clab = ((self.tps, '횡격벽') if g['axis'] == 'x'
                           else (self.lps, '종격벽'))
            outer = name in WALL_OUTER
            for h in grp:
                for j, d in enumerate(cross):
                    if not (h['u'] + h['hu'] > d['lo'] + EPS and
                            h['u'] - h['hu'] < d['hi'] - EPS):
                        continue
                    if outer:
                        self.warns.append(
                            f"{name} 유공 #{h['i']} ({ulab}={h['u']:.3f}m) 가 "
                            f"{clab} #{j + 1} ({d['lo']:.3f}~{d['hi']:.3f}m) 바로 "
                            f"앞에 놓입니다. 체적 산정에는 영향이 없으나 유공이 "
                            f"격벽으로 막히므로 배치를 재검토하십시오.")
                    else:
                        raise InputError(
                            f"[유공] {name} 의 유공 #{h['i']} "
                            f"({ulab}={h['u']:.3f}m) 가 {clab} #{j + 1} "
                            f"({d['lo']:.3f}~{d['hi']:.3f}m) 와 만나는 교차부를 "
                            f"관통합니다.\n격벽 교차부는 이미 중복 공제하므로 "
                            f"유공까지 겹치면 체적이 이중으로 빠집니다. "
                            f"유공을 교차부 밖으로 옮기십시오.")
        self.face_area = sum(g['face'] for g in self.wall_geoms)

    # ------------------------------------------------------------------
    # 헌치 - 셀 단위로 4면에 생성
    #   x면 헌치(y방향 압출)는 각 단부에서 a 만큼 잘라 y면 헌치와의 중복을 없앤다.
    #   이때 코너마다 정확히 a²b/6 이 누락되며, hn_corner 로 되살릴 수 있다.
    # ------------------------------------------------------------------
    def _cell_haunches(self, cell):
        sp = self.sp
        a, b = sp.hn_a, sp.hn_b
        x0, x1, y0, y1 = cell['x0'], cell['x1'], cell['y0'], cell['y1']
        out = []
        lx, ly = x1 - x0, y1 - y0
        ylen = ly - 2 * a
        if ylen > EPS:
            out.append((f"헌치 {cell['name']} 전면측",
                        TriPrism('y', x0, sp.tb, +a, +b, y0 + a, ylen)))
            out.append((f"헌치 {cell['name']} 후면측",
                        TriPrism('y', x1, sp.tb, -a, +b, y0 + a, ylen)))
        if lx > EPS:
            out.append((f"헌치 {cell['name']} 좌측",
                        TriPrism('x', y0, sp.tb, +a, +b, x0, lx)))
            out.append((f"헌치 {cell['name']} 우측",
                        TriPrism('x', y1, sp.tb, -a, +b, x0, lx)))
        return out

    def _cell_corner_lump(self, cell):
        """헌치 코너 보정 4개를 하나로 묶은 보정항.

        코너 국부좌표(u = 벽면에서 셀 안쪽 x, v = 벽면에서 셀 안쪽 y,
        w = 저판 상면에서 위)에서 누락 영역은
            { u/a + w/b ≤ 1 } ∩ { v/a + w/b > 1 }
        이며 적분하면 체적 a²b/6, 도심 (a/4, 3a/4, b/2) 이다.
        """
        sp = self.sp
        a, b = sp.hn_a, sp.hn_b
        x0, x1, y0, y1 = cell['x0'], cell['x1'], cell['y0'], cell['y1']
        v1 = a * a * b / 6.0
        pts = [(x0 + a / 4.0, y0 + 3 * a / 4.0),
               (x1 - a / 4.0, y0 + 3 * a / 4.0),
               (x0 + a / 4.0, y1 - 3 * a / 4.0),
               (x1 - a / 4.0, y1 - 3 * a / 4.0)]
        vt = 4 * v1
        cx = sum(p[0] for p in pts) / 4.0
        cy = sum(p[1] for p in pts) / 4.0
        cz = sp.tb + b / 2.0
        return Lump(vt, cx, cy, cz,
                    f"4 × {a:.3f}² × {b:.3f} / 6")

    # ------------------------------------------------------------------
    # 수직(우각부) 헌치 - 셀의 4개 연직 모서리
    #
    #   격벽-격벽, 격벽-외벽, 외벽-외벽이 만나는 연직 모서리의 45° 헌치다.
    #   수평 헌치 상단(z = tb + hn_b)부터 세우므로 두 헌치가 절대 겹치지 않는다.
    #   높이 w 에서 수평 헌치가 차지하는 폭은 s = a(1 − w/b) 이고 w ≥ b 이면
    #   s = 0 이기 때문이다. 실무 계산서의 "수직헌치 = 삼각단면 × (벽체높이 −
    #   수평헌치높이)" 관행과도 일치한다.
    #
    #   코너를 이루는 두 벽 중 낮은 쪽까지만 세운다 (감격벽 대응).
    # ------------------------------------------------------------------
    def _vh_z0(self):
        """수직 헌치 시작 높이 - 수평 헌치가 있으면 그 상단부터."""
        sp = self.sp
        return sp.tb + (sp.hn_b if sp.hn_on else 0.0)

    def _cell_corners(self, cell):
        """셀의 4개 연직 코너 [(x, y, du부호, dv부호, 코너높이), ...].

        코너높이는 그 코너를 이루는 두 벽 높이 중 낮은 쪽이다 (저판 상면 기준).
        """
        x0, x1, y0, y1 = cell['x0'], cell['x1'], cell['y0'], cell['y1']
        hx0, hx1 = cell['hx0'], cell['hx1']
        hy0, hy1 = cell['hy0'], cell['hy1']
        return [
            ('좌전', x0, y0, +1, +1, min(hx0, hy0)),
            ('좌후', x1, y0, -1, +1, min(hx1, hy0)),
            ('우전', x0, y1, +1, -1, min(hx0, hy1)),
            ('우후', x1, y1, -1, -1, min(hx1, hy1)),
        ]

    def cell_vh_top(self, cell):
        """그 셀에서 수직 헌치가 올라갈 수 있는 최고 높이.

        덮개가 있으면 실제로도 헌치는 덮개 밑면까지만 올라간다.
        """
        cv = cell.get('cover', 0.0)
        return self.sp.H - cv if cv > EPS else None

    def _cell_vhaunches(self, cell, z_top=None):
        """셀의 수직 헌치 [(이름, TriPrism), ...].

        z_top 을 주면 그 높이까지만 잘라 만든다 (속채움 공제에 쓴다).
        """
        sp = self.sp
        c = sp.vh_c
        z0 = self._vh_z0()
        cap = self.cell_vh_top(cell)
        out = []
        for tag, x, y, sx, sy, hcor in self._cell_corners(cell):
            z1 = sp.tb + hcor
            if cap is not None:
                z1 = min(z1, cap)              # 덮개 밑면에서 멈춘다
            if z_top is not None:
                z1 = min(z1, z_top)
            h = z1 - z0
            if h <= EPS:
                continue
            out.append((f"헌치 {cell['name']} 연직 {tag}",
                        TriPrism('z', x, y, sx * c, sy * c, z0, h)))
        return out

    @staticmethod
    def _vh_corner_integrals(m):
        """수직 헌치 코너 보정에 쓰는 무차원 적분 I0~I3.

        높이 w 에서 빠지는 영역은 (s, s) 를 직각점으로 하는 닮은 삼각형이며
        (s = a(1 − w/b), 다리 = c·k, k = max(0, 1 − 2s/c)), τ = 1 − w/b 로
        치환하면 k = max(0, 1 − mτ), m = 2a/c 가 된다.
        """
        U = min(1.0, 1.0 / m)
        I0 = U - m * U ** 2 + m ** 2 * U ** 3 / 3.0
        I1 = (U - (2 * m + 1) * U ** 2 / 2.0
              + (m ** 2 + 2 * m) * U ** 3 / 3.0 - m ** 2 * U ** 4 / 4.0)
        I2 = U ** 2 / 2.0 - 2 * m * U ** 3 / 3.0 + m ** 2 * U ** 4 / 4.0
        I3 = (1.0 - (1.0 - m * U) ** 4) / (4.0 * m)
        return I0, I1, I2, I3

    def _cell_vcorner_lump(self, cell):
        """수직 헌치를 수평 헌치 상단부터 세우면서 빠지는 4 코너분 보정.

        w < b 구간에서 수직 헌치 삼각형 {u + v ≤ c} 가 수평 헌치가 만드는
        L자 {u ≤ s} ∪ {v ≤ s} 밖으로 삐져나오는 부분이다. 기존 a²b/6 보정과는
        영역이 겹치지 않아(u ≤ s 인지 아닌지로 갈린다) 그냥 더하면 된다.

            V   = (c²/2) · b · I0
            z̄  = tb + b · I1/I0
            ū = v̄ = (a·I2 + (c/3)·I3) / I0      (코너에서 셀 안쪽으로)

        코너 높이가 수평 헌치 상단에 못 미치는 감격벽 코너는 애초에 수직 헌치가
        서지 않으므로 보정에서도 뺀다.
        """
        sp = self.sp
        a, b, c = sp.hn_a, sp.hn_b, sp.vh_c
        if a <= EPS or b <= EPS or c <= EPS:
            return None
        m = 2.0 * a / c
        I0, I1, I2, I3 = self._vh_corner_integrals(m)
        if I0 <= 0:
            return None
        v1 = c * c / 2.0 * b * I0
        d_in = (a * I2 + c / 3.0 * I3) / I0      # 코너에서 셀 안쪽 거리
        cz = sp.tb + b * I1 / I0
        z0 = self._vh_z0()
        cap = self.cell_vh_top(cell)
        pts = [(x + sx * d_in, y + sy * d_in)
               for _tag, x, y, sx, sy, hcor in self._cell_corners(cell)
               if (min(sp.tb + hcor, cap if cap is not None else 1e18)
                   - z0) > EPS]
        if not pts:
            return None
        return Lump(v1 * len(pts),
                    sum(p[0] for p in pts) / len(pts),
                    sum(p[1] for p in pts) / len(pts), cz,
                    f"{len(pts)} × {c:.3f}²/2 × {b:.3f} × {I0:.5f}")

    # ------------------------------------------------------------------
    # 속채움 검증 - 셀별 채움 상단
    # ------------------------------------------------------------------
    def _hole_face_cells(self):
        """유공 대상 벽에 접한 셀의 (ix, iy) 판정 집합.

        x축 벽은 셀의 ix 로, y축 벽은 iy 로 인접 여부를 가린다.
        """
        ixs, iys = set(), set()
        nx, ny = len(self.lps) + 1, len(self.tps) + 1
        for g in getattr(self, 'wall_geoms', []):
            name, ax = g['name'], g['axis']
            if name == WALL_FRONT:
                ixs.add(0)
            elif name == WALL_REAR:
                ixs.add(nx - 1)
            elif name == WALL_LEFT:
                iys.add(0)
            elif name == WALL_RIGHT:
                iys.add(ny - 1)
            else:
                try:
                    k = int(str(name).split('#')[-1]) - 1
                except Exception:
                    continue
                (ixs if ax == 'x' else iys).update({k, k + 1})
        return ixs, iys

    def _validate_fill(self):
        """셀별 채움 상단 z 를 검증하고 경고를 모은다.

        채움 상단은 저판 저면 기준 절대 표고이다. 0 이하이면 채우지 않는다.
        """
        sp = self.sp
        tb, H = sp.tb, sp.H
        low_haunch = []
        blocked = []
        z_hole_min = min((h['v'] - h['hv'] for h in self.holes), default=None)
        face_ix, face_iy = (self._hole_face_cells() if self.holes
                            else (set(), set()))

        for c in self.cells:
            # --- 유수실 덮개 ---
            cv = c['cover']
            if cv > EPS:
                if cv > self.hw + EPS:
                    raise InputError(
                        f"[셀] {c['name']} 의 덮개 두께 {cv:.3f}m 가 벽체 높이 "
                        f"{self.hw:.3f}m 를 넘습니다.")
            ft = c['fill_top']
            if ft <= EPS:
                continue
            if ft < tb - EPS or ft > H + EPS:
                raise InputError(
                    f"[셀] {c['name']} ({c['ctype']}) 의 채움 상단 {ft:.3f}m 가 "
                    f"저판 상면 {tb:.3f}m ~ 정단 {H:.3f}m 범위를 벗어납니다.")
            if cv > EPS and ft > H - cv + EPS:
                raise InputError(
                    f"[셀] {c['name']} 의 채움 상단 {ft:.3f}m 가 덮개 밑면 "
                    f"{H - cv:.3f}m (정단 {H:.3f} − 덮개 {cv:.3f}) 보다 높습니다. "
                    f"채움을 낮추거나 덮개를 얇게 하십시오.")
            if sp.hn_on and ft < tb + sp.hn_b - EPS:
                low_haunch.append(f"{c['name']} ({ft:.3f}m)")
            if z_hole_min is not None and ft > z_hole_min + EPS \
                    and (c['ix'] in face_ix or c['iy'] in face_iy):
                blocked.append(f"{c['name']} ({ft:.3f}m)")

        if low_haunch:
            self.warns.append(
                f"다음 셀의 채움 상단이 헌치 상단 {tb + sp.hn_b:.3f}m 보다 낮습니다 "
                f"- {', '.join(low_haunch)}. 헌치 전량을 속채움에서 공제했으므로 "
                f"해당 셀의 속채움 체적이 과소 산정됩니다. 헌치를 끄거나 채움 상단을 "
                f"올려 확인하십시오.")
        if blocked:
            self.warns.append(
                f"다음 셀의 채움 상단이 최하단 유공의 하단 {z_hole_min:.3f}m 보다 "
                f"높습니다 - {', '.join(blocked)}. 체적 산정에는 영향이 없으나 "
                f"속채움이 유공을 막으므로 채움 높이를 재검토하십시오.")

    # ------------------------------------------------------------------
    # 부재 분해
    # ------------------------------------------------------------------
    def build(self):
        sp = self.sp
        self.parts = []
        self.warns = []
        self._validate_basic()

        self.lps = self._build_partitions(sp.lp_mode, sp.lp_n, sp.lp_t, sp.lp_h,
                                          sp.lp_list, self.x_in0, self.x_in1, '종격벽')
        self.tps = self._build_partitions(sp.tp_mode, sp.tp_n, sp.tp_t, sp.tp_h,
                                          sp.tp_list, self.y_in0, self.y_in1, '횡격벽')
        self._make_cells()
        self._build_holes()

        B, L, H, tb = sp.B, sp.L, sp.H, sp.tb
        tf, tr = sp.tf, sp.tr
        ts_l, ts_r = sp.ts_l, sp.ts_r
        hw = self.hw
        P = self.parts

        def add(name, group, sign, prim, note='', inner=True, roll=''):
            P.append(Part(name, group, sign, prim, note, inner, roll))

        # --- (1) 저판 ---
        add('저판', GRP_CONC, +1, Box(0, 0, 0, B, L, tb), 'B × L × 저판두께')

        # --- (2) 외벽 : 측벽은 전·후면벽 사이만 차지시켜 중복을 없앤다 ---
        add('전면벽', GRP_CONC, +1, Box(0, 0, tb, tf, L, hw), 'tf × L × 벽체높이')
        add('후면벽', GRP_CONC, +1, Box(B - tr, 0, tb, tr, L, hw), 'tr × L × 벽체높이')
        add('좌측벽', GRP_CONC, +1, Box(tf, 0, tb, B - tf - tr, ts_l, hw),
            '순폭 × ts_l × 벽체높이')
        add('우측벽', GRP_CONC, +1, Box(tf, L - ts_r, tb, B - tf - tr, ts_r, hw),
            '순폭 × ts_r × 벽체높이')

        # --- (3) 격벽 + 교차부 공제 ---
        for i, d in enumerate(self.lps):
            add(f'종격벽 #{i + 1}', GRP_CONC, +1,
                Box(d['lo'], self.y_in0, tb, d['t'], self.y_in1 - self.y_in0, d['h']),
                f"중심 x={d['pos']:.3f}")
        for j, d in enumerate(self.tps):
            add(f'횡격벽 #{j + 1}', GRP_CONC, +1,
                Box(self.x_in0, d['lo'], tb, self.x_in1 - self.x_in0, d['t'], d['h']),
                f"중심 y={d['pos']:.3f}")
        n_cross = 0
        for i, a in enumerate(self.lps):
            for j, b in enumerate(self.tps):
                hh = min(a['h'], b['h'])
                add(f'격벽 교차 L#{i + 1}×T#{j + 1}', GRP_CONC, -1,
                    Box(a['lo'], b['lo'], tb, a['t'], b['t'], hh),
                    '종·횡격벽 중복 공제')
                n_cross += 1

        # --- (4) 유공 공제 ---
        v_hole = 0.0
        for h in self.holes:
            ul = 'y' if h['axis'] == 'x' else 'x'
            add(f"유공 #{h['i']}", GRP_CONC, -1, h['prim'],
                f"{ul}={h['u']:.3f}, z={h['v']:.3f}",
                roll=f"유공 ({h['wall']})")
            v_hole += h['prim'].volume()

        # --- (5) 헌치 (+ 코너 보정) 및 셀 ---
        v_hn = 0.0
        if sp.hn_on:
            if sp.hn_a <= 0 or sp.hn_b <= 0:
                raise InputError(f"[헌치] 다리 길이가 0 이하입니다. "
                                 f"(a={sp.hn_a:.3f}, b={sp.hn_b:.3f})")
            if sp.hn_b > hw + EPS:
                raise InputError(f"[헌치] 연직 다리 {sp.hn_b:.3f}m 가 "
                                 f"벽체 높이 {hw:.3f}m 를 넘습니다.")
            for c in self.cells:
                if (c['x1'] - c['x0']) < 2 * sp.hn_a - EPS or \
                        (c['y1'] - c['y0']) < 2 * sp.hn_a - EPS:
                    raise InputError(
                        f"[헌치] 셀 {c['name']} 의 순치수 "
                        f"({c['x1'] - c['x0']:.3f} × {c['y1'] - c['y0']:.3f}m) 가 "
                        f"헌치 수평 다리 2×{sp.hn_a:.3f}m 보다 작습니다.")
        if sp.vh_on:
            if sp.vh_c <= 0:
                raise InputError(f"[수직 헌치] 다리 {sp.vh_c:.3f}m 가 0 이하입니다.")
            for c in self.cells:
                if (c['x1'] - c['x0']) < 2 * sp.vh_c - EPS or \
                        (c['y1'] - c['y0']) < 2 * sp.vh_c - EPS:
                    raise InputError(
                        f"[수직 헌치] 셀 {c['name']} 의 순치수 "
                        f"({c['x1'] - c['x0']:.3f} × {c['y1'] - c['y0']:.3f}m) 가 "
                        f"수직 헌치 2×{sp.vh_c:.3f}m 보다 작습니다. "
                        f"양쪽 코너 헌치가 서로 겹칩니다.")
            if self._vh_z0() >= tb + hw - EPS:
                raise InputError(
                    f"[수직 헌치] 시작 높이 {self._vh_z0():.3f}m 가 벽체 상단 "
                    f"{tb + hw:.3f}m 이상이라 세울 자리가 없습니다. "
                    f"수평 헌치 연직 다리 b 를 줄이십시오.")

        # 속채움 상단은 셀별로 다를 수 있으므로 부재를 만들기 전에 모두 검증한다
        self._validate_fill()

        v_cover = 0.0
        for c in self.cells:
            rh = f"헌치 {c['name']}"
            rv = f"수직헌치 {c['name']}"
            rf = f"속채움 {c['name']}"
            # 유수실 덮개 - 벽체 상단에 얹히는 슬래브 (셀 순개구부만 차지한다)
            if c['cover'] > EPS:
                pr = Box(c['x0'], c['y0'], H - c['cover'],
                         c['x1'] - c['x0'], c['y1'] - c['y0'], c['cover'])
                add(f"덮개 {c['name']}", GRP_CONC, +1, pr,
                    f"두께 {c['cover']:.3f}, 밑면 z={H - c['cover']:.3f}",
                    roll='유수실 덮개')
                v_cover += pr.volume()
            hns = self._cell_haunches(c) if sp.hn_on else []
            for nm, pr in hns:
                add(nm, GRP_CONC, +1, pr, '저판-벽체 접합부', roll=rh)
                v_hn += pr.volume()
            if sp.hn_on and sp.hn_corner:
                lump = self._cell_corner_lump(c)
                add(f"{rh} 코너 보정", GRP_CONC, +1, lump, '4 코너 합산', roll=rh)
                v_hn += lump.volume()

            # 수직(우각부) 헌치 - 수평 헌치 상단부터라 서로 겹치지 않는다
            vhs = self._cell_vhaunches(c) if sp.vh_on else []
            for nm, pr in vhs:
                add(nm, GRP_CONC, +1, pr, '연직 모서리 우각부', roll=rv)
                v_hn += pr.volume()
            vlump = (self._cell_vcorner_lump(c)
                     if (sp.vh_on and sp.hn_on and sp.hn_corner) else None)
            if vlump is not None:
                add(f"{rv} 코너 보정", GRP_CONC, +1, vlump,
                    '수평 헌치 상단 아래 누락분', roll=rv)
                v_hn += vlump.volume()

            # 속채움 - 셀별 채움 상단까지. 셀 공극에서 헌치 점유분을 뺀다.
            ht = c['fill_top'] - tb
            if ht > EPS:
                add(rf, GRP_FILL, +1,
                    Box(c['x0'], c['y0'], tb, c['x1'] - c['x0'], c['y1'] - c['y0'], ht),
                    f"셀 {c['x1'] - c['x0']:.3f} × {c['y1'] - c['y0']:.3f}"
                    f" × {ht:.3f}", roll=rf)
                for nm, pr in hns:
                    add(f"{nm} 공제", GRP_FILL, -1, pr, '속채움에서 헌치 제외', roll=rf)
                if sp.hn_on and sp.hn_corner:
                    add(f"{rh} 코너 보정 공제", GRP_FILL, -1,
                        self._cell_corner_lump(c), '속채움에서 헌치 제외', roll=rf)
                # 수직 헌치는 채움 상단에서 잘라 정확히 공제한다
                for nm, pr in (self._cell_vhaunches(c, z_top=c['fill_top'])
                               if sp.vh_on else []):
                    add(f"{nm} 공제", GRP_FILL, -1, pr,
                        '속채움에서 수직 헌치 제외', roll=rf)
                if vlump is not None:
                    add(f"{rv} 코너 보정 공제", GRP_FILL, -1, vlump,
                        '속채움에서 수직 헌치 제외', roll=rf)

        # --- (6) 풋팅 (링 방식 - 모서리 중복 원천 차단) ---
        wf, wr, ws, tft = sp.ft_front, sp.ft_rear, sp.ft_side, sp.ft_t
        if max(wf, wr, ws) > EPS:
            if tft <= 0:
                raise InputError("[풋팅] 폭을 입력했으나 풋팅 두께가 0 입니다.")
            if tft > tb + EPS:
                raise InputError(f"[풋팅] 두께 {tft:.3f}m 가 저판 두께 "
                                 f"{tb:.3f}m 를 넘습니다.")
            add('풋팅(외곽)', GRP_CONC, +1,
                Box(-wf, -ws, 0, B + wf + wr, L + 2 * ws, tft),
                '전·후·측 풋팅 일체', inner=False)
            add('풋팅(본체 공제)', GRP_CONC, -1, Box(0, 0, 0, B, L, tft),
                '저판과 중복되는 부분', inner=False)

        # --- (7) 전단키 ---
        ks = sorted(sp.keys, key=lambda k: k.x)
        for i, k in enumerate(ks):
            if k.w <= 0 or k.d <= 0:
                raise InputError(f"[전단키] #{i + 1} 의 폭·깊이가 0 이하입니다.")
            if k.x < -EPS or k.x + k.w > B + EPS:
                raise InputError(f"[전단키] #{i + 1} 이 저판 폭 "
                                 f"[0, {B:.3f}] 을 벗어납니다.")
            if i > 0 and k.x < ks[i - 1].x + ks[i - 1].w - EPS:
                raise InputError(f"[전단키] #{i} 과 #{i + 1} 이 서로 겹칩니다.")
            add(f'전단키 #{i + 1}', GRP_CONC, +1, Box(k.x, 0, -k.d, k.w, L, k.d),
                f"저판 저면 아래 {k.d:.3f}m", inner=False)

        # --- (8) 상치콘크리트 ---
        if sp.cp_on:
            z0 = sp.cp_z0 if sp.cp_z0 > EPS else H
            if sp.cp_w <= 0 or sp.cp_h <= 0:
                raise InputError("[상치] 폭·높이가 0 이하입니다.")
            add('상치 본체', GRP_COPING, +1, Box(sp.cp_x0, 0, z0, sp.cp_w, L, sp.cp_h),
                f"저면 z={z0:.3f}", inner=False)
            top_z = z0 + sp.cp_h
            front_x = sp.cp_x0
            if sp.cp_par_on:
                if sp.cp_par_w <= 0 or sp.cp_par_h <= 0:
                    raise InputError("[상치] 파라펫 폭·높이가 0 이하입니다.")
                if sp.cp_par_w > sp.cp_w + EPS:
                    raise InputError(f"[상치] 파라펫 폭 {sp.cp_par_w:.3f}m 가 "
                                     f"상치 폭 {sp.cp_w:.3f}m 보다 큽니다.")
                px = sp.cp_x0 if sp.cp_par_side == SIDE_FRONT \
                    else sp.cp_x0 + sp.cp_w - sp.cp_par_w
                add('상치 파라펫', GRP_COPING, +1,
                    Box(px, 0, top_z, sp.cp_par_w, L, sp.cp_par_h),
                    f"{sp.cp_par_side} 흉벽", inner=False)
                if sp.cp_par_side == SIDE_FRONT:
                    front_x, top_z = px, top_z + sp.cp_par_h
            if sp.cp_sl_a > EPS and sp.cp_sl_b > EPS:
                lim_w = sp.cp_par_w if (sp.cp_par_on and sp.cp_par_side == SIDE_FRONT) \
                    else sp.cp_w
                lim_h = sp.cp_par_h if (sp.cp_par_on and sp.cp_par_side == SIDE_FRONT) \
                    else sp.cp_h
                if sp.cp_sl_a > lim_w + EPS or sp.cp_sl_b > lim_h + EPS:
                    raise InputError(
                        f"[상치] 전면 경사 {sp.cp_sl_a:.3f}×{sp.cp_sl_b:.3f}m 가 "
                        f"대상 부재 치수 {lim_w:.3f}×{lim_h:.3f}m 를 넘습니다.")
                add('상치 전면경사 공제', GRP_COPING, -1,
                    TriPrism('y', front_x, top_z, +sp.cp_sl_a, -sp.cp_sl_b, 0, L),
                    '전면 상단 모서리', inner=False)

        # --- 검산용 집계 ---
        v_cells_box = sum((c['x1'] - c['x0']) * (c['y1'] - c['y0']) * hw
                          for c in self.cells)
        v_extra = 0.0
        for d in self.lps:
            v_extra += d['t'] * (self.y_in1 - self.y_in0) * (hw - d['h'])
        for d in self.tps:
            v_extra += d['t'] * (self.x_in1 - self.x_in0) * (hw - d['h'])
        for a in self.lps:
            for b in self.tps:
                v_extra -= a['t'] * b['t'] * (hw - min(a['h'], b['h']))

        self.q = dict(v_hn=v_hn, v_hole=v_hole, v_cells_box=v_cells_box,
                      v_extra=v_extra, n_cross=n_cross, v_cover=v_cover)
        self._run_check()
        return self

    # ------------------------------------------------------------------
    # 자동 검산
    # ------------------------------------------------------------------
    def _run_check(self):
        sp = self.sp
        v_outer = sp.B * sp.L * sp.H
        v_conc_in = sum(p.v for p in self.parts
                        if p.group == GRP_CONC and p.inner)
        # 경로 ① 부재 분해에서 유도한 공극
        void_parts = v_outer - v_conc_in
        # 경로 ② 셀 격자에서 유도한 공극 (덮개는 셀 공극을 막는다)
        void_cells = (self.q['v_cells_box'] + self.q['v_extra']
                      - self.q['v_hn'] + self.q['v_hole']
                      - self.q.get('v_cover', 0.0))
        d = void_parts - void_cells

        # 좌우 대칭 확인
        V, cx, cy, cz = mass_props([p for p in self.parts if p.group == GRP_CONC])
        sym_err = abs(cy - sp.L / 2.0)

        self.check = dict(
            v_outer=v_outer, v_conc_in=v_conc_in,
            void_parts=void_parts, void_cells=void_cells,
            delta=d, rel=(abs(d) / v_outer if v_outer > EPS else 0.0),
            ok=(abs(d) <= max(1e-6, v_outer * 1e-9)),
            sym_err=sym_err, sym_ok=(sym_err <= 1e-6),
        )

    # ------------------------------------------------------------------
    # 결과 조회
    # ------------------------------------------------------------------
    def group_parts(self, group):
        return [p for p in self.parts if p.group == group]

    def group_props(self, group):
        return mass_props(self.group_parts(group))

    @staticmethod
    def _hole_face(h):
        return (math.pi * h['hu'] ** 2 if h['circle']
                else 4 * h['hu'] * h['hv'])

    def opening_ratio(self):
        """개구율 = Σ 유공 정면적 / 대상 벽 정면적 (전 대상 벽 합계)."""
        if not self.holes:
            return 0.0, 0.0, getattr(self, 'face_area', 0.0)
        a = sum(self._hole_face(h) for h in self.holes)
        face = getattr(self, 'face_area', 0.0)
        return (a / face if face > EPS else 0.0), a, face

    def opening_by_wall(self):
        """벽별 개구율 [(벽이름, 개수, 유공면적, 정면적, 개구율), ...]."""
        out = []
        for g in getattr(self, 'wall_geoms', []):
            hs = [h for h in self.holes if h['wall'] == g['name']]
            a = sum(self._hole_face(h) for h in hs)
            out.append((g['name'], len(hs), a, g['face'],
                        a / g['face'] if g['face'] > EPS else 0.0))
        return out

    def combined(self):
        """단위중량 가중 합성 무게중심. use_gamma 가 꺼져 있으면 None."""
        sp = self.sp
        if not sp.use_gamma:
            return None
        rows = []
        for g, gam in ((GRP_CONC, sp.g_conc), (GRP_FILL, sp.g_fill),
                       (GRP_COPING, sp.g_cop)):
            V, cx, cy, cz = self.group_props(g)
            if abs(V) < 1e-12:
                continue
            rows.append((g, V, gam, V * gam, cx, cy, cz))
        W = sum(r[3] for r in rows)
        if abs(W) < 1e-12:
            return None
        gx = sum(r[3] * r[4] for r in rows) / W
        gy = sum(r[3] * r[5] for r in rows) / W
        gz = sum(r[3] * r[6] for r in rows) / W
        return dict(rows=rows, W=W, gx=gx, gy=gy, gz=gz)

    # ------------------------------------------------------------------
    # 몬테카를로 형상 검증
    #
    #   해석적 분해와 완전히 다른 경로로 체적·도심을 구한다. 아울러 모든
    #   표본점의 부호 합이 0 또는 1 인지 확인해 부재 중복·과공제를 직접
    #   찾아낸다 (합이 2 이면 이중 계상, −1 이면 과공제).
    #   ※ Lump(코너 보정)는 점 포함 판정이 불가능하므로 제외한다.
    # ------------------------------------------------------------------
    def monte_carlo(self, groups=(GRP_CONC,), n=2_000_000, seed=20260810,
                    batch=200_000):
        parts = [p for p in self.parts
                 if p.group in groups and p.prim.samplable]
        skipped = [p for p in self.parts
                   if p.group in groups and not p.prim.samplable]
        if not parts:
            return None
        bb = [p.prim.bbox() for p in parts]
        x0 = min(b[0] for b in bb); x1 = max(b[1] for b in bb)
        y0 = min(b[2] for b in bb); y1 = max(b[3] for b in bb)
        z0 = min(b[4] for b in bb); z1 = max(b[5] for b in bb)
        vbb = (x1 - x0) * (y1 - y0) * (z1 - z0)
        if vbb <= EPS:
            return None

        rng = np.random.default_rng(seed)
        tot = 0.0
        tot2 = 0.0
        sx = sy = sz = 0.0
        bad_over = 0     # 부호 합 ≥ 2  → 이중 계상
        bad_under = 0    # 부호 합 < 0  → 과공제
        done = 0
        while done < n:
            m = min(batch, n - done)
            P = np.empty((m, 3))
            P[:, 0] = rng.uniform(x0, x1, m)
            P[:, 1] = rng.uniform(y0, y1, m)
            P[:, 2] = rng.uniform(z0, z1, m)
            acc = np.zeros(m, dtype=np.int32)
            for p in parts:
                acc += p.sign * p.prim.contains(P).astype(np.int32)
            bad_over += int(np.count_nonzero(acc >= 2))
            bad_under += int(np.count_nonzero(acc < 0))
            w = acc.astype(np.float64)
            tot += w.sum()
            tot2 += (w * w).sum()
            sx += (w * P[:, 0]).sum()
            sy += (w * P[:, 1]).sum()
            sz += (w * P[:, 2]).sum()
            done += m

        if abs(tot) < EPS:
            return None
        mean = tot / n
        var = max(0.0, tot2 / n - mean * mean)
        # 표본 평균의 1σ → 체적의 1σ. 오차가 이 범위 안이면 통계 요동일 뿐이다.
        sigma = vbb * math.sqrt(var / n)
        return dict(n=n, V=mean * vbb, sigma=sigma,
                    cx=sx / tot, cy=sy / tot, cz=sz / tot,
                    bad_over=bad_over, bad_under=bad_under,
                    skipped=[p.name for p in skipped], vbb=vbb)


# ==========================================
# 3. 계산서 생성
#    폴더 표준 서식 : [수식] → [대입] → [계산] → → 결과
# ==========================================

W = 118
LINE1 = "=" * W
LINE2 = "-" * W
LINE3 = "·" * (W // _AMB_W)      # 애매폭 문자라 글꼴에 따라 개수가 달라진다

# 부재 표 열 폭 - 합계 + 구분 공백 = W 가 되도록 맞춘다
#   산출식 30 : TriPrism 의 '1/2 × 0.500 × 0.500 × 3.550' 이 30 칸을 쓴다
CW = (24, 4, 30, 11, 7, 7, 8, 9, 10)
# §3 셀 표 열 폭
CWC = (6, 7, 15, 15, 15, 11, 9, 8, 14, 9)


class ReportBuilder:
    """CaissonEngine 의 결과를 계산서 문자열로 조립한다."""

    def __init__(self, eng: CaissonEngine, detail=False):
        self.e = eng
        self.sp = eng.sp
        self.detail = detail
        self.buf = []

    def _w(self, s=''):
        self.buf.append(s)

    def build(self):
        self.buf = []
        self._header()
        self._sec0_axis()
        self._sec1_input()
        self._sec2_parts()
        self._sec3_cells()
        self._sec4_groups()
        self._sec5_summary()
        self._sec6_check()
        self._sec7_notes()
        return "\n".join(self.buf)

    # ------------------------------------------------------------------
    def layout_errors(self):
        """표 정렬 자체 점검 - 폭이 넘치는 줄과 머리글을 돌려준다.

        표는 _dw() 의 표시폭 가정 위에 서 있어서 문구를 늘리거나 애매폭 기호를
        넣으면 조용히 어긋난다. 개발 중 그것을 바로 드러내기 위한 점검이며
        계산 결과에는 영향을 주지 않는다.
        """
        bad = []
        for i, ln in enumerate(self.build().splitlines(), start=1):
            if _dw(ln) > W:
                bad.append(f"{i}행 폭 {_dw(ln)} > {W} : {ln[:40]}")
        heads = [
            (CW, ('부재명', '부호', '산출식', '체적 (m³)', 'Xg', 'Yg', 'Zg',
                  'V·Xg', 'V·Zg')),
            (CWC, ('셀', '구분', 'x 범위', 'y 범위', '순치수 (x × y)',
                   '채움상단 z', '채움높이', '덮개 t', '속채움 V (m³)', 'Zg')),
        ]
        for widths, texts in heads:
            for t, wd in zip(texts, widths):
                if _dw(t) > wd:
                    bad.append(f"머리글 '{t}' 폭 {_dw(t)} > 열폭 {wd}")
        return bad

    # ------------------------------------------------------------------
    def _header(self):
        e, sp = self.e, self.sp
        self._w(LINE1)
        self._w(_pad("케이슨 체적 · 무게중심 산정 계산서", W, 'center'))
        self._w(_pad("[ 부호 있는 기본도형 분해 - 해석적 정해 ]", W, 'center'))
        self._w(LINE1)
        self._w(f" 프로젝트   : {sp.project}")
        self._w(f" 케이슨명   : {sp.name}")
        walls = [g['name'] for g in getattr(e, 'wall_geoms', [])]
        self._w(f" 형식       : {sp.ctype}케이슨"
                f"{'  (유공 ' + ' · '.join(walls) + ')' if walls else ''}")
        self._w(f" 외곽 제원  : 폭 B = {sp.B:.3f} m × 길이 L = {sp.L:.3f} m"
                f" × 높이 H = {sp.H:.3f} m")
        self._w(f" 셀 구성    : 종격벽 {len(e.lps)} 매 × 횡격벽 {len(e.tps)} 매"
                f"  →  셀 {len(e.cells)} 개"
                f"  (속채움 {sum(1 for c in e.cells if c['ctype'] == CELL_FILL)} /"
                f" 유수실 {sum(1 for c in e.cells if c['ctype'] == CELL_CHAMBER)} /"
                f" 공셀 {sum(1 for c in e.cells if c['ctype'] == CELL_EMPTY)})")
        self._w(f" 속채움     : 채움 적용 {sum(1 for c in e.cells if c['fill_top'] > sp.tb + EPS)} 셀"
                f"  (셀별 채움 상단 z 지정 - 유수실도 채울 수 있다. §3 참조)")
        n_cov = sum(1 for c in e.cells if c['cover'] > EPS)
        if n_cov:
            v_cov = sum(p.v for p in e.parts if p.roll == '유수실 덮개')
            self._w(f" 덮개       : {n_cov} 셀에 콘크리트 덮개"
                    f"  (합 {v_cov:.3f} m³, 정단 아래로 얹힌다)")
        self._w(LINE1)
        self._w(" ※ 표는 고정폭 한글 글꼴에서 정렬됩니다."
                " (D2Coding · 나눔고딕코딩 · 돋움체 · 굴림체)")
        self._w(LINE1)
        if e.warns:
            self._w()
            self._w(" ※ 경고")
            for wmsg in e.warns:
                head, *rest = _wrap(wmsg, W - 5, '')
                self._w(f"   · {head}")
                for ln in rest:
                    self._w(f"     {ln}")
        self._w()

    # ------------------------------------------------------------------
    def _sec0_axis(self):
        sp = self.sp
        self._w("■ 0. 좌표계 정의 및 부호 규약")
        self._w(LINE2)
        self._w("  · x : 케이슨 폭 방향.   전면(해측) 외면 = 0,  후면 외면 = B")
        self._w("  · y : 케이슨 길이 방향. 좌측 단부 외면 = 0,  우측 단부 외면 = L")
        self._w("  · z : 연직 방향.        저판 저면 = 0,       케이슨 정단 = H")
        self._w("        전단키는 z < 0, 상치콘크리트는 z > H 구간을 차지한다.")
        self._w()
        self._w("  · 부재 부호  (+) 추가,  (−) 공제. 체적과 도심은 다음 누적식 하나로 구한다.")
        self._w("        V  = Σ sign·V_i           Cg = Σ (sign·V_i·c_i) / V")
        self._w("    메시·복셀 근사를 쓰지 않으므로 아래 값은 모두 해석적 정해이다.")
        self._w()
        self._w(f"  · 무게중심 표기 : Xg 는 전면 외면에서의 거리, Zg 는 저판 저면에서의 높이,")
        self._w(f"                    e = Xg − B/2 = Xg − {sp.B / 2:.3f} 는 폭 중앙 기준 편심이다.")
        self._w()

    # ------------------------------------------------------------------
    def _kv(self, label, value, unit='m', note=''):
        s = f"  - {_pad(label, 26)} : {value:>10}"
        if unit:
            s += f" {unit}"
        if note:
            s += f"   {note}"
        self._w(s)

    def _sec1_input(self):
        e, sp = self.e, self.sp
        self._w("■ 1. 입력 제원")
        self._w(LINE2)
        self._w("  ▶ 기본")
        self._kv("폭 B", f"{sp.B:.3f}")
        self._kv("길이 L", f"{sp.L:.3f}")
        self._kv("높이 H", f"{sp.H:.3f}", 'm', f"(벽체 높이 = H − 저판 = {e.hw:.3f} m)")
        self._kv("저판 두께", f"{sp.tb:.3f}")
        self._kv("전면벽 두께 tf", f"{sp.tf:.3f}")
        self._kv("후면벽 두께 tr", f"{sp.tr:.3f}")
        self._kv("좌측벽 두께 ts_l", f"{sp.ts_l:.3f}", 'm', '(y = 0 쪽)')
        self._kv("우측벽 두께 ts_r", f"{sp.ts_r:.3f}", 'm', '(y = L 쪽)')
        self._kv("내부 순치수", f"{e.x_in1 - e.x_in0:.3f}", 'm',
                 f"× {e.y_in1 - e.y_in0:.3f} m  (x × y)")

        self._w()
        self._w("  ▶ 격벽")
        if e.lps:
            self._w(f"      종격벽 {len(e.lps)} 매 ({sp.lp_mode})")
            for i, d in enumerate(e.lps):
                self._w(f"        #{i + 1}  중심 x = {d['pos']:7.3f} m,"
                        f"  두께 = {d['t']:.3f} m,  높이 = {d['h']:.3f} m"
                        f"  →  {d['lo']:.3f} ~ {d['hi']:.3f} m")
        else:
            self._w("      종격벽 없음")
        if e.tps:
            self._w(f"      횡격벽 {len(e.tps)} 매 ({sp.tp_mode})")
            for j, d in enumerate(e.tps):
                self._w(f"        #{j + 1}  중심 y = {d['pos']:7.3f} m,"
                        f"  두께 = {d['t']:.3f} m,  높이 = {d['h']:.3f} m"
                        f"  →  {d['lo']:.3f} ~ {d['hi']:.3f} m")
        else:
            self._w("      횡격벽 없음")
        self._w(f"      교차부 공제 : {len(e.lps)} × {len(e.tps)} = "
                f"{e.q['n_cross']} 개소")

        self._w()
        self._w("  ▶ 유공")
        if not sp.hole_on:
            self._w("      없음 (무공)")
        else:
            r, area, face = e.opening_ratio()
            self._w(f"      형식 : {sp.hole_shape}   배치 : {sp.hole_mode}"
                    f"{' / 열=' + sp.col_mode if sp.hole_mode == HOLE_GRID else ''}"
                    f"{f' / 간격 {sp.col_gap:.3f} m' if (sp.col_mode == COL_CELL and sp.col_gap > EPS) else ''}")
            if sp.hole_shape == HOLE_CIRCLE:
                self._w(f"      직경 D = {sp.hole_d:.3f} m   전체 {len(e.holes)} 개")
            else:
                self._w(f"      폭 × 높이 = {sp.hole_w:.3f} × {sp.hole_hh:.3f} m"
                        f"   전체 {len(e.holes)} 개")
            self._w()
            hd = (_pad('대상 벽체', 14) + ' ' + _pad('관통축', 7, 'center') + ' ' +
                  _pad('개수', 6, 'right') + ' ' + _pad('유공면적', 12, 'right') +
                  ' ' + _pad('벽 정면적', 12, 'right') + ' ' +
                  _pad('개구율', 10, 'right'))
            self._w('      ' + hd)
            self._w('      ' + LINE3[:len(hd) // _AMB_W])
            for nm, n, a_, f_, rr in e.opening_by_wall():
                self._w('      ' + _pad(nm, 14) + ' ' +
                        _pad(wall_axis(nm) + ' 방향', 7, 'center') + ' ' +
                        _pad(f"{n}", 6, 'right') + ' ' +
                        _pad(f"{a_:.3f}", 12, 'right') + ' ' +
                        _pad(f"{f_:.3f}", 12, 'right') + ' ' +
                        _pad(f"{rr * 100:.2f} %", 10, 'right'))
            self._w('      ' + LINE3[:len(hd) // _AMB_W])
            self._w('      ' + _pad('합계', 14) + ' ' + _pad('', 7) + ' ' +
                    _pad(f"{len(e.holes)}", 6, 'right') + ' ' +
                    _pad(f"{area:.3f}", 12, 'right') + ' ' +
                    _pad(f"{face:.3f}", 12, 'right') + ' ' +
                    _pad(f"{r * 100:.2f} %", 10, 'right'))
            self._w()
            zs = sorted({round(h['v'], 4) for h in e.holes})
            self._w(f"      행 z = {', '.join(f'{v:.3f}' for v in zs)}")
            for g in e.wall_geoms:
                us = sorted({round(h['u'], 4) for h in e.holes
                             if h['wall'] == g['name']})
                ul = 'y' if g['axis'] == 'x' else 'x'
                self._w(f"      {g['name']} 열 {ul} = "
                        f"{', '.join(f'{v:.3f}' for v in us)}")

        self._w()
        self._w("  ▶ 풋팅 · 헌치 · 전단키")
        if max(sp.ft_front, sp.ft_rear, sp.ft_side) > EPS:
            self._w(f"      풋팅 폭 전면/후면/측면 = {sp.ft_front:.3f} /"
                    f" {sp.ft_rear:.3f} / {sp.ft_side:.3f} m,"
                    f"  두께 = {sp.ft_t:.3f} m")
            self._w("      ※ 모서리 중복을 없애기 위해 링(외곽 − 본체) 방식으로 산정한다.")
        else:
            self._w("      풋팅 없음")
        n_c = 4 * len(e.cells)
        if sp.hn_on:
            self._w(f"      수평 헌치 (저판-벽체) : 수평 a = {sp.hn_a:.3f} m,"
                    f"  연직 b = {sp.hn_b:.3f} m   (모든 셀 4면)")
        else:
            self._w("      수평 헌치 없음")
        if sp.vh_on:
            z0 = e._vh_z0()
            v1 = sp.vh_c ** 2 / 2.0
            self._w(f"      수직 헌치 (우각부)     : 다리 c = {sp.vh_c:.3f} m (45°),"
                    f"  단면적 c²/2 = {v1:.5f} m²")
            self._w(f"        {n_c} 코너 (셀 {len(e.cells)} × 4). z = {z0:.3f} m"
                    f"{' (수평 헌치 상단)' if sp.hn_on else ' (저판 상면)'}"
                    f" 부터 코너를 이루는 두 벽 중 낮은 쪽까지.")
            vv = sum(p.v for p in e.parts if p.roll.startswith('수직헌치'))
            self._w(f"        수직 헌치 합계 = {vv:.3f} m³")
        else:
            self._w("      수직 헌치 없음")
        if sp.hn_on or sp.vh_on:
            v1 = sp.hn_a ** 2 * sp.hn_b / 6.0
            vl = sum(p.v for p in e.parts
                     if p.group == GRP_CONC and not p.prim.samplable)
            if sp.hn_corner and sp.hn_on:
                self._w(f"      헌치 코너 보정 적용 : 수평 {n_c} × a²b/6 ="
                        f" {n_c * v1:.3f} m³"
                        + (f",  수직 보정 포함 총 {vl:.3f} m³" if sp.vh_on else ''))
            elif sp.hn_on:
                self._w(f"      헌치 코너 보정 미적용 : 누락량 ≥ {n_c} × a²b/6"
                        f" = {n_c * v1:.3f} m³ (과소 산정)")
        if sp.keys:
            for i, k in enumerate(sorted(sp.keys, key=lambda q: q.x)):
                self._w(f"      전단키 #{i + 1}  x = {k.x:.3f} ~ {k.x + k.w:.3f} m,"
                        f"  깊이 = {k.d:.3f} m")
        else:
            self._w("      전단키 없음")

        self._w()
        self._w("  ▶ 셀 (속채움)")
        self._w(f"      기본 채움 상단 z = {sp.fill_top:.3f} m"
                f"  (저판 상면 {sp.tb:.3f} m 에서 {sp.fill_top - sp.tb:.3f} m)")
        self._w("      ※ 채움 높이는 셀별로 지정한다. 위 값은 셀별 지정이 없는")
        self._w("         속채움 셀에만 적용되는 기본값이다. 셀별 값은 §3 참조.")
        fills = [c for c in e.cells if c['fill_top'] > sp.tb + EPS]
        tops = sorted({round(c['fill_top'], 4) for c in fills})
        if not fills:
            self._w("      채움 적용 셀 없음")
        else:
            self._w(f"      채움 적용 {len(fills)} 셀 / 전체 {len(e.cells)} 셀,"
                    f"  채움 상단 z = {', '.join(f'{v:.3f}' for v in tops)} m")
            ch = [c['name'] for c in fills if c['ctype'] == CELL_CHAMBER]
            if ch:
                self._w(f"      유수실 속채움 : {', '.join(ch)}")

        self._w()
        self._w("  ▶ 상치콘크리트")
        if sp.cp_on:
            z0 = sp.cp_z0 if sp.cp_z0 > EPS else sp.H
            self._w(f"      본체 : x = {sp.cp_x0:.3f} ~ {sp.cp_x0 + sp.cp_w:.3f} m,"
                    f"  z = {z0:.3f} ~ {z0 + sp.cp_h:.3f} m,  길이 = {sp.L:.3f} m")
            if sp.cp_par_on:
                self._w(f"      파라펫({sp.cp_par_side}) : 폭 {sp.cp_par_w:.3f} m,"
                        f"  높이 {sp.cp_par_h:.3f} m")
            if sp.cp_sl_a > EPS and sp.cp_sl_b > EPS:
                self._w(f"      전면 상단 경사 : {sp.cp_sl_a:.3f} (수평)"
                        f" × {sp.cp_sl_b:.3f} (연직) 공제")
        else:
            self._w("      없음")

        self._w()
        self._w("  ▶ 재료 단위중량   [KDS 64 10 10 재료표]")
        if sp.use_gamma:
            self._kv("케이슨 콘크리트", f"{sp.g_conc:.2f}", 'kN/m³')
            self._kv("속채움", f"{sp.g_fill:.2f}", 'kN/m³')
            self._kv("상치콘크리트", f"{sp.g_cop:.2f}", 'kN/m³')
            self._w("      ※ 단위중량은 재료가 다른 부재군을 하나의 무게중심으로")
            self._w("         합성하기 위한 가중치로만 쓴다. 중량·부력 검토는 범위 외이다.")
        else:
            self._w("      미사용 - 부재군별 체적과 도심만 산정한다.")
        self._w()

    # ------------------------------------------------------------------
    # 부재 표
    # ------------------------------------------------------------------
    def _rows(self, group):
        """(이름, 부호, 산출식, V, cx, cy, cz) 목록. 요약 모드면 roll 단위로 묶는다."""
        parts = self.e.group_parts(group)
        rows = []
        seen = set()
        for p in parts:
            if (not self.detail) and p.roll:
                if p.roll in seen:
                    continue
                seen.add(p.roll)
                grp = [q for q in parts if q.roll == p.roll]
                V, cx, cy, cz = mass_props(grp)
                rows.append((p.roll, ' ', f"{len(grp)} 개 소계", V, cx, cy, cz))
            else:
                c = p.c
                rows.append((p.name, '+' if p.sign > 0 else '−',
                             p.prim.desc(), p.v, c[0], c[1], c[2]))
        return rows

    def _table_head(self):
        h = (_pad('부재명', CW[0]) + ' ' + _pad('부호', CW[1], 'center') + ' ' +
             _pad('산출식', CW[2]) + ' ' + _pad('체적 (m³)', CW[3], 'right') + ' ' +
             _pad('Xg', CW[4], 'right') + ' ' + _pad('Yg', CW[5], 'right') + ' ' +
             _pad('Zg', CW[6], 'right') + ' ' + _pad('V·Xg', CW[7], 'right') + ' ' +
             _pad('V·Zg', CW[8], 'right'))
        self._w(h)
        self._w(LINE3)

    def _table_row(self, nm, sg, fm, V, cx, cy, cz):
        # 이름·산출식은 폭에 맞춰 자른다 - 한 행이라도 넘치면 표가 어긋나 보인다
        self._w(_cut(nm, CW[0]) + ' ' +
                _pad(sg, CW[1], 'center') + ' ' +
                _cut(fm, CW[2]) + ' ' +
                _pad(f"{V:.3f}", CW[3], 'right') + ' ' +
                _pad(f"{cx:.3f}", CW[4], 'right') + ' ' +
                _pad(f"{cy:.3f}", CW[5], 'right') + ' ' +
                _pad(f"{cz:.3f}", CW[6], 'right') + ' ' +
                _pad(f"{V * cx:.2f}", CW[7], 'right') + ' ' +
                _pad(f"{V * cz:.2f}", CW[8], 'right'))

    def _sec2_parts(self):
        e = self.e
        self._w("■ 2. 부재별 체적 및 도심")
        self._w(LINE2)
        self._w(f"  표기 방식 : {'전체 (모든 부재 개별 표기)' if self.detail else '요약 (헌치·속채움·유공은 소계로 묶음)'}")
        self._w()
        for g in (GRP_CONC, GRP_FILL, GRP_COPING):
            rows = self._rows(g)
            if not rows:
                continue
            self._w(f"  ▶ {GRP_LABEL[g]}")
            self._table_head()
            for r in rows:
                self._table_row(*r)
            self._w(LINE3)
            V, cx, cy, cz = e.group_props(g)
            self._table_row(f"{GRP_LABEL[g]} 계", ' ', f"{len(rows)} 항목",
                            V, cx, cy, cz)
            self._w()

    # ------------------------------------------------------------------
    def _sec3_cells(self):
        e, sp = self.e, self.sp
        self._w("■ 3. 셀별 제원 및 속채움")
        self._w(LINE2)
        self._w("  ※ 채움 상단 z 는 저판 저면 기준 절대 표고이며 셀마다 따로 지정한다.")
        self._w("     구분이 유수실이어도 채움 상단을 주면 그 높이까지 채운다.")
        self._w("     덮개 t 는 유수실 상부 콘크리트 두께로, 정단에서 그만큼 아래로 얹힌다.")
        self._w("     '-' 는 없는 것이다. 속채움 체적은 셀 공극에서 헌치 점유분을 뺀 값이다.")
        self._w()
        cwc = CWC
        head = ' '.join(_pad(t, w, a) for t, w, a in
                        zip(('셀', '구분', 'x 범위', 'y 범위', '순치수 (x × y)',
                             '채움상단 z', '채움높이', '덮개 t', '속채움 V (m³)', 'Zg'),
                            cwc,
                            ('left', 'left', 'left', 'left', 'right',
                             'right', 'right', 'right', 'right', 'right')))
        self._w(head)
        self._w(LINE3)
        tot = tcov = 0.0
        for c in e.cells:
            fp = [p for p in e.parts
                  if p.group == GRP_FILL and p.roll == f"속채움 {c['name']}"]
            V, cx, cy, cz = mass_props(fp) if fp else (0.0, 0.0, 0.0, 0.0)
            tot += V
            cv = c['cover']
            tcov += cv * (c['x1'] - c['x0']) * (c['y1'] - c['y0'])
            self._w(_pad(c['name'], cwc[0]) + ' ' + _pad(c['ctype'], cwc[1]) + ' ' +
                    _pad(f"{c['x0']:.3f} ~ {c['x1']:.3f}", cwc[2]) + ' ' +
                    _pad(f"{c['y0']:.3f} ~ {c['y1']:.3f}", cwc[3]) + ' ' +
                    _pad(f"{c['x1'] - c['x0']:.3f} × {c['y1'] - c['y0']:.3f}",
                         cwc[4], 'right') + ' ' +
                    _pad(f"{c['fill_top']:.3f}" if fp else '-', cwc[5], 'right') + ' ' +
                    _pad(f"{c['fill_top'] - sp.tb:.3f}" if fp else '-',
                         cwc[6], 'right') + ' ' +
                    _pad(f"{cv:.3f}" if cv > EPS else '-', cwc[7], 'right') + ' ' +
                    _pad(f"{V:.3f}" if fp else '-', cwc[8], 'right') + ' ' +
                    _pad(f"{cz:.3f}" if fp else '-', cwc[9], 'right'))
        self._w(LINE3)
        # 앞 8 열 + 그 사이 구분 공백 7 칸 = 속채움 V 열 바로 앞까지
        self._w(_pad('속채움 합계', sum(cwc[:8]) + 7, 'right') +
                ' ' + _pad(f"{tot:.3f}", cwc[8], 'right'))
        if tcov > EPS:
            self._w(_pad('덮개 합계', sum(cwc[:8]) + 7, 'right') +
                    ' ' + _pad(f"{tcov:.3f}", cwc[8], 'right'))
        self._w()

    # ------------------------------------------------------------------
    def _sec4_groups(self):
        e = self.e
        self._w("■ 4. 부재군별 집계")
        self._w(LINE2)
        head = (_pad('부재군', 18) + ' ' + _pad('체적 (m³)', 14, 'right') + ' ' +
                _pad('Xg (m)', 12, 'right') + ' ' + _pad('Yg (m)', 12, 'right') + ' ' +
                _pad('Zg (m)', 12, 'right') + ' ' + _pad('e = Xg−B/2', 14, 'right'))
        self._w(head)
        self._w(LINE3)
        for g in (GRP_CONC, GRP_FILL, GRP_COPING):
            V, cx, cy, cz = e.group_props(g)
            if abs(V) < 1e-12:
                continue
            self._w(_pad(GRP_LABEL[g], 18) + ' ' +
                    _pad(f"{V:.3f}", 14, 'right') + ' ' +
                    _pad(f"{cx:.3f}", 12, 'right') + ' ' +
                    _pad(f"{cy:.3f}", 12, 'right') + ' ' +
                    _pad(f"{cz:.3f}", 12, 'right') + ' ' +
                    _pad(f"{cx - self.sp.B / 2:+.3f}", 14, 'right'))
        self._w(LINE3)
        Vt, xt, yt, zt = mass_props(e.parts)
        self._w(_pad('전 부재 (체적 가중)', 18) + ' ' +
                _pad(f"{Vt:.3f}", 14, 'right') + ' ' +
                _pad(f"{xt:.3f}", 12, 'right') + ' ' +
                _pad(f"{yt:.3f}", 12, 'right') + ' ' +
                _pad(f"{zt:.3f}", 12, 'right') + ' ' +
                _pad(f"{xt - self.sp.B / 2:+.3f}", 14, 'right'))
        self._w("  ※ '전 부재' 행은 재료를 구분하지 않은 단순 체적 가중 도심이다.")
        self._w("     재료가 다른 부재군의 무게중심은 아래 §5 의 합성값을 쓴다.")
        self._w()

    # ------------------------------------------------------------------
    def _sec5_summary(self):
        e, sp = self.e, self.sp
        self._w("■ 5. 종합")
        self._w(LINE2)
        Vc = e.group_props(GRP_CONC)
        Vf = e.group_props(GRP_FILL)
        Vp = e.group_props(GRP_COPING)
        self._w(f"  · 케이슨 콘크리트 체적   V = {Vc[0]:12.3f} m³")
        self._w(f"  · 속채움 체적            V = {Vf[0]:12.3f} m³")
        self._w(f"  · 상치콘크리트 체적      V = {Vp[0]:12.3f} m³")
        self._w(f"    {'-' * 46}")
        self._w(f"  · 합 계                  V = {Vc[0] + Vf[0] + Vp[0]:12.3f} m³")
        self._w()
        cb = e.combined()
        if cb is None:
            self._w("  · 합성 무게중심 : 단위중량 미사용 - 부재군별 도심(§4)을 그대로 쓴다.")
            self._w()
            return
        self._w("  · 합성 무게중심   [수식]  Xg = Σ(V·γ·x) / Σ(V·γ)")
        self._w()
        head = (_pad('부재군', 18) + ' ' + _pad('V (m³)', 12, 'right') + ' ' +
                _pad('γ (kN/m³)', 12, 'right') + ' ' + _pad('W = V·γ (kN)', 14, 'right') + ' ' +
                _pad('Xg', 9, 'right') + ' ' + _pad('Yg', 9, 'right') + ' ' +
                _pad('Zg', 9, 'right'))
        self._w("    " + head)
        self._w("    " + "·" * ((_dw(head)) // _AMB_W))
        for g, V, gam, Wt, cx, cy, cz in cb['rows']:
            self._w("    " + _pad(GRP_LABEL[g], 18) + ' ' +
                    _pad(f"{V:.3f}", 12, 'right') + ' ' +
                    _pad(f"{gam:.2f}", 12, 'right') + ' ' +
                    _pad(f"{Wt:.1f}", 14, 'right') + ' ' +
                    _pad(f"{cx:.3f}", 9, 'right') + ' ' +
                    _pad(f"{cy:.3f}", 9, 'right') + ' ' +
                    _pad(f"{cz:.3f}", 9, 'right'))
        self._w("    " + "·" * ((_dw(head)) // _AMB_W))
        self._w("    " + _pad('합 계', 18) + ' ' +
                _pad('', 12) + ' ' + _pad('', 12) + ' ' +
                _pad(f"{cb['W']:.1f}", 14, 'right') + ' ' +
                _pad(f"{cb['gx']:.3f}", 9, 'right') + ' ' +
                _pad(f"{cb['gy']:.3f}", 9, 'right') + ' ' +
                _pad(f"{cb['gz']:.3f}", 9, 'right'))
        self._w()
        self._w(f"    → 무게중심  Xg = {cb['gx']:.3f} m (전면 외면 기준)")
        self._w(f"                Yg = {cb['gy']:.3f} m (좌측 단부 기준)")
        self._w(f"                Zg = {cb['gz']:.3f} m (저판 저면 기준)")
        self._w(f"                e  = Xg − B/2 = {cb['gx']:.3f} − {sp.B / 2:.3f}"
                f" = {cb['gx'] - sp.B / 2:+.3f} m  (폭 중앙 기준 편심)")
        if sp.hole_on:
            r, _, _ = e.opening_ratio()
            self._w(f"                개구율 = {r * 100:.2f} %")
        self._w()

    # ------------------------------------------------------------------
    def _sec6_check(self):
        e, sp = self.e, self.sp
        ck = e.check
        q = e.q
        self._w("■ 6. 자동 검산")
        self._w(LINE2)
        self._w("  케이슨 외곽 직육면체의 공극을 서로 다른 두 경로로 산정해 대조한다.")
        self._w("  격벽 교차부 중복공제 누락, 부재 이중 계상 같은 오류가 여기서 드러난다.")
        self._w()
        self._w("  ▶ 경로 ① 부재 분해")
        self._w("      ※ 풋팅·전단키·상치는 외곽 직육면체 밖이므로 이 검산에서 제외한다.")
        self._w(f"      [수식] V공극 = B·L·H − Σ(외곽 내부 콘크리트 부재)")
        self._w(f"      [대입] = {sp.B:.3f} × {sp.L:.3f} × {sp.H:.3f} − {ck['v_conc_in']:.6f}")
        self._w(f"      [계산] = {ck['v_outer']:.6f} − {ck['v_conc_in']:.6f}")
        self._w(f"      → {ck['void_parts']:.6f} m³")
        self._w()
        self._w("  ▶ 경로 ② 셀 격자")
        self._w(f"      [수식] V공극 = Σ(셀 순면적 × 벽체높이) + 감격벽 상부"
                f" − 헌치 + 유공")
        self._w(f"      [대입] = {q['v_cells_box']:.6f} + {q['v_extra']:.6f}"
                f" − {q['v_hn']:.6f} + {q['v_hole']:.6f}")
        self._w(f"      → {ck['void_cells']:.6f} m³")
        self._w()
        self._w(f"  ▶ 잔차 Δ = ① − ② = {ck['delta']:+.9f} m³"
                f"   (외곽체적 대비 {ck['rel'] * 100:.9f} %)")
        self._w(f"      → {'O.K  두 경로가 일치한다.' if ck['ok'] else 'N.G  ★ 부재 분해에 중복 또는 누락이 있다. 결과를 쓰지 마시오.'}")
        self._w()
        self._w(f"  ▶ 좌우 대칭 : 콘크리트 Yg = {mass_props(e.group_parts(GRP_CONC))[2]:.6f} m"
                f"  vs  L/2 = {sp.L / 2:.6f} m   (차 {ck['sym_err']:.9f} m)")
        if ck['sym_ok']:
            self._w("      → O.K  좌우 대칭 입력과 부합한다.")
        else:
            self._w("      → 좌우 비대칭. 격벽·유공·전단키를 비대칭으로 배치했다면 정상이다.")
        self._w()
        mc = getattr(e, 'mc', None)
        if mc:
            V_lump = sum(p.v for p in e.parts
                         if p.group == GRP_CONC and not p.prim.samplable)
            Va = e.group_props(GRP_CONC)[0] - V_lump
            err = abs(mc['V'] - Va) / Va * 100 if Va > EPS else 0.0
            nsig = abs(mc['V'] - Va) / mc['sigma'] if mc['sigma'] > EPS else 0.0
            self._w("  ▶ 몬테카를로 형상 검증 (해석적 분해와 독립한 경로)")
            self._w(f"      표본 {mc['n']:,} 점 / 바운딩박스 {mc['vbb']:.3f} m³")
            self._w(f"      해석해 V = {Va:.3f} m³   몬테카를로 V = {mc['V']:.3f}"
                    f" ± {mc['sigma']:.3f} m³ (1σ)")
            self._w(f"      차 = {err:.3f} % = {nsig:.2f} σ"
                    f"   (표본 추출에 따른 통계 요동이며 3σ 이내면 정상이다)")
            self._w(f"      부호합 ≥ 2 (이중 계상) : {mc['bad_over']:,} 점")
            self._w(f"      부호합 <  0 (과공제)   : {mc['bad_under']:,} 점")
            if mc['skipped']:
                self._w(f"      ※ 코너 보정 {len(mc['skipped'])} 개(합 {V_lump:.3f} m³)는"
                        f" 점 포함 판정이 불가능해 양쪽 모두에서 제외했다.")
            ok = (mc['bad_over'] == 0 and mc['bad_under'] == 0
                  and abs(mc['V'] - Va) <= max(3 * mc['sigma'], Va * 0.005))
            self._w(f"      → {'O.K' if ok else 'N.G  ★ 부재 분해를 재검토하시오.'}")
            self._w()

    # ------------------------------------------------------------------
    def _sec7_notes(self):
        sp = self.sp
        self._w("■ 7. 적용 가정 및 본 프로그램의 범위 외")
        self._w(LINE2)
        self._w("  ▶ 적용 가정")
        self._w("    · 측벽은 전·후면벽 사이 구간만 차지시켜 모서리 중복을 없앴다.")
        self._w("    · 종격벽 × 횡격벽 교차부는 전 조합을 개별 부재로 공제했다 (§2 참조).")
        self._w("    · 풋팅은 링(외곽 − 본체) 방식으로 산정해 모서리 중복이 발생하지 않는다.")
        if sp.hn_on:
            a, b = sp.hn_a, sp.hn_b
            self._w("    · 헌치 교차부(코너)는 y방향 헌치를 우선하고 x방향 헌치를 각 단부에서")
            self._w(f"      a = {a:.3f} m 만큼 절단해 중복을 제거했다. 이때 코너마다 누락되는")
            self._w(f"      체적은 적분 결과 정확히 a²b/6 = {a * a * b / 6:.5f} m³ 이며,")
            if sp.hn_corner:
                self._w("      [헌치 코너 보정] 을 켜 정확한 도심 (a/4, 3a/4, b/2) 과 함께 되살렸다.")
            else:
                self._w("      [헌치 코너 보정] 이 꺼져 있어 그만큼 과소 산정된다.")
        if sp.vh_on:
            c = sp.vh_c
            z0 = self.e._vh_z0()
            self._w(f"    · 수직(우각부) 헌치는 셀의 4개 연직 모서리에 다리 c = {c:.3f} m 의")
            self._w("      45° 삼각단면으로 세운다. 격벽-격벽, 격벽-외벽, 외벽-외벽 모서리가")
            self._w("      모두 해당하며, 코너를 이루는 두 벽 중 낮은 쪽까지만 올린다(감격벽).")
            if sp.hn_on:
                self._w(f"      수평 헌치 상단 z = {z0:.3f} m 부터 세우므로 두 헌치는 겹치지 않는다.")
                self._w(f"      높이 w 에서 수평 헌치가 차지하는 폭은 s = a(1 − w/b) 이고")
                self._w(f"      w ≥ b 이면 s = 0 이기 때문이다.")
                if sp.hn_corner:
                    m_ = 2 * sp.hn_a / c if c > EPS else 0.0
                    I0 = self.e._vh_corner_integrals(m_)[0] if m_ > 0 else 0.0
                    self._w(f"      그 아래 구간에서 수직 헌치가 수평 헌치 밖으로 삐져나오는 부분은")
                    self._w(f"      코너당 (c²/2)·b·I0 = {c * c / 2:.5f} × {sp.hn_b:.3f}"
                            f" × {I0:.5f} = {c * c / 2 * sp.hn_b * I0:.6f} m³ 이며,")
                    self._w("      [헌치 코너 보정] 이 정확한 도심과 함께 되살린다.")
            self._w("    · 속채움은 수평 헌치뿐 아니라 수직 헌치 점유분도 공제한다. 채움 상단이")
            self._w("      수직 헌치 중간을 자르면 그 높이에서 잘라 정확히 공제한다.")
        self._w("    · 유공은 대상 벽체에서만 공제한다. 유공이 격벽 바로 앞에 놓이면")
        self._w("      체적은 정확하나 설계상 막히므로 경고를 표시한다.")
        self._w("    · [3D 모델 저장] 의 CAD 메시는 서로 겹치지 않는 분해로 만든다.")
        self._w("      풋팅은 링 4상자, 횡격벽은 종격벽 사이로 쪼개고, 헌치는 실제")
        self._w("      형상인 마이터 액자와 우각부 로프트로, 속채움은 그 헌치를 뺀")
        self._w("      팔각형 로프트로 만든다. 따라서 부재군별 체적이 계산서와 같으며,")
        self._w(f"      유일한 차이는 원형 유공을 {CIRCLE_SEG} 각형으로 근사한 분이다.")
        self._w("      저장할 때 부재군별 체적 대조표와 판정을 함께 표시한다.")
        self._w("    · 속채움은 셀(격실)마다 채움 상단 z 를 따로 지정하며, 구분이 유수실이어도")
        self._w("      값을 주면 채운다. 각 셀의 속채움에서 그 셀의 헌치 점유분을 공제한다.")
        self._w("      채움이 유공 하단보다 높으면 체적은 정확하나 유공이 막히므로 경고한다.")
        self._w()
        self._w("  ▶ 범위 외")
        self._w("    중량 · 부력 · 수중중량 · 흘수 · 경심고 GM · 진수/예항 안정,")
        self._w("    활동 · 전도 · 지지력 검토, 철근량, 부재 응력.")
        self._w("    ※ 단위중량은 재료가 다른 부재군을 합성하기 위한 가중치로만 쓴다.")
        self._w()
        self._w(LINE1)
        self._w(_pad("이상", W, 'center'))
        self._w(LINE1)


# ==========================================
# 4. 도해
#    입력 제원을 눈으로 검증하기 위한 것이다. 치수는 계산서를 따른다.
# ==========================================

C_BASE = '#95A5A6'
C_WALL = '#5DADE2'
C_PART = '#58D68D'
C_FOOT = '#7F8C8D'
C_KEY = '#616A6B'
C_COP = '#F5B041'
C_HAUNCH = '#AF7AC5'
C_HOLE = '#E74C3C'
C_CG = '#C0392B'
CELL_COLOR = {CELL_FILL: '#F9E79F', CELL_CHAMBER: '#AED6F1', CELL_EMPTY: '#FDFEFE'}


def _cat(name):
    """부재명 → 도해 분류."""
    if name.startswith('저판'):
        return 'base'
    if name.startswith(('전면벽', '후면벽', '좌측벽', '우측벽')):
        return 'wall'
    if name.startswith(('종격벽', '횡격벽')):
        return 'part'
    if name.startswith('풋팅'):
        return 'foot'
    if name.startswith('전단키'):
        return 'key'
    if name.startswith('상치'):
        return 'cop'
    if name.startswith('헌치'):
        return 'haunch'
    if name.startswith('덮개'):
        return 'cover'
    return 'etc'


CAT_COLOR = {'base': C_BASE, 'wall': C_WALL, 'part': C_PART, 'foot': C_FOOT,
             'key': C_KEY, 'cop': C_COP, 'haunch': C_HAUNCH, 'cover': '#5499C7',
             'etc': '#BDC3C7'}

# ------------------------------------------------------------------
# 도해 표시 그룹 — 네 뷰 모두 같은 기준으로 켜고 끈다
# ------------------------------------------------------------------
SHOW_BODY = 'body'        # 저판 · 외벽 · 격벽
SHOW_HAUNCH = 'haunch'    # 수평 · 수직 헌치
SHOW_FILL = 'fill'        # 속채움
SHOW_COPING = 'coping'    # 상치콘크리트 · 파라펫
SHOW_FOOT = 'foot'        # 풋팅 · 전단키
SHOW_HOLE = 'hole'        # 유공
SHOW_ALL = (SHOW_BODY, SHOW_HAUNCH, SHOW_FILL, SHOW_COPING, SHOW_FOOT, SHOW_HOLE)
SHOW_LABEL = {SHOW_BODY: '케이슨 본체', SHOW_HAUNCH: '헌치', SHOW_FILL: '속채움',
              SHOW_COPING: '상치콘크리트', SHOW_FOOT: '풋팅·전단키',
              SHOW_HOLE: '유공'}

#: _cat() 분류 → 표시 그룹
CAT_SHOW = {'base': SHOW_BODY, 'wall': SHOW_BODY, 'part': SHOW_BODY,
            'haunch': SHOW_HAUNCH, 'foot': SHOW_FOOT, 'key': SHOW_FOOT,
            'cop': SHOW_COPING, 'cover': SHOW_BODY, 'etc': SHOW_BODY}


def _uvw_to_xyz(axis, u, v, w):
    iu, iv, iw = _AXIS_UVW[axis]
    p = [0.0, 0.0, 0.0]
    p[iu], p[iv], p[iw] = u, v, w
    return tuple(p)


def _box_polys(b: Box):
    x0, y0, z0 = b.x0, b.y0, b.z0
    x1, y1, z1 = x0 + b.dx, y0 + b.dy, z0 + b.dz
    return [
        [(x0, y0, z0), (x1, y0, z0), (x1, y1, z0), (x0, y1, z0)],
        [(x0, y0, z1), (x1, y0, z1), (x1, y1, z1), (x0, y1, z1)],
        [(x0, y0, z0), (x1, y0, z0), (x1, y0, z1), (x0, y0, z1)],
        [(x0, y1, z0), (x1, y1, z0), (x1, y1, z1), (x0, y1, z1)],
        [(x0, y0, z0), (x0, y1, z0), (x0, y1, z1), (x0, y0, z1)],
        [(x1, y0, z0), (x1, y1, z0), (x1, y1, z1), (x1, y0, z1)],
    ]


def _tri_polys(t: TriPrism):
    ax = t.axis
    w0, w1 = t.w0, t.w0 + t.h
    A = (t.u0, t.v0)
    Bv = (t.u0 + t.du, t.v0)
    Cv = (t.u0, t.v0 + t.dv)
    f = lambda p, w: _uvw_to_xyz(ax, p[0], p[1], w)
    return [
        [f(A, w0), f(Bv, w0), f(Cv, w0)],
        [f(A, w1), f(Bv, w1), f(Cv, w1)],
        [f(A, w0), f(Bv, w0), f(Bv, w1), f(A, w1)],
        [f(Bv, w0), f(Cv, w0), f(Cv, w1), f(Bv, w1)],
        [f(Cv, w0), f(A, w0), f(A, w1), f(Cv, w1)],
    ]


class Plotter:
    """엔진 결과를 3D / 평면 / 정면 / 측면 도해로 그린다."""

    def __init__(self, eng: CaissonEngine, show=None):
        self.e = eng
        self.sp = eng.sp
        #: 표시할 부재군. 없으면 전부 그린다.
        self.show = set(SHOW_ALL if show is None else show)

    def on(self, key):
        return key in self.show

    def _hidden_note(self):
        """제목에 덧붙일 '무엇을 뺀 그림인지' 문구."""
        off = [SHOW_LABEL[k] for k in SHOW_ALL if k not in self.show]
        return f"  ({'·'.join(off)} 숨김)" if off else ''

    def _cg(self):
        """도해에 표시할 무게중심. 단위중량을 쓰면 합성값, 아니면 콘크리트 도심."""
        cb = self.e.combined()
        if cb:
            return cb['gx'], cb['gy'], cb['gz'], '합성 무게중심'
        V, x, y, z = self.e.group_props(GRP_CONC)
        return x, y, z, '콘크리트 도심'

    # ------------------------------------------------------------------
    def draw_3d(self, ax):
        """3D 형상.

        내부(격벽·유공)를 확인하는 것이 목적이므로 면은 거의 투명하게 두고
        모서리 선으로 형상을 읽게 한다. 면을 불투명하게 칠하면 바깥 벽이
        내부를 전부 가려 검증에 쓸 수 없다.
        """
        e, sp = self.e, self.sp
        ax.clear()
        used = set()
        for p in e.parts:
            if p.sign < 0 or p.group == GRP_FILL:
                continue
            cat = _cat(p.name)
            if not self.on(CAT_SHOW.get(cat, SHOW_BODY)):
                continue
            if isinstance(p.prim, Box):
                polys = _box_polys(p.prim)
            elif isinstance(p.prim, TriPrism):
                polys = _tri_polys(p.prim)
            else:
                continue
            alpha = 0.28 if cat in ('base', 'foot', 'key') else 0.11
            col = Poly3DCollection(polys, facecolors=CAT_COLOR[cat],
                                   edgecolors=CAT_COLOR[cat], linewidths=0.8,
                                   alpha=alpha)
            ax.add_collection3d(col)
            used.add(cat)

        if self.on(SHOW_FILL):
            for p in e.parts:
                if p.group != GRP_FILL or p.sign < 0 or not isinstance(p.prim, Box):
                    continue
                ax.add_collection3d(Poly3DCollection(
                    _box_polys(p.prim), facecolors='#F7DC6F',
                    edgecolors='#B7950B', linewidths=0.4, alpha=0.10))
                used.add('fill')

        # 유공 - 전면/후면 두 원을 테두리로 표시
        for h in (e.holes if self.on(SHOW_HOLE) else []):
            th = np.linspace(0, 2 * np.pi, 48)
            for ww in (h['w0'], h['w0'] + h['t']):
                if h['circle']:
                    uu = h['u'] + h['hu'] * np.cos(th)
                    vv = h['v'] + h['hu'] * np.sin(th)
                else:
                    uu = np.array([h['u'] - h['hu'], h['u'] + h['hu'],
                                   h['u'] + h['hu'], h['u'] - h['hu'],
                                   h['u'] - h['hu']])
                    vv = np.array([h['v'] - h['hv'], h['v'] - h['hv'],
                                   h['v'] + h['hv'], h['v'] + h['hv'],
                                   h['v'] - h['hv']])
                wl = np.full_like(uu, ww, dtype=float)
                # 축에 따라 (u, v, w) → (x, y, z)
                if h['axis'] == 'x':
                    ax.plot(wl, uu, vv, color=C_HOLE, lw=1.2)
                else:
                    ax.plot(uu, wl, vv, color=C_HOLE, lw=1.2)
        if e.holes and self.on(SHOW_HOLE):
            used.add('hole')

        z_lo = min(0.0, -max([k.d for k in sp.keys], default=0.0))
        z_hi = sp.H + ((sp.cp_h + (sp.cp_par_h if sp.cp_par_on else 0))
                       if sp.cp_on else 0)

        gx, gy, gz, glabel = self._cg()
        ax.plot([gx, gx], [gy, gy], [z_lo, gz], ls=':', c=C_CG, lw=1.2)
        ax.plot([gx, gx], [0, gy], [gz, gz], ls=':', c=C_CG, lw=1.2)
        ax.plot([0, gx], [gy, gy], [gz, gz], ls=':', c=C_CG, lw=1.2)
        ax.plot([gx], [gy], [gz], marker='X', ms=13, c=C_CG, ls='none')
        ax.text(gx, gy, gz, f"  G ({gx:.2f}, {gy:.2f}, {gz:.2f})",
                color=C_CG, fontsize=8.5, weight='bold')

        pad = max(sp.B, sp.L) * 0.03
        ax.set_xlim(-sp.ft_front - pad, sp.B + sp.ft_rear + pad)
        ax.set_ylim(-pad, sp.L + pad)
        ax.set_zlim(z_lo - pad, z_hi + pad)
        ax.set_box_aspect((sp.B, sp.L, max(z_hi - z_lo, 1.0)))
        ax.set_xlabel('x - 폭 (전면 → 후면)', fontsize=8)
        ax.set_ylabel('y - 길이', fontsize=8)
        ax.set_zlabel('z - 높이', fontsize=8)
        ax.tick_params(labelsize=7)
        ax.set_title(f"3D 형상 - {glabel} G 표시{self._hidden_note()}", fontsize=10)
        # 전면벽(x = 0)이 보이는 방향에서 내려다본다
        ax.view_init(elev=22, azim=-125)

        labels = [('base', '저판'), ('wall', '외벽'), ('part', '격벽'),
                  ('haunch', '헌치'), ('foot', '풋팅'), ('key', '전단키'),
                  ('cop', '상치'), ('fill', '속채움'), ('hole', '유공')]
        handles = []
        for k, lb in labels:
            if k not in used:
                continue
            c = C_HOLE if k == 'hole' else ('#F7DC6F' if k == 'fill' else CAT_COLOR[k])
            handles.append(plt.Line2D([], [], color=c, lw=3, label=lb))
        handles.append(plt.Line2D([], [], color=C_CG, marker='X', ls='none',
                                  ms=9, label='무게중심 G'))
        ax.legend(handles=handles, loc='upper left', fontsize=7.5,
                  framealpha=0.85, ncol=2)

    # ------------------------------------------------------------------
    def draw_plan(self, ax):
        """평면도 (x-y) - 격벽 배치와 셀 구분."""
        e, sp = self.e, self.sp
        ax.clear()
        if self.on(SHOW_BODY):
            ax.add_patch(Rectangle((0, 0), sp.B, sp.L, fc='#D5D8DC',
                                   ec='#2C3E50', lw=1.2))
        for c in e.cells:
            filled = c['fill_top'] > sp.tb + EPS
            if self.on(SHOW_BODY) or (self.on(SHOW_FILL) and filled):
                # 본체를 끄고 속채움만 볼 때는 채워진 셀만 남긴다
                fc = (CELL_COLOR.get(c['ctype'], '#FFFFFF')
                      if self.on(SHOW_BODY) else '#F7DC6F')
                ax.add_patch(Rectangle((c['x0'], c['y0']),
                                       c['x1'] - c['x0'], c['y1'] - c['y0'],
                                       fc=fc, ec='#7F8C8D', lw=0.6))
                lb = f"{c['name']}\n{c['ctype']}" if self.on(SHOW_BODY) else c['name']
                if filled and self.on(SHOW_FILL):
                    lb += f"\n채움 z={c['fill_top']:.2f}"
                if c['cover'] > EPS and self.on(SHOW_BODY):
                    lb += f"\n덮개 t={c['cover']:.2f}"
                ax.text((c['x0'] + c['x1']) / 2, (c['y0'] + c['y1']) / 2, lb,
                        ha='center', va='center', fontsize=7, color='#2C3E50')
        # 수직(우각부) 헌치 - 평면에서 셀 4모서리의 45° 삼각형으로 보인다
        if sp.vh_on and e.cells and self.on(SHOW_HAUNCH):
            vc = sp.vh_c
            for c in e.cells:
                for _tag, x, y, sx, sy, hcor in e._cell_corners(c):
                    if sp.tb + hcor - e._vh_z0() <= EPS:
                        continue        # 감격벽이라 헌치가 서지 않는 코너
                    ax.add_patch(MplPolygon(
                        [(x, y), (x + sx * vc, y), (x, y + sy * vc)],
                        fc=C_HAUNCH, ec='#4A235A', lw=0.5, alpha=0.85))
        if max(sp.ft_front, sp.ft_rear, sp.ft_side) > EPS and self.on(SHOW_FOOT):
            ax.add_patch(Rectangle((-sp.ft_front, -sp.ft_side),
                                   sp.B + sp.ft_front + sp.ft_rear,
                                   sp.L + 2 * sp.ft_side,
                                   fc='none', ec=C_FOOT, lw=1.0, ls='--'))
        # 유공은 관통공이므로 평면에서는 직사각형이다. 축과 무관하게 도형의
        # 바운딩박스를 그대로 쓰면 x축·y축 벽 모두 맞는다.
        for h in (e.holes if self.on(SHOW_HOLE) else []):
            bx0, bx1, by0, by1, _z0, _z1 = h['prim'].bbox()
            ax.add_patch(Rectangle((bx0, by0), bx1 - bx0, by1 - by0,
                                   fc='#FDEDEC', ec=C_HOLE, lw=0.9))
        gx, gy, gz, glabel = self._cg()
        ax.plot([gx], [gy], marker='X', ms=11, color=C_CG, zorder=10)
        ax.axvline(gx, ls=':', c=C_CG, lw=0.8)
        ax.axhline(gy, ls=':', c=C_CG, lw=0.8)
        ax.text(gx, gy, f"  G ({gx:.2f}, {gy:.2f})", color=C_CG,
                fontsize=8, weight='bold', va='bottom')
        pad = max(sp.B, sp.L) * 0.06
        ax.set_xlim(-sp.ft_front - pad, sp.B + sp.ft_rear + pad)
        ax.set_ylim(-pad, sp.L + pad)
        ax.set_aspect('equal')
        ax.set_xlabel('x - 폭 (전면 0 → 후면 B)')
        ax.set_ylabel('y - 길이')
        ax.set_title(f"평면도 - 셀 {len(e.cells)} 개 "
                     f"(속채움 {sum(1 for c in e.cells if c['ctype'] == CELL_FILL)} / "
                     f"유수실 {sum(1 for c in e.cells if c['ctype'] == CELL_CHAMBER)} / "
                     f"공셀 {sum(1 for c in e.cells if c['ctype'] == CELL_EMPTY)}),  "
                     f"채움 적용 {sum(1 for c in e.cells if c['fill_top'] > sp.tb + EPS)} 셀"
                     + self._hidden_note(), fontsize=10)
        ax.grid(alpha=0.25, ls=':')

    # ------------------------------------------------------------------
    def draw_front(self, ax):
        """정면도 (y-z) - 유공 배치."""
        e, sp = self.e, self.sp
        ax.clear()
        if self.on(SHOW_BODY):
            ax.add_patch(Rectangle((0, 0), sp.L, sp.tb, fc=C_BASE,
                                   ec='#2C3E50', lw=1.0))
            ax.add_patch(Rectangle((0, sp.tb), sp.L, e.hw, fc=C_WALL,
                                   ec='#2C3E50', lw=1.2, alpha=0.75))
        if sp.cp_on and self.on(SHOW_COPING):
            z0 = sp.cp_z0 if sp.cp_z0 > EPS else sp.H
            ax.add_patch(Rectangle((0, z0), sp.L, sp.cp_h, fc=C_COP,
                                   ec='#2C3E50', lw=1.0, alpha=0.8))
            if sp.cp_par_on:
                ax.add_patch(Rectangle((0, z0 + sp.cp_h), sp.L, sp.cp_par_h,
                                       fc=C_COP, ec='#2C3E50', lw=1.0, alpha=0.6))
        for j, d in enumerate(e.tps if self.on(SHOW_BODY) else []):
            ax.add_patch(Rectangle((d['lo'], sp.tb), d['t'], d['h'],
                                   fc='none', ec=C_PART, lw=1.0, ls='--'))
            ax.text(d['pos'], sp.tb + d['h'], f"T#{j + 1}", ha='center',
                    va='bottom', fontsize=7, color='#1E8449')
        # 속채움 - 정면에서는 벽 뒤라 보이지 않으므로 칸별 채움 상단을 파선으로 알린다
        if self.on(SHOW_FILL):
            seen = set()
            for c in e.cells:
                key = (round(c['y0'], 6), round(c['y1'], 6),
                       round(c['fill_top'], 6))
                if c['fill_top'] <= sp.tb + EPS or key in seen:
                    continue
                seen.add(key)
                ax.plot([c['y0'], c['y1']], [c['fill_top']] * 2,
                        ls='--', c='#B7950B', lw=1.1, zorder=6)
                # 상단선 아래(벽체 안쪽)에 적는다 - 채움이 정단까지면 위쪽은
                # 상치·격벽 표기와 겹친다
                ax.text((c['y0'] + c['y1']) / 2, c['fill_top'],
                        f"채움 z={c['fill_top']:.2f}", ha='center', va='top',
                        fontsize=6.5, color='#7D6608', zorder=6)
        # 정면도는 y-z 평면이므로 x축 관통 벽(전면·후면·종격벽)의 유공만 보인다
        for h in (e.holes if self.on(SHOW_HOLE) else []):
            if h['axis'] != 'x':
                continue
            if h['circle']:
                ax.add_patch(Circle((h['u'], h['v']), h['hu'],
                                    fc='#FDEDEC', ec=C_HOLE, lw=1.2))
            else:
                ax.add_patch(Rectangle((h['u'] - h['hu'], h['v'] - h['hv']),
                                       2 * h['hu'], 2 * h['hv'],
                                       fc='#FDEDEC', ec=C_HOLE, lw=1.2))
            ax.text(h['u'], h['v'], str(h['i']), ha='center', va='center',
                    fontsize=7, color=C_HOLE)
        r, area, face = e.opening_ratio()
        pad = sp.L * 0.05
        top = sp.H + ((sp.cp_h + (sp.cp_par_h if sp.cp_par_on else 0))
                      if sp.cp_on else 0)
        ax.set_xlim(-pad, sp.L + pad)
        ax.set_ylim(-pad, top + pad)
        ax.set_aspect('equal')
        ax.set_xlabel('y - 길이')
        ax.set_ylabel('z - 높이 (저판 저면 = 0)')
        nx = sum(1 for h in e.holes if h['axis'] == 'x')
        wl = ' · '.join(g['name'] for g in getattr(e, 'wall_geoms', [])
                        if g['axis'] == 'x')
        ttl = (f"정면도 ({wl}) - 유공 {nx} 개,  개구율 {r * 100:.2f} %"
               if nx else "정면도 - x축 관통 유공 없음")
        ax.set_title(ttl + self._hidden_note(), fontsize=10)
        ax.grid(alpha=0.25, ls=':')

    # ------------------------------------------------------------------
    def draw_side(self, ax):
        """측면도 (x-z) - 셀 중앙을 지나는 횡단면."""
        e, sp = self.e, self.sp
        ax.clear()
        # 절단 위치 : 첫 셀의 y 중앙 (횡격벽을 피해 종격벽이 보이도록)
        ycut = ((e.cells[0]['y0'] + e.cells[0]['y1']) / 2.0
                if e.cells else sp.L / 2.0)

        if max(sp.ft_front, sp.ft_rear) > EPS and self.on(SHOW_FOOT):
            ax.add_patch(Rectangle((-sp.ft_front, 0),
                                   sp.B + sp.ft_front + sp.ft_rear, sp.ft_t,
                                   fc=C_FOOT, ec='#2C3E50', lw=1.0))
        if self.on(SHOW_BODY):
            ax.add_patch(Rectangle((0, 0), sp.B, sp.tb, fc=C_BASE,
                                   ec='#2C3E50', lw=1.2))
        for i, k in enumerate(sorted(sp.keys, key=lambda q: q.x)
                              if self.on(SHOW_FOOT) else []):
            ax.add_patch(Rectangle((k.x, -k.d), k.w, k.d, fc=C_KEY,
                                   ec='#2C3E50', lw=1.0))
        if self.on(SHOW_BODY):
            ax.add_patch(Rectangle((0, sp.tb), sp.tf, e.hw, fc=C_WALL,
                                   ec='#2C3E50', lw=1.2))
            ax.add_patch(Rectangle((sp.B - sp.tr, sp.tb), sp.tr, e.hw, fc=C_WALL,
                                   ec='#2C3E50', lw=1.2))
        for i, d in enumerate(e.lps if self.on(SHOW_BODY) else []):
            ax.add_patch(Rectangle((d['lo'], sp.tb), d['t'], d['h'], fc=C_PART,
                                   ec='#2C3E50', lw=1.0))
            ax.text(d['pos'], sp.tb + d['h'], f"L#{i + 1}", ha='center',
                    va='bottom', fontsize=7, color='#1E8449')
        # 유공 - x축 벽은 절단면에 걸리는 것만, y축 벽(측벽·횡격벽)은 x-z 평면에
        # 그대로 보이므로 전부 그린다
        for h in (e.holes if self.on(SHOW_HOLE) else []):
            if h['axis'] == 'y':
                if h['circle']:
                    ax.add_patch(Circle((h['u'], h['v']), h['hu'],
                                        fc='#FDEDEC', ec=C_HOLE, lw=1.0, ls='--'))
                else:
                    ax.add_patch(Rectangle((h['u'] - h['hu'], h['v'] - h['hv']),
                                           2 * h['hu'], 2 * h['hv'],
                                           fc='#FDEDEC', ec=C_HOLE, lw=1.0,
                                           ls='--'))
            elif abs(h['u'] - ycut) <= h['hu']:
                zz = h['hv'] if not h['circle'] else \
                    math.sqrt(max(0.0, h['hu'] ** 2 - (h['u'] - ycut) ** 2))
                ax.add_patch(Rectangle((h['w0'], h['v'] - zz), h['t'], 2 * zz,
                                       fc='#FDEDEC', ec=C_HOLE, lw=1.0))
        # 속채움 - 셀별 채움 상단까지. 헌치보다 먼저 그려야 헌치가 가려지지 않는다
        for c in (e.cells if self.on(SHOW_FILL) else []):
            if not (c['y0'] <= ycut <= c['y1']) or c['fill_top'] <= sp.tb + EPS:
                continue
            ax.add_patch(Rectangle((c['x0'], sp.tb), c['x1'] - c['x0'],
                                   c['fill_top'] - sp.tb, fc='#F7DC6F',
                                   ec='#B7950B', lw=0.6, alpha=0.45))
            # 채움 상단선 바로 아래(채움 안쪽)에 적어 상치·정단 표기와 겹치지 않게 한다
            ax.text((c['x0'] + c['x1']) / 2, c['fill_top'],
                    f"채움 z={c['fill_top']:.2f}", ha='center', va='top',
                    fontsize=7, color='#7D6608')
        # 유수실 덮개 (절단면이 지나는 셀)
        if self.on(SHOW_BODY):
            for c in e.cells:
                if not (c['y0'] <= ycut <= c['y1']) or c['cover'] <= EPS:
                    continue
                ax.add_patch(Rectangle((c['x0'], sp.H - c['cover']),
                                       c['x1'] - c['x0'], c['cover'],
                                       fc=C_WALL, ec='#2C3E50', lw=1.0))
                ax.text((c['x0'] + c['x1']) / 2, sp.H - c['cover'],
                        f"덮개 t={c['cover']:.2f}", ha='center', va='top',
                        fontsize=7, color='#1B4F72')
        # 헌치 (절단면이 지나는 셀)
        if sp.hn_on and self.on(SHOW_HAUNCH):
            for c in e.cells:
                if not (c['y0'] <= ycut <= c['y1']):
                    continue
                for x, s in ((c['x0'], +1), (c['x1'], -1)):
                    ax.add_patch(MplPolygon(
                        [(x, sp.tb), (x + s * sp.hn_a, sp.tb),
                         (x, sp.tb + sp.hn_b)],
                        fc=C_HAUNCH, ec='#4A235A', lw=0.7))
        # 상치 - 전면 상단 경사를 잘라낸 다각형으로 그린다
        if sp.cp_on and self.on(SHOW_COPING):
            z0 = sp.cp_z0 if sp.cp_z0 > EPS else sp.H
            ax.add_patch(Rectangle((sp.cp_x0, z0), sp.cp_w, sp.cp_h, fc=C_COP,
                                   ec='#2C3E50', lw=1.2))
            fx, tz, lw_, lh_ = sp.cp_x0, z0 + sp.cp_h, sp.cp_w, sp.cp_h
            if sp.cp_par_on:
                px = sp.cp_x0 if sp.cp_par_side == SIDE_FRONT \
                    else sp.cp_x0 + sp.cp_w - sp.cp_par_w
                ax.add_patch(Rectangle((px, z0 + sp.cp_h), sp.cp_par_w,
                                       sp.cp_par_h, fc=C_COP, ec='#2C3E50', lw=1.2))
                if sp.cp_par_side == SIDE_FRONT:
                    fx, tz = px, z0 + sp.cp_h + sp.cp_par_h
            if sp.cp_sl_a > EPS and sp.cp_sl_b > EPS:
                ax.add_patch(MplPolygon(
                    [(fx, tz), (fx + sp.cp_sl_a, tz), (fx, tz - sp.cp_sl_b)],
                    fc='white', ec='#2C3E50', lw=1.0))

        gx, gy, gz, glabel = self._cg()
        ax.plot([gx], [gz], marker='X', ms=12, color=C_CG, zorder=12)
        ax.axvline(gx, ls=':', c=C_CG, lw=0.9)
        ax.axhline(gz, ls=':', c=C_CG, lw=0.9)
        ax.axvline(sp.B / 2, ls='-.', c='#2C3E50', lw=0.8)
        zoff = sp.H * 0.05
        ax.annotate('', xy=(gx, gz - zoff), xytext=(sp.B / 2, gz - zoff),
                    arrowprops=dict(arrowstyle='<->', color=C_CG, lw=1.0))
        ax.text((gx + sp.B / 2) / 2, gz - zoff, f"e = {gx - sp.B / 2:+.3f}",
                ha='center', va='top', fontsize=8, color=C_CG,
                bbox=dict(fc='white', ec='none', alpha=0.75, pad=1))
        ax.text(gx, gz, f"  G ({gx:.2f}, {gz:.2f})", color=C_CG,
                fontsize=8.5, weight='bold', va='bottom')

        z_lo = min(0.0, -max([k.d for k in sp.keys], default=0.0))
        z_hi = sp.H + ((sp.cp_h + (sp.cp_par_h if sp.cp_par_on else 0))
                       if sp.cp_on else 0)
        pad = sp.B * 0.10
        ax.set_xlim(-sp.ft_front - pad, sp.B + sp.ft_rear + pad)
        ax.set_ylim(z_lo - pad, z_hi + pad)
        ax.set_aspect('equal')
        ax.set_xlabel('x - 폭 (전면 0 → 후면 B).  일점쇄선 = 폭 중앙 B/2')
        ax.set_ylabel('z - 높이 (저판 저면 = 0)')
        ax.set_title(f"측면도 - y = {ycut:.3f} m 절단,  {glabel} G"
                     + self._hidden_note(), fontsize=10)
        ax.grid(alpha=0.25, ls=':')


# ==========================================
# 4-B. CAD 출력 (DXF)
#
#   외부 라이브러리 없이 DXF R12(AC1009) ASCII 를 직접 쓴다. R12 는 구조가
#   단순하고 AutoCAD·BricsCAD·ZWCAD·DraftSight·QCAD·LibreCAD 가 모두 읽는다.
#   폴리라인 대신 LINE 을 여러 개 쓰므로 호환성 문제가 생길 여지가 없다.
#
#   단위는 m (1 m = 1 도면단위) — 프로그램 좌표를 그대로 옮긴다.
# ==========================================

#: 레이어 이름은 ASCII 로 둔다 (한글 레이어명은 CAD·버전에 따라 깨진다)
DXF_LAYERS = (
    # (이름, ACI 색, 선종류, 설명)
    ('OUTLINE',      7, 'CONTINUOUS', '외곽선'),
    ('CAISSON-BASE', 8, 'CONTINUOUS', '저판'),
    ('CAISSON-WALL', 5, 'CONTINUOUS', '전·후·측벽'),
    ('CAISSON-PART', 3, 'CONTINUOUS', '종·횡격벽'),
    ('HAUNCH',       6, 'CONTINUOUS', '수평·수직 헌치'),
    ('FILL',         2, 'DASHED',     '속채움'),
    ('COVER',        4, 'CONTINUOUS', '유수실 덮개'),
    ('COPING',      30, 'CONTINUOUS', '상치·파라펫'),
    ('FOOTING',      9, 'DASHED',     '풋팅'),
    ('SHEARKEY',     8, 'CONTINUOUS', '전단키'),
    ('HOLE',         1, 'CONTINUOUS', '유공'),
    ('CG',           1, 'CONTINUOUS', '무게중심'),
    ('CENTERLINE',   4, 'DASHED',     '중심선'),
    ('TEXT',         7, 'CONTINUOUS', '문자'),
)


class DxfWriter:
    """DXF R12(AC1009) ASCII 작성기. LINE · CIRCLE · TEXT 만 쓴다.

    한글 문자는 CP949 로 기록하고 헤더에 $DWGCODEPAGE = ANSI_949 를 적는다.
    국내 CAD 는 이 조합을 그대로 읽는다.
    """

    def __init__(self):
        self.ents = []
        self.layers = list(DXF_LAYERS)
        self._lim = [None, None, None, None]     # xmin, ymin, xmax, ymax
        self._zlim = [None, None]                # zmin, zmax (3D 모델용)

    # ---------- 내부 ----------
    def _g(self, code, value):
        self.ents.append(f"{code}\n{value}")

    def _touch(self, x, y):
        lo_x, lo_y, hi_x, hi_y = self._lim
        self._lim = [x if lo_x is None else min(lo_x, x),
                     y if lo_y is None else min(lo_y, y),
                     x if hi_x is None else max(hi_x, x),
                     y if hi_y is None else max(hi_y, y)]

    # ---------- 엔티티 ----------
    def line(self, layer, p1, p2):
        self._g(0, 'LINE'); self._g(8, layer)
        self._g(10, f"{p1[0]:.6f}"); self._g(20, f"{p1[1]:.6f}"); self._g(30, '0.0')
        self._g(11, f"{p2[0]:.6f}"); self._g(21, f"{p2[1]:.6f}"); self._g(31, '0.0')
        self._touch(*p1); self._touch(*p2)

    def poly(self, layer, pts, close=True):
        """점 목록을 LINE 여러 개로 그린다 (사각형·삼각형·다각형 공용)."""
        n = len(pts)
        if n < 2:
            return
        for i in range(n - 1):
            self.line(layer, pts[i], pts[i + 1])
        if close:
            self.line(layer, pts[-1], pts[0])

    def rect(self, layer, x, y, dx, dy):
        self.poly(layer, [(x, y), (x + dx, y), (x + dx, y + dy), (x, y + dy)])

    def circle(self, layer, c, r):
        if r <= 0:
            return
        self._g(0, 'CIRCLE'); self._g(8, layer)
        self._g(10, f"{c[0]:.6f}"); self._g(20, f"{c[1]:.6f}"); self._g(30, '0.0')
        self._g(40, f"{r:.6f}")
        self._touch(c[0] - r, c[1] - r); self._touch(c[0] + r, c[1] + r)

    def cross(self, layer, c, r):
        """무게중심 표시용 십자."""
        self.line(layer, (c[0] - r, c[1]), (c[0] + r, c[1]))
        self.line(layer, (c[0], c[1] - r), (c[0], c[1] + r))
        self.circle(layer, c, r * 0.55)

    def polyface(self, layer, mesh):
        """닫힌 메시 하나를 POLYFACE MESH 로 쓴다 (부재 하나 = CAD 객체 하나).

        POLYLINE(70=64) + 정점 VERTEX(70=192) + 면 VERTEX(70=128) + SEQEND.
        면의 정점 번호(71~74)는 1부터 센다.
        """
        if not mesh.faces:
            return
        self._g(0, 'POLYLINE'); self._g(8, layer)
        self._g(66, 1)
        self._g(10, '0.0'); self._g(20, '0.0'); self._g(30, '0.0')
        self._g(70, 64)
        self._g(71, len(mesh.verts)); self._g(72, len(mesh.faces))
        for x, y, z in mesh.verts:
            self._g(0, 'VERTEX'); self._g(8, layer)
            self._g(10, f"{x:.6f}"); self._g(20, f"{y:.6f}"); self._g(30, f"{z:.6f}")
            self._g(70, 192)
            self._touch(x, y)
            lo, hi = self._zlim
            self._zlim = [z if lo is None else min(lo, z),
                          z if hi is None else max(hi, z)]
        for f in mesh.faces:
            self._g(0, 'VERTEX'); self._g(8, layer)
            self._g(10, '0.0'); self._g(20, '0.0'); self._g(30, '0.0')
            self._g(70, 128)
            for k, code in zip(f, (71, 72, 73, 74)):
                self._g(code, k + 1)
        self._g(0, 'SEQEND'); self._g(8, layer)

    #: TEXT 의 72(수평 정렬) 코드
    _HALIGN = {'left': 0, 'center': 1, 'right': 2}

    def text(self, layer, p, h, s, align='left', valign=0):
        """문자. 정렬을 쓰면 그룹 11/21 정렬점을 반드시 함께 적어야 한다."""
        s = str(s).replace('\n', ' ')
        self._g(0, 'TEXT'); self._g(8, layer)
        self._g(10, f"{p[0]:.6f}"); self._g(20, f"{p[1]:.6f}"); self._g(30, '0.0')
        self._g(40, f"{h:.6f}")
        self._g(1, s)
        ha = self._HALIGN.get(align, 0)
        if ha or valign:
            self._g(72, ha); self._g(73, valign)
            self._g(11, f"{p[0]:.6f}"); self._g(21, f"{p[1]:.6f}"); self._g(31, '0.0')
        self._touch(p[0], p[1])
        self._touch(p[0] + h * _dw(s) * 0.5, p[1] + h)

    # ---------- 저장 ----------
    def _header(self):
        lo_x, lo_y, hi_x, hi_y = [v if v is not None else 0.0 for v in self._lim]
        lo_z, hi_z = [v if v is not None else 0.0 for v in self._zlim]
        out = ['0', 'SECTION', '2', 'HEADER',
               '9', '$ACADVER', '1', 'AC1009',
               '9', '$DWGCODEPAGE', '3', 'ANSI_949',
               '9', '$EXTMIN', '10', f"{lo_x:.6f}", '20', f"{lo_y:.6f}",
               '30', f"{lo_z:.6f}",
               '9', '$EXTMAX', '10', f"{hi_x:.6f}", '20', f"{hi_y:.6f}",
               '30', f"{hi_z:.6f}",
               '9', '$LTSCALE', '40', '1.0',
               '0', 'ENDSEC']
        return out

    def _tables(self):
        out = ['0', 'SECTION', '2', 'TABLES']
        # 선종류 — 파선 하나만 정의한다
        out += ['0', 'TABLE', '2', 'LTYPE', '70', '2',
                '0', 'LTYPE', '2', 'CONTINUOUS', '70', '0',
                '3', 'Solid line', '72', '65', '73', '0', '40', '0.0',
                '0', 'LTYPE', '2', 'DASHED', '70', '0',
                '3', '__ __ __ __', '72', '65', '73', '2', '40', '0.75',
                '49', '0.5', '49', '-0.25',
                '0', 'ENDTAB']
        out += ['0', 'TABLE', '2', 'LAYER', '70', str(len(self.layers) + 1),
                '0', 'LAYER', '2', '0', '70', '0', '62', '7', '6', 'CONTINUOUS']
        for name, aci, lt, _desc in self.layers:
            out += ['0', 'LAYER', '2', name, '70', '0',
                    '62', str(aci), '6', lt]
        out += ['0', 'ENDTAB', '0', 'ENDSEC']
        return out

    def save(self, path):
        body = self._header() + self._tables()
        body += ['0', 'SECTION', '2', 'ENTITIES']
        body += '\n'.join(self.ents).split('\n') if self.ents else []
        body += ['0', 'ENDSEC', '0', 'EOF']
        txt = '\n'.join(body) + '\n'
        # 국내 CAD 표준 조합 : CP949 본문 + $DWGCODEPAGE ANSI_949
        with open(path, 'w', encoding='cp949', errors='replace', newline='\r\n') as f:
            f.write(txt)


class DxfExporter:
    """엔진 결과를 평면도·정면도·측면도 3도면이 담긴 DXF 한 장으로 내보낸다.

    화면 도해와 같은 엔진 값을 쓰므로 도면과 화면이 어긋날 일이 없다.
    표시 설정과 무관하게 전 부재를 담고 레이어로 나눈다 — CAD 에서 레이어를
    켜고 끄는 편이 다시 저장하는 것보다 빠르기 때문이다.
    """

    def __init__(self, eng: CaissonEngine):
        self.e = eng
        self.sp = eng.sp
        s = max(eng.sp.B, eng.sp.L)
        self.th = s / 60.0          # 일반 문자 높이
        self.tt = s / 25.0          # 제목 문자 높이
        self.gap = s * 0.35         # 도면 사이 간격

    # ------------------------------------------------------------------
    def _cg(self):
        cb = self.e.combined()
        if cb:
            return cb['gx'], cb['gy'], cb['gz']
        _V, x, y, z = self.e.group_props(GRP_CONC)
        return x, y, z

    def _z_top(self):
        sp = self.sp
        return sp.H + ((sp.cp_h + (sp.cp_par_h if sp.cp_par_on else 0))
                       if sp.cp_on else 0.0)

    def _title(self, d, ox, oy, w, text):
        """도면 아래에 제목을 넣고 밑줄을 긋는다."""
        y = oy - self.tt * 2.2
        d.text('TEXT', (ox + w / 2.0, y), self.tt, text, align='center')
        d.line('TEXT', (ox + w / 2.0 - w * 0.32, y - self.tt * 0.45),
               (ox + w / 2.0 + w * 0.32, y - self.tt * 0.45))

    # ------------------------------------------------------------------
    def _plan(self, d, ox, oy):
        """평면도 (x-y). 원점 이동만 하고 좌표계는 그대로."""
        e, sp = self.e, self.sp

        def P(x, y):
            return (ox + x, oy + y)

        if max(sp.ft_front, sp.ft_rear, sp.ft_side) > EPS:
            d.poly('FOOTING', [P(-sp.ft_front, -sp.ft_side),
                               P(sp.B + sp.ft_rear, -sp.ft_side),
                               P(sp.B + sp.ft_rear, sp.L + sp.ft_side),
                               P(-sp.ft_front, sp.L + sp.ft_side)])
        d.poly('OUTLINE', [P(0, 0), P(sp.B, 0), P(sp.B, sp.L), P(0, sp.L)])
        # 외벽 내면
        d.poly('CAISSON-WALL', [P(e.x_in0, e.y_in0), P(e.x_in1, e.y_in0),
                                P(e.x_in1, e.y_in1), P(e.x_in0, e.y_in1)])
        for i, dd in enumerate(e.lps):
            d.rect('CAISSON-PART', *P(dd['lo'], e.y_in0),
                   dd['t'], e.y_in1 - e.y_in0)
            d.text('TEXT', P(dd['pos'], e.y_in1 + self.th * 0.4), self.th,
                   f"L#{i + 1}", align='center')
        for j, dd in enumerate(e.tps):
            d.rect('CAISSON-PART', *P(e.x_in0, dd['lo']),
                   e.x_in1 - e.x_in0, dd['t'])
            d.text('TEXT', P(e.x_in1 + self.th * 0.4, dd['pos']), self.th,
                   f"T#{j + 1}")
        for c in e.cells:
            cx, cy = (c['x0'] + c['x1']) / 2.0, (c['y0'] + c['y1']) / 2.0
            d.text('TEXT', P(cx, cy + self.th * 0.6), self.th,
                   c['name'], align='center')
            d.text('TEXT', P(cx, cy - self.th * 0.6), self.th,
                   c['ctype'], align='center')
            if c['fill_top'] > sp.tb + EPS:
                d.text('FILL', P(cx, cy - self.th * 2.0), self.th,
                       f"채움 z={c['fill_top']:.3f}", align='center')
            if c['cover'] > EPS:
                d.rect('COVER', *P(c['x0'], c['y0']),
                       c['x1'] - c['x0'], c['y1'] - c['y0'])
                d.text('COVER', P(cx, cy - self.th * 3.2), self.th,
                       f"덮개 t={c['cover']:.3f}", align='center')
        # 수직 헌치
        if sp.vh_on:
            for c in e.cells:
                for _tag, x, y, sx, sy, hcor in e._cell_corners(c):
                    if sp.tb + hcor - e._vh_z0() <= EPS:
                        continue
                    d.poly('HAUNCH', [P(x, y), P(x + sx * sp.vh_c, y),
                                      P(x, y + sy * sp.vh_c)])
        for h in e.holes:                       # 축과 무관하게 바운딩박스로
            bx0, bx1, by0, by1, _z0, _z1 = h['prim'].bbox()
            d.rect('HOLE', *P(bx0, by0), bx1 - bx0, by1 - by0)
        gx, gy, _gz = self._cg()
        d.cross('CG', P(gx, gy), self.th * 1.2)
        d.text('CG', P(gx + self.th, gy + self.th), self.th,
               f"G ({gx:.3f}, {gy:.3f})")
        w = sp.B + sp.ft_front + sp.ft_rear
        self._title(d, ox - sp.ft_front, oy - sp.ft_side, w,
                    f"평 면 도   (셀 {len(e.cells)} 개)")
        return w

    # ------------------------------------------------------------------
    def _front(self, d, ox, oy):
        """정면도 (y-z). 도면 x = 케이슨 y, 도면 y = z."""
        e, sp = self.e, self.sp

        def P(y, z):
            return (ox + y, oy + z)

        d.rect('CAISSON-BASE', *P(0, 0), sp.L, sp.tb)
        d.rect('CAISSON-WALL', *P(0, sp.tb), sp.L, e.hw)
        for j, dd in enumerate(e.tps):
            d.rect('CAISSON-PART', *P(dd['lo'], sp.tb), dd['t'], dd['h'])
            d.text('TEXT', P(dd['pos'], sp.tb + dd['h'] + self.th * 0.3),
                   self.th, f"T#{j + 1}", align='center')
        if sp.cp_on:
            z0 = sp.cp_z0 if sp.cp_z0 > EPS else sp.H
            d.rect('COPING', *P(0, z0), sp.L, sp.cp_h)
            if sp.cp_par_on:
                d.rect('COPING', *P(0, z0 + sp.cp_h), sp.L, sp.cp_par_h)
        # 속채움 상단 (벽 뒤라 보이지 않으므로 파선으로 알린다)
        seen = set()
        for c in e.cells:
            key = (round(c['y0'], 6), round(c['y1'], 6), round(c['fill_top'], 6))
            if c['fill_top'] <= sp.tb + EPS or key in seen:
                continue
            seen.add(key)
            d.line('FILL', P(c['y0'], c['fill_top']), P(c['y1'], c['fill_top']))
            d.text('FILL', P((c['y0'] + c['y1']) / 2.0,
                             c['fill_top'] - self.th * 1.3), self.th,
                   f"채움 z={c['fill_top']:.3f}", align='center')
        # 정면도는 y-z 평면이라 x축 관통 벽의 유공만 보인다
        nx = 0
        for h in e.holes:
            if h['axis'] != 'x':
                continue
            nx += 1
            if h['circle']:
                d.circle('HOLE', P(h['u'], h['v']), h['hu'])
            else:
                d.rect('HOLE', *P(h['u'] - h['hu'], h['v'] - h['hv']),
                       2 * h['hu'], 2 * h['hv'])
            d.text('HOLE', P(h['u'], h['v']), self.th * 0.8, str(h['i']),
                   align='center')
        r, _a, _f = e.opening_ratio()
        wl = ' · '.join(g['name'] for g in getattr(e, 'wall_geoms', [])
                        if g['axis'] == 'x')
        ttl = (f"정 면 도   ({wl})   유공 {nx} 개,  개구율 {r * 100:.2f} %"
               if nx else "정 면 도   (x축 관통 유공 없음)")
        self._title(d, ox, oy, sp.L, ttl)
        return sp.L

    # ------------------------------------------------------------------
    def _side(self, d, ox, oy):
        """측면도 (x-z). 첫 셀의 y 중앙을 지나는 횡단면."""
        e, sp = self.e, self.sp
        ycut = ((e.cells[0]['y0'] + e.cells[0]['y1']) / 2.0
                if e.cells else sp.L / 2.0)

        def P(x, z):
            return (ox + x, oy + z)

        if max(sp.ft_front, sp.ft_rear) > EPS:
            d.rect('FOOTING', *P(-sp.ft_front, 0),
                   sp.B + sp.ft_front + sp.ft_rear, sp.ft_t)
        d.rect('CAISSON-BASE', *P(0, 0), sp.B, sp.tb)
        for k in sorted(sp.keys, key=lambda q: q.x):
            d.rect('SHEARKEY', *P(k.x, -k.d), k.w, k.d)
        d.rect('CAISSON-WALL', *P(0, sp.tb), sp.tf, e.hw)
        d.rect('CAISSON-WALL', *P(sp.B - sp.tr, sp.tb), sp.tr, e.hw)
        for i, dd in enumerate(e.lps):
            d.rect('CAISSON-PART', *P(dd['lo'], sp.tb), dd['t'], dd['h'])
            d.text('TEXT', P(dd['pos'], sp.tb + dd['h'] + self.th * 0.3),
                   self.th, f"L#{i + 1}", align='center')
        # y축 관통 벽(측벽·횡격벽)의 유공은 x-z 평면에 그대로 보인다
        for h in e.holes:
            if h['axis'] == 'y':
                if h['circle']:
                    d.circle('HOLE', P(h['u'], h['v']), h['hu'])
                else:
                    d.rect('HOLE', *P(h['u'] - h['hu'], h['v'] - h['hv']),
                           2 * h['hu'], 2 * h['hv'])
            elif abs(h['u'] - ycut) <= h['hu']:
                zz = h['hv'] if not h['circle'] else \
                    math.sqrt(max(0.0, h['hu'] ** 2 - (h['u'] - ycut) ** 2))
                d.rect('HOLE', *P(h['w0'], h['v'] - zz), h['t'], 2 * zz)
        for c in e.cells:
            if not (c['y0'] <= ycut <= c['y1']):
                continue
            if c['cover'] > EPS:
                d.rect('COVER', *P(c['x0'], sp.H - c['cover']),
                       c['x1'] - c['x0'], c['cover'])
                d.text('COVER', P((c['x0'] + c['x1']) / 2.0,
                                  sp.H - c['cover'] - self.th * 1.2), self.th,
                       f"덮개 t={c['cover']:.3f}", align='center')
            if c['fill_top'] <= sp.tb + EPS:
                continue
            d.rect('FILL', *P(c['x0'], sp.tb),
                   c['x1'] - c['x0'], c['fill_top'] - sp.tb)
            d.text('FILL', P((c['x0'] + c['x1']) / 2.0,
                             c['fill_top'] - self.th * 1.2), self.th,
                   f"채움 z={c['fill_top']:.3f}", align='center')
        if sp.hn_on:
            for c in e.cells:
                if not (c['y0'] <= ycut <= c['y1']):
                    continue
                for x, s in ((c['x0'], +1), (c['x1'], -1)):
                    d.poly('HAUNCH', [P(x, sp.tb), P(x + s * sp.hn_a, sp.tb),
                                      P(x, sp.tb + sp.hn_b)])
        if sp.cp_on:
            z0 = sp.cp_z0 if sp.cp_z0 > EPS else sp.H
            d.rect('COPING', *P(sp.cp_x0, z0), sp.cp_w, sp.cp_h)
            fx, tz = sp.cp_x0, z0 + sp.cp_h
            if sp.cp_par_on:
                px = sp.cp_x0 if sp.cp_par_side == SIDE_FRONT \
                    else sp.cp_x0 + sp.cp_w - sp.cp_par_w
                d.rect('COPING', *P(px, z0 + sp.cp_h), sp.cp_par_w, sp.cp_par_h)
                if sp.cp_par_side == SIDE_FRONT:
                    fx, tz = px, z0 + sp.cp_h + sp.cp_par_h
            if sp.cp_sl_a > EPS and sp.cp_sl_b > EPS:
                # 경사로 잘려 나간 자리를 실제 마감선으로 그린다
                d.line('COPING', P(fx + sp.cp_sl_a, tz), P(fx, tz - sp.cp_sl_b))
        gx, _gy, gz = self._cg()
        z_lo = min(0.0, -max([k.d for k in sp.keys], default=0.0))
        z_hi = self._z_top()
        d.line('CENTERLINE', P(sp.B / 2.0, z_lo - self.th),
               P(sp.B / 2.0, z_hi + self.th))
        d.text('CENTERLINE', P(sp.B / 2.0, z_hi + self.th * 1.4), self.th,
               'B/2', align='center')
        d.cross('CG', P(gx, gz), self.th * 1.2)
        d.text('CG', P(gx + self.th, gz + self.th), self.th,
               f"G ({gx:.3f}, {gz:.3f})")
        d.line('CG', P(sp.B / 2.0, gz), P(gx, gz))
        d.text('CG', P((gx + sp.B / 2.0) / 2.0, gz - self.th * 1.3), self.th,
               f"e = {gx - sp.B / 2.0:+.3f}", align='center')
        w = sp.B + sp.ft_front + sp.ft_rear
        self._title(d, ox - sp.ft_front, oy + z_lo, w,
                    f"측 면 도   (y = {ycut:.3f} m 절단)")
        return w

    # ------------------------------------------------------------------
    def save(self, path):
        sp = self.sp
        d = DxfWriter()
        ox = 0.0
        ox += self._plan(d, ox + sp.ft_front, 0.0) + self.gap
        ox += self._front(d, ox, 0.0) + self.gap
        self._side(d, ox + sp.ft_front, 0.0)
        # 표제 - 세 도면 위쪽에 한 줄
        lo_x, _lo_y, hi_x, hi_y = [v if v is not None else 0.0 for v in d._lim]
        d.text('TEXT', ((lo_x + hi_x) / 2.0, hi_y + self.tt * 2.0),
               self.tt * 1.3,
               f"{sp.project}   {sp.name}   (단위 : m)", align='center')
        d.save(path)
        return [nm for nm, _c, _lt, _ds in DXF_LAYERS]


# ==========================================
# 4-C. 3D 모델 출력 (DXF POLYFACE MESH)
#
#   엔진이 들고 있는 부호 있는 기본도형 분해를 그대로 닫힌 면 메시로 옮긴다.
#   겉모양에 드러나는 공제(유공·상치 전면경사)는 실제로 뚫고, 속에서만 겹치는
#   공제(격벽 교차부·풋팅 본체·속채움 헌치)는 겉면이 같으므로 생략한다.
#
#   ※ 3D 모델은 형상 확인용이다. 체적은 계산서 값을 써야 한다.
# ==========================================

CIRCLE_SEG = 36          # 원형 근사 분할수 (10°)


class Mesh:
    """닫힌 다면체 하나. 정점 목록 + 면(정점 인덱스) 목록."""

    def __init__(self, name='', layer=''):
        self.name = name
        self.layer = layer
        self.verts = []
        self.faces = []
        self._idx = {}       # 반올림 좌표 → 정점 번호 (중복 정점 제거)

    # ---------- 조립 ----------
    def add_vert(self, p):
        key = (round(p[0], 9), round(p[1], 9), round(p[2], 9))
        i = self._idx.get(key)
        if i is None:
            i = len(self.verts)
            self.verts.append((float(p[0]), float(p[1]), float(p[2])))
            self._idx[key] = i
        return i

    def add_face(self, *pts):
        """좌표 3~4개로 면 하나. 중복 정점은 지우고 3개 미만이면 버린다."""
        idx = []
        for p in pts:
            i = self.add_vert(p)
            if not idx or idx[-1] != i:
                idx.append(i)
        if len(idx) > 2 and idx[0] == idx[-1]:
            idx.pop()
        if len(idx) >= 3:
            self.faces.append(tuple(idx[:4]))

    def extend(self, other):
        for f in other.faces:
            self.add_face(*[other.verts[i] for i in f])

    # ---------- 검증 ----------
    def _tris(self):
        for f in self.faces:
            yield f[0], f[1], f[2]
            if len(f) == 4:
                yield f[0], f[2], f[3]

    def volume(self):
        """발산정리 부호 체적. 면이 바깥을 보면 양수다."""
        v6 = 0.0
        for a, b, c in self._tris():
            (ax, ay, az), (bx, by, bz), (cx, cy, cz) = \
                self.verts[a], self.verts[b], self.verts[c]
            v6 += (ax * (by * cz - bz * cy)
                   - ay * (bx * cz - bz * cx)
                   + az * (bx * cy - by * cx))
        return v6 / 6.0

    def orient_outward(self):
        """부호 체적이 음수면 전 면을 뒤집어 바깥 방향으로 맞춘다."""
        if self.volume() < 0:
            self.faces = [tuple(reversed(f)) for f in self.faces]
        return self

    def drop_internal(self):
        """정확히 겹친 면 쌍(안쪽 막)을 지운다.

        유공 두 개가 딱 맞닿게 배치되면 두 배럴 면이 같은 자리에 반대 방향으로
        생긴다. 체적에는 영향이 없지만 다양체가 아니게 되므로 걷어낸다.
        """
        seen, drop = {}, set()
        for i, f in enumerate(self.faces):
            key = tuple(sorted(f))
            j = seen.pop(key, None)
            if j is None:
                seen[key] = i
            else:
                drop.add(i)
                drop.add(j)
        if drop:
            self.faces = [f for i, f in enumerate(self.faces) if i not in drop]
        return self

    def open_edges(self):
        """닫히지 않은 모서리 목록. 정상이면 빈 목록이다.

        모든 모서리는 정확히 두 면에 서로 반대 방향으로 쓰여야 한다.
        """
        cnt = {}
        for f in self.faces:
            n = len(f)
            for i in range(n):
                a, b = f[i], f[(i + 1) % n]
                cnt[(a, b)] = cnt.get((a, b), 0) + 1
        bad = []
        for (a, b), c in cnt.items():
            if c != 1 or cnt.get((b, a), 0) != 1:
                bad.append((a, b, c, cnt.get((b, a), 0)))
        return bad


# ------------------------------------------------------------------
# 기본도형 → 메시
# ------------------------------------------------------------------
def mesh_box(b: Box, name='', layer=''):
    m = Mesh(name, layer)
    x0, y0, z0 = b.x0, b.y0, b.z0
    x1, y1, z1 = x0 + b.dx, y0 + b.dy, z0 + b.dz
    P = [(x0, y0, z0), (x1, y0, z0), (x1, y1, z0), (x0, y1, z0),
         (x0, y0, z1), (x1, y0, z1), (x1, y1, z1), (x0, y1, z1)]
    for a, bb, c, d in ((0, 3, 2, 1), (4, 5, 6, 7), (0, 1, 5, 4),
                        (1, 2, 6, 5), (2, 3, 7, 6), (3, 0, 4, 7)):
        m.add_face(P[a], P[bb], P[c], P[d])
    return m.orient_outward()


def mesh_triprism(t: TriPrism, name='', layer=''):
    """직각삼각형 단면 압출체. du·dv 부호에 따라 방향이 뒤집히므로 마지막에 정리."""
    m = Mesh(name, layer)
    w0, w1 = t.w0, t.w0 + t.h
    uv = [(t.u0, t.v0), (t.u0 + t.du, t.v0), (t.u0, t.v0 + t.dv)]
    A = [_uvw_to_xyz(t.axis, u, v, w0) for u, v in uv]
    B = [_uvw_to_xyz(t.axis, u, v, w1) for u, v in uv]
    m.add_face(A[0], A[2], A[1])
    m.add_face(B[0], B[1], B[2])
    for i in range(3):
        j = (i + 1) % 3
        m.add_face(A[i], A[j], B[j], B[i])
    return m.orient_outward()


def mesh_cyl(c: Cyl, name='', layer='', n=CIRCLE_SEG):
    """원기둥을 n각형 기둥으로 근사."""
    m = Mesh(name, layer)
    w0, w1 = c.w0, c.w0 + c.h
    ring = [(c.u + c.r * math.cos(2 * math.pi * k / n),
             c.v + c.r * math.sin(2 * math.pi * k / n)) for k in range(n)]
    A = [_uvw_to_xyz(c.axis, u, v, w0) for u, v in ring]
    B = [_uvw_to_xyz(c.axis, u, v, w1) for u, v in ring]
    for k in range(1, n - 1):
        m.add_face(A[0], A[k + 1], A[k])
        m.add_face(B[0], B[k], B[k + 1])
    for k in range(n):
        j = (k + 1) % n
        m.add_face(A[k], A[j], B[j], B[k])
    return m.orient_outward()


def ngon_volume(r, h, n=CIRCLE_SEG):
    """n각형 기둥의 해석 체적 — 원기둥 근사 오차를 검증할 때 쓴다."""
    return 0.5 * n * r * r * math.sin(2 * math.pi / n) * h


def hole_polygon(h, n=CIRCLE_SEG):
    """유공 하나의 (u, v) 단면 다각형. 원형은 n각형, 사각형은 4점. 반시계 방향."""
    if h['circle']:
        r = h['hu']
        return [(h['u'] + r * math.cos(2 * math.pi * k / n),
                 h['v'] + r * math.sin(2 * math.pi * k / n)) for k in range(n)]
    hu, hv = h['hu'], h['hv']
    return [(h['u'] - hu, h['v'] - hv), (h['u'] + hu, h['v'] - hv),
            (h['u'] + hu, h['v'] + hv), (h['u'] - hu, h['v'] + hv)]


def _poly_span_at_z(poly, z):
    """볼록 다각형과 수평선 z 의 교차 구간 (yL, yR). 만나지 않으면 None."""
    ys = []
    n = len(poly)
    for i in range(n):
        (y1, z1), (y2, z2) = poly[i], poly[(i + 1) % n]
        if abs(z2 - z1) < EPS:
            if abs(z1 - z) < EPS:
                ys += [y1, y2]
            continue
        s = (z - z1) / (z2 - z1)
        if -EPS <= s <= 1.0 + EPS:
            ys.append(y1 + s * (y2 - y1))
    return (min(ys), max(ys)) if ys else None


def _wall_breakpoints(polys, z, Y0, Y1):
    """높이 z 에서 벽면을 끊는 y 값. 그 높이에 '닿기만' 하는 구멍도 포함한다.

    띠 위아래가 같은 끊는 점을 쓰게 만들어야 면이 T 자로 어긋나지 않는다.
    """
    pts = {round(Y0, 9), round(Y1, 9)}
    for poly in polys:
        pz = [p[1] for p in poly]
        if z < min(pz) - EPS or z > max(pz) + EPS:
            continue
        s = _poly_span_at_z(poly, z)
        if s:
            pts.add(round(s[0], 9))
            pts.add(round(s[1], 9))
    return sorted(pts)


def _zip_strip(m, bot, top, za, zb, xf, xb, axis='x'):
    """아래 사슬과 위 사슬 사이를 삼각형으로 메운다.

    정점을 새로 만들지 않고 두 사슬의 점만 쓰므로, 이웃한 띠·배럴과 모서리가
    정확히 맞물린다. bot·top 은 왼쪽→오른쪽으로 정렬된 u 값 목록이다.
    """
    def W(w, u, v):
        return _uvw_to_xyz(axis, u, v, w)

    bot = [y for i, y in enumerate(bot) if i == 0 or abs(y - bot[i - 1]) > EPS]
    top = [y for i, y in enumerate(top) if i == 0 or abs(y - top[i - 1]) > EPS]
    if len(bot) < 2 and len(top) < 2:
        return                                   # 폭이 0 인 조각

    def par(ch):
        w = ch[-1] - ch[0]
        if len(ch) < 2 or abs(w) < EPS:
            return [0.0] * len(ch)
        return [(y - ch[0]) / w for y in ch]

    pb, pt = par(bot), par(top)
    i = j = 0
    while i < len(bot) - 1 or j < len(top) - 1:
        if j >= len(top) - 1 or (i < len(bot) - 1 and pb[i + 1] <= pt[j + 1]):
            A, B, C = (bot[i], za), (bot[i + 1], za), (top[j], zb)
            i += 1
        else:
            A, B, C = (bot[i], za), (top[j + 1], zb), (top[j], zb)
            j += 1
        # A→B→C 는 u 우측·v 상방 기준 반시계 = +w 방향
        m.add_face(W(xf, *A), W(xf, *C), W(xf, *B))
        m.add_face(W(xb, *A), W(xb, *B), W(xb, *C))


def mesh_perforated_wall(x0, t, Y0, Y1, Z0, Z1, polys, name='', layer='',
                         axis='x'):
    """유공을 실제로 뚫은 벽 하나의 닫힌 메시.

    벽의 국부좌표 (u, v, w) 로 조립하고 마지막에 축으로 옮긴다.
      axis='x' : w=x, (u,v)=(y,z)      전면·후면·종격벽
      axis='y' : w=y, (u,v)=(x,z)      좌측·우측·횡격벽
    인수 이름의 x0/Y/Z 는 각각 w0 / u범위 / v범위 를 뜻한다.

    구멍 뚫린 앞·뒷면은 일반 다각형 삼각분할기 없이 **v 방향 띠로 잘라** 채운다.
    띠 경계를 구멍 다각형의 정점 v 값에 두면 띠 안에서 구멍 경계가 v 에 대해
    선형이므로, 구멍 사이에 남는 벽은 두 직선 사슬 사이의 띠가 된다. 그 사이를
    정점을 더하지 않고 삼각형으로 메우면 배럴·이웃 띠와 모서리가 맞물린다.
    """
    m = Mesh(name, layer)
    x1 = x0 + t

    def W(w, u, v):
        return _uvw_to_xyz(axis, u, v, w)

    # --- 띠 경계 : 벽 상·하단 + 모든 구멍 정점의 z ---
    #   값을 반올림하면 안 된다. 띠 경계가 구멍 정점의 z 와 정확히 같아야
    #   _poly_span_at_z 가 그 정점의 y 를 그대로 돌려주고, 그래야 면과 배럴이
    #   같은 점을 쓴다. 겹치는 값만 오차 범위로 합친다.
    cand = sorted([Z0, Z1] + [z for poly in polys for _y, z in poly
                              if Z0 - EPS < z < Z1 + EPS])
    zs = []
    for z in cand:
        if not zs or z - zs[-1] > 1e-9:
            zs.append(z)
    bp = [_wall_breakpoints(polys, z, Y0, Y1) for z in zs]

    for iz in range(len(zs) - 1):
        za, zb = zs[iz], zs[iz + 1]
        if zb - za < EPS:
            continue
        spans = []
        for poly in polys:
            pz = [p[1] for p in poly]
            if min(pz) > zb - EPS or max(pz) < za + EPS:
                continue                         # 이 띠와 겹치지 않는 구멍
            sa, sb = _poly_span_at_z(poly, za), _poly_span_at_z(poly, zb)
            if sa and sb:
                spans.append((sa[0], sa[1], sb[0], sb[1]))
        spans.sort(key=lambda s: s[0])
        pa, pb = Y0, Y0                          # 직전 구멍의 오른쪽 끝
        regions = []
        for La, Ra, Lb, Rb in spans:
            regions.append((pa, La, pb, Lb))
            pa, pb = Ra, Rb
        regions.append((pa, Y1, pb, Y1))
        for ya0, ya1, yb0, yb1 in regions:
            bot = [ya0] + [c for c in bp[iz] if ya0 + EPS < c < ya1 - EPS] + [ya1]
            top = [yb0] + [c for c in bp[iz + 1] if yb0 + EPS < c < yb1 - EPS] + [yb1]
            _zip_strip(m, bot, top, za, zb, x0, x1, axis)

    # --- 바깥 둘레 ---
    #   좌·우면은 앞·뒷면이 띠로 잘려 있으므로 같은 띠로 나눠야 맞물린다
    for za, zb in zip(zs[:-1], zs[1:]):
        if zb - za < EPS:
            continue
        m.add_face(W(x0, Y0, za), W(x1, Y0, za), W(x1, Y0, zb), W(x0, Y0, zb))
        m.add_face(W(x0, Y1, za), W(x0, Y1, zb), W(x1, Y1, zb), W(x1, Y1, za))
    # 상·하면은 끝 띠가 통짜라 그대로 맞물린다
    m.add_face(W(x0, Y0, Z0), W(x0, Y1, Z0), W(x1, Y1, Z0), W(x1, Y0, Z0))
    m.add_face(W(x0, Y0, Z1), W(x1, Y0, Z1), W(x1, Y1, Z1), W(x0, Y1, Z1))

    # --- 구멍 안쪽 면 (배럴). 바깥 방향은 구멍 속을 향한다 ---
    for poly in polys:
        n = len(poly)
        for k in range(n):
            (y1, z1), (y2, z2) = poly[k], poly[(k + 1) % n]
            m.add_face(W(x0, y2, z2), W(x0, y1, z1),
                       W(x1, y1, z1), W(x1, y2, z2))
    return m.drop_internal().orient_outward()


def mesh_extrude_y(profile, y0, y1, name='', layer=''):
    """x-z 평면의 볼록 프로필을 y 방향으로 압출한다.

    상치·파라펫처럼 단면이 일정한 부재에 쓴다. 전면 경사(모따기)를 프로필에서
    미리 잘라 두면 공제 부재 없이 정확한 형상이 나온다.
    """
    m = Mesh(name, layer)
    n = len(profile)
    A = [(x, y0, z) for x, z in profile]
    B = [(x, y1, z) for x, z in profile]
    for k in range(1, n - 1):          # 볼록이므로 부채꼴 삼각분할로 충분하다
        m.add_face(A[0], A[k], A[k + 1])
        m.add_face(B[0], B[k + 1], B[k])
    for k in range(n):
        j = (k + 1) % n
        m.add_face(A[k], B[k], B[j], A[j])
    return m.orient_outward()


def mesh_loft(sections, name='', layer=''):
    """볼록 다각형 단면을 z 방향으로 이어 붙인 로프트.

    sections = [(z, [(x, y), ...]), ...] — 단면마다 점 개수가 같아야 한다.
    점이 겹쳐 생기는 축퇴(모따기 다리가 0 인 경우 등)는 Mesh.add_face 가 정리한다.
    헌치·속채움처럼 단면 치수가 높이에 따라 선형으로 변하는 부재에 쓴다.
    """
    m = Mesh(name, layer)
    if len(sections) < 2:
        return m
    rings = [[(x, y, z) for x, y in pts] for z, pts in sections]
    n = len(rings[0])
    for a, b in zip(rings[:-1], rings[1:]):
        for k in range(n):
            j = (k + 1) % n
            m.add_face(a[k], b[k], b[j], a[j])
    bot, top = rings[0], rings[-1]
    for k in range(1, n - 1):                # 볼록이므로 부채꼴 삼각분할
        m.add_face(bot[0], bot[k], bot[k + 1])
        m.add_face(top[0], top[k + 1], top[k])
    return m.drop_internal().orient_outward()


def mesh_haunch_wedge(cell, a, b, tb, side, name='', layer=''):
    """수평 헌치 한 쪽 - 45° 마이터로 접합된 쐐기.

    계산서는 'x면 헌치를 a 만큼 자르고 코너 Lump 로 되살리는' 분해를 쓰지만,
    3D 에서는 실제 형상인 마이터 액자로 만든다. 네 쪽을 합치면 체적이
    ab(lx+ly) - 4a²b/3 로 계산서와 정확히 같고 Lump 가 필요 없다.

    높이 w 의 단면은 마이터 평면(코너 대각선)에 잘린 사다리꼴이고 폭
    s = a(1 - w/b) 가 선형이므로 로프트로 정확히 표현된다. w = b 에서는
    선분으로 축퇴하며 Mesh.add_face 가 정리한다.

    side : 'x0' | 'x1' | 'y0' | 'y1'
    """
    x0, x1, y0, y1 = cell['x0'], cell['x1'], cell['y0'], cell['y1']

    def sect(s):
        if side == 'x0':
            return [(x0, y0), (x0 + s, y0 + s), (x0 + s, y1 - s), (x0, y1)]
        if side == 'x1':
            return [(x1, y0), (x1 - s, y0 + s), (x1 - s, y1 - s), (x1, y1)]
        if side == 'y0':
            return [(x0, y0), (x1, y0), (x1 - s, y0 + s), (x0 + s, y0 + s)]
        return [(x0, y1), (x0 + s, y1 - s), (x1 - s, y1 - s), (x1, y1)]

    return mesh_loft([(tb, sect(a)), (tb + b, sect(0.0))], name, layer)


def haunch_section(sp, w):
    """저판 상면에서 w 만큼 위에서 헌치가 차지하는 크기.

    s : 수평 헌치가 셀 둘레를 먹는 폭        s = a(1 - w/b)
    g : 수직 헌치 삼각형의 다리              g = max(0, c - 2s)
    둘 다 w 에 대해 구간별 선형이라 로프트로 정확히 표현된다.
    """
    s = (sp.hn_a * max(0.0, 1.0 - w / sp.hn_b)
         if (sp.hn_on and sp.hn_b > EPS) else 0.0)
    g = max(0.0, sp.vh_c - 2.0 * s) if sp.vh_on else 0.0
    return s, g


def haunch_knots(sp, w_top):
    """헌치 단면이 꺾이는 높이. s·g 가 구간별로만 선형이므로 여기서 끊는다."""
    ks = {0.0, w_top}
    if sp.hn_on and sp.hn_b > EPS:
        if 0.0 < sp.hn_b < w_top:
            ks.add(sp.hn_b)                    # 수평 헌치 상단
        if sp.vh_on and sp.hn_a > EPS:
            wg = sp.hn_b * (1.0 - sp.vh_c / (2.0 * sp.hn_a))
            if 0.0 < wg < min(sp.hn_b, w_top):
                ks.add(wg)                     # 수직 헌치 다리가 살아나는 곳
    return sorted(k for k in ks if 0.0 - EPS <= k <= w_top + EPS)


def mesh_vhaunch_loft(cell, corner, sp, z_top, name='', layer=''):
    """수직(우각부) 헌치 한 모서리 - 실제 형상 그대로의 로프트.

    높이 w 에서 직각점 (s, s), 다리 g 인 삼각형이다. 빗면은 u + v = c 로
    w 와 무관한 고정 평면이라 옆면이 모두 평면이 된다. 계산서의
    '프리즘 + 코너 보정 Lump' 와 체적이 정확히 같다.
    """
    _tag, X, Y, sx, sy, _h = corner
    tb = sp.tb
    raw = []
    for w in haunch_knots(sp, z_top - tb):
        s, g = haunch_section(sp, w)
        p0 = (X + sx * s, Y + sy * s)
        if g <= EPS:
            raw.append((tb + w, [p0, p0, p0], False))
        else:
            raw.append((tb + w, [p0, (X + sx * (s + g), Y + sy * s),
                                 (X + sx * s, Y + sy * (s + g))], True))
    # g 는 높이에 따라 단조 증가하므로 살아 있는 구간은 하나다.
    # 그 앞의 축퇴 단면 하나를 남겨 두어야 아래쪽 뾰족한 끝이 닫힌다.
    live = [i for i, r in enumerate(raw) if r[2]]
    if not live:
        return Mesh(name, layer)
    lo = max(0, live[0] - 1)
    secs = [(z, pts) for z, pts, _on in raw[lo:]]
    if len(secs) < 2:
        return Mesh(name, layer)
    return mesh_loft(secs, name, layer)


def mesh_fill_loft(cell, corners, sp, fill_top, name='', layer=''):
    """속채움 - 셀 공극에서 헌치 형상을 뺀 실제 형상.

    높이 w 의 단면은 안쪽으로 s 만큼 줄인 직사각형의 네 귀를 g 로 모따기한
    팔각형이다. 옆면은 u = s 평면, 모따기면은 u + v = c 평면이라 전부 평면이다.

    감격벽이면 코너마다 수직 헌치가 끝나는 높이가 다르므로 모따기를 코너별로
    따로 계산한다. 헌치가 끝나는 높이에서는 같은 z 에 단면을 두 번 넣어
    수평한 마감면을 만든다.
    """
    x0, x1, y0, y1 = cell['x0'], cell['x1'], cell['y0'], cell['y1']
    tb = sp.tb
    ht = fill_top - tb
    if ht <= EPS:
        return Mesh(name, layer)
    #: 코너 이름 → 수직 헌치가 끝나는 높이 (저판 상면 기준)
    ctop = {(round(cx, 9), round(cy, 9)): ch
            for _t, cx, cy, _sx, _sy, ch in corners}

    def sect(w, upper):
        """upper=True 면 그 높이에서 헌치가 끝난 것으로 본다 (모따기 없음)."""
        s, g = haunch_section(sp, w)
        ax0, ax1, ay0, ay1 = x0 + s, x1 - s, y0 + s, y1 - s
        if ax1 - ax0 <= EPS or ay1 - ay0 <= EPS:
            return [((ax0 + ax1) / 2, (ay0 + ay1) / 2)] * 8
        gm = min((ax1 - ax0) / 2.0, (ay1 - ay0) / 2.0)

        def gc(cx, cy):
            h = ctop.get((round(cx, 9), round(cy, 9)), 0.0)
            if w > h + EPS or (upper and abs(w - h) <= EPS):
                return 0.0                     # 그 코너의 헌치는 끝났다
            return min(g, gm)

        g00, g10 = gc(x0, y0), gc(x1, y0)
        g01, g11 = gc(x0, y1), gc(x1, y1)
        return [(ax0 + g00, ay0), (ax1 - g10, ay0), (ax1, ay0 + g10),
                (ax1, ay1 - g11), (ax1 - g11, ay1), (ax0 + g01, ay1),
                (ax0, ay1 - g01), (ax0, ay0 + g00)]

    knots = set(haunch_knots(sp, ht))
    for h in ctop.values():                    # 헌치가 끝나는 높이도 경계다
        if EPS < h < ht - EPS:
            knots.add(h)
    secs = []
    for w in sorted(knots):
        secs.append((tb + w, sect(w, False)))
        if any(abs(w - h) <= EPS for h in ctop.values()) and w < ht - EPS:
            secs.append((tb + w, sect(w, True)))   # 헌치 끝 - 수평 마감면
    if len(secs) < 2:
        return Mesh(name, layer)
    return mesh_loft(secs, name, layer)


def mesh_footing_ring(sp, name='', layer=''):
    """풋팅 - 저판과 겹치지 않는 링 4상자."""
    wf, wr, ws, t = sp.ft_front, sp.ft_rear, sp.ft_side, sp.ft_t
    out = []
    for tag, (X, Y, dx, dy) in (
            ('전면', (-wf, -ws, wf, sp.L + 2 * ws)),
            ('후면', (sp.B, -ws, wr, sp.L + 2 * ws)),
            ('좌측', (0.0, -ws, sp.B, ws)),
            ('우측', (0.0, sp.L, sp.B, ws))):
        if dx <= EPS or dy <= EPS:
            continue
        out.append(mesh_box(Box(X, Y, 0.0, dx, dy, t), f"{name} {tag}", layer))
    return out


#: _cat() 분류 → 3D 레이어 (2D 도면과 같은 이름을 쓴다)
CAT_LAYER = {'base': 'CAISSON-BASE', 'wall': 'CAISSON-WALL',
             'part': 'CAISSON-PART', 'haunch': 'HAUNCH', 'foot': 'FOOTING',
             'key': 'SHEARKEY', 'cop': 'COPING', 'cover': 'COVER',
             'etc': 'OUTLINE'}


class Dxf3DExporter:
    """엔진의 부호 있는 도형 분해를 3D 메시로 옮겨 DXF 로 내보낸다.

    겉모양에 드러나는 공제(유공 · 상치 전면경사)는 실제로 뚫고, 속에서만 겹치는
    공제(격벽 교차부 · 풋팅 본체 · 속채움 헌치)는 겉면이 같으므로 생략한다.
    Lump(헌치 코너 보정)는 형상이 없어 메시로 만들 수 없다.
    """

    def __init__(self, eng: CaissonEngine):
        self.e = eng
        self.sp = eng.sp
        self.meshes = []
        self.skipped_lump = 0.0
        self.omitted = 0.0

    # ------------------------------------------------------------------
    def _hole_wall_names(self):
        """유공이 뚫린 부재 이름 → 그 벽의 기하."""
        return {g['name']: g for g in getattr(self.e, 'wall_geoms', [])}

    def _coping_profiles(self):
        """(부재명 → x-z 프로필) — 전면 경사를 프로필에서 미리 잘라 둔다.

        경사가 본체 것인지 파라펫 것인지는 build() 와 같은 규칙으로 정한다.
        """
        sp = self.sp
        if not sp.cp_on:
            return {}
        z0 = sp.cp_z0 if sp.cp_z0 > EPS else sp.H
        boxes = {'상치 본체': (sp.cp_x0, z0, sp.cp_w, sp.cp_h)}
        front_x, top_z = sp.cp_x0, z0 + sp.cp_h
        if sp.cp_par_on:
            px = (sp.cp_x0 if sp.cp_par_side == SIDE_FRONT
                  else sp.cp_x0 + sp.cp_w - sp.cp_par_w)
            boxes['상치 파라펫'] = (px, z0 + sp.cp_h, sp.cp_par_w, sp.cp_par_h)
            if sp.cp_par_side == SIDE_FRONT:
                front_x, top_z = px, z0 + sp.cp_h + sp.cp_par_h
        cut = (sp.cp_sl_a > EPS and sp.cp_sl_b > EPS)
        out = {}
        for nm, (X, Z, Wd, Ht) in boxes.items():
            pts = [(X, Z), (X + Wd, Z), (X + Wd, Z + Ht), (X, Z + Ht)]
            if cut and abs(X - front_x) < EPS and abs(Z + Ht - top_z) < EPS:
                pts = [(X, Z), (X + Wd, Z), (X + Wd, Z + Ht),
                       (X + sp.cp_sl_a, Z + Ht), (X, Z + Ht - sp.cp_sl_b)]
            out[nm] = pts
        return out

    # ------------------------------------------------------------------
    #: parts 를 그대로 옮기면 서로 파고들어 체적이 이중 계상되는 부재.
    #: 이들은 셀·격벽에서 겹치지 않게 새로 조립한다.
    REBUILT = ('횡격벽', '풋팅', '헌치', '속채움')

    def build(self):
        """겹치지 않는 분해로 메시를 만든다.

        서로 파고드는 솔리드를 그대로 내보내면 겉모양은 맞아도 체적을 합했을 때
        겹친 부분이 두 번 세어진다. 아래 다섯 가지를 새로 조립해 부재군별 체적이
        계산서와 일치하게 만든다.
        """
        e, sp = self.e, self.sp
        self.meshes = []
        self.skipped_lump = 0.0
        self.omitted = 0.0
        hw = self._hole_wall_names()
        cop = self._coping_profiles()

        def wall_mesh(g, name, layer, u0=None, u1=None):
            """유공 뚫린 벽 하나. u0/u1 을 주면 그 구간만 (횡격벽 분할용)."""
            a, b = (g['u0'] if u0 is None else u0), (g['u1'] if u1 is None else u1)
            polys = [hole_polygon(h) for h in e.holes
                     if h['wall'] == g['name']
                     and a - EPS <= h['u'] - h['hu']
                     and h['u'] + h['hu'] <= b + EPS]
            return mesh_perforated_wall(g['w0'], g['t'], a, b, g['v0'], g['v1'],
                                        polys, name, layer, axis=g['axis'])

        # --- (1) 겹치지 않는 부재는 그대로 옮긴다 ---
        for p in e.parts:
            if p.sign < 0:
                continue                        # 공제는 형상에 이미 반영돼 있다
            if not p.prim.samplable:
                continue                        # Lump - 헌치 형상에 흡수했다
            if p.name.startswith(self.REBUILT):
                continue                        # 아래에서 새로 조립한다
            layer = ('FILL' if p.group == GRP_FILL
                     else CAT_LAYER.get(_cat(p.name), 'OUTLINE'))
            if p.name in hw:
                mesh = wall_mesh(hw[p.name], p.name, layer)
            elif p.name in cop:
                mesh = mesh_extrude_y(cop[p.name], 0.0, sp.L, p.name, layer)
            elif isinstance(p.prim, Box):
                mesh = mesh_box(p.prim, p.name, layer)
            elif isinstance(p.prim, TriPrism):
                mesh = mesh_triprism(p.prim, p.name, layer)
            elif isinstance(p.prim, Cyl):
                mesh = mesh_cyl(p.prim, p.name, layer)
            else:
                continue
            self.meshes.append(mesh)

        # --- (2) 횡격벽 : 종격벽 사이 구간으로 쪼갠다 (교차부 이중 계상 제거) ---
        #   유공이 뚫린 횡격벽이면 구간마다 그 안에 든 유공을 함께 뚫는다.
        #   유공이 교차부를 관통하는 배치는 엔진에서 이미 오류로 막았다.
        for j, d in enumerate(e.tps):
            nm = f"횡격벽 #{j + 1}"
            g = hw.get(nm)
            for k, (xa, xb) in enumerate(e.xspans):
                if xb - xa <= EPS:
                    continue
                if g is not None:
                    self.meshes.append(
                        wall_mesh(g, f"{nm}-{k + 1}", 'CAISSON-PART', xa, xb))
                    continue
                self.meshes.append(mesh_box(
                    Box(xa, d['lo'], sp.tb, xb - xa, d['t'], d['h']),
                    f"{nm}-{k + 1}", 'CAISSON-PART'))
            for i, q in enumerate(e.lps):       # 감격벽이면 위쪽만 교차부에 얹는다
                dh = d['h'] - q['h']
                if dh > EPS:
                    self.meshes.append(mesh_box(
                        Box(q['lo'], d['lo'], sp.tb + q['h'], q['t'], d['t'], dh),
                        f"{nm} 교차상부 L#{i + 1}", 'CAISSON-PART'))

        # --- (3) 풋팅 : 링 4상자 (저판과 겹치지 않는다) ---
        if max(sp.ft_front, sp.ft_rear, sp.ft_side) > EPS and sp.ft_t > EPS:
            self.meshes += mesh_footing_ring(sp, '풋팅', 'FOOTING')

        # --- (4)(5) 헌치와 속채움 : 셀마다 실제 형상으로 ---
        for c in e.cells:
            cors = e._cell_corners(c)
            if sp.hn_on:
                for side in ('x0', 'x1', 'y0', 'y1'):
                    self.meshes.append(mesh_haunch_wedge(
                        c, sp.hn_a, sp.hn_b, sp.tb, side,
                        f"헌치 {c['name']} {side}", 'HAUNCH'))
            if sp.vh_on:
                cap = e.cell_vh_top(c)          # 덮개가 있으면 그 밑면까지만
                for cor in cors:
                    top = sp.tb + cor[5]
                    if cap is not None:
                        top = min(top, cap)
                    msh = mesh_vhaunch_loft(
                        c, cor, sp, top,
                        f"헌치 {c['name']} 연직 {cor[0]}", 'HAUNCH')
                    if msh.faces:
                        self.meshes.append(msh)
            msh = mesh_fill_loft(c, cors if sp.vh_on else [], sp, c['fill_top'],
                                 f"속채움 {c['name']}", 'FILL')
            if msh.faces:
                self.meshes.append(msh)
        return self

    def group_volume(self):
        """레이어별 메시 체적 합계."""
        out = {}
        for m in self.meshes:
            out[m.layer] = out.get(m.layer, 0.0) + m.volume()
        return out

    def problems(self):
        """닫히지 않았거나 방향이 뒤집힌 메시 목록 — 자체 점검용."""
        bad = []
        for m in self.meshes:
            oe = m.open_edges()
            if oe or m.volume() <= 0:
                bad.append((m.name, len(oe), m.volume()))
        return bad

    #: 3D 레이어 → 계산서 부재군
    LAYER_GROUP = {'CAISSON-BASE': GRP_CONC, 'CAISSON-WALL': GRP_CONC,
                   'CAISSON-PART': GRP_CONC, 'HAUNCH': GRP_CONC,
                   'FOOTING': GRP_CONC, 'SHEARKEY': GRP_CONC,
                   'OUTLINE': GRP_CONC, 'COVER': GRP_CONC,
                   'FILL': GRP_FILL, 'COPING': GRP_COPING}

    def circle_gap(self):
        """원형 유공을 N각형으로 근사해 '덜 판' 만큼. 콘크리트가 그만큼 남는다."""
        return sum(h['prim'].volume() - ngon_volume(h['hu'], h['t'])
                   for h in self.e.holes if h['circle'])

    def reconcile(self):
        """3D 모델 체적과 계산서 체적을 부재군별로 대조한다.

        계산서 §6 자동검산과 같은 취지다. 겹치지 않는 분해로 만들었으므로
        원형 유공 근사분 말고는 차이가 없어야 한다.
        """
        if not self.meshes:
            self.build()
        gv = self.group_volume()
        m3d = {}
        for lay, v in gv.items():
            g = self.LAYER_GROUP.get(lay, GRP_CONC)
            m3d[g] = m3d.get(g, 0.0) + v
        gap = self.circle_gap()
        rows, tot3, totr = [], 0.0, 0.0
        for g in (GRP_CONC, GRP_FILL, GRP_COPING):
            v3, vr = m3d.get(g, 0.0), self.e.group_props(g)[0]
            if abs(v3) < 1e-12 and abs(vr) < 1e-12:
                continue
            rows.append((GRP_LABEL[g], v3, vr, v3 - vr))
            tot3 += v3
            totr += vr
        d = tot3 - totr
        tol = abs(gap) + max(1e-6, abs(totr) * 1e-9)
        return dict(rows=rows, v3d=tot3, v_rep=totr, delta=d,
                    gap=gap, ok=(abs(d - gap) <= tol))

    def reconcile_text(self):
        """대조표를 사람이 읽을 문자열로."""
        r = self.reconcile()
        w = 62
        out = [' 3D 모델 ↔ 계산서 체적 대조 (m³) '.center(w, '-'),
               f"{'부재군':<12s}{'3D 모델':>15s}{'계산서':>15s}{'차':>14s}"]
        for nm, v3, vr, dd in r['rows']:
            out.append(f"{nm:<12s}{v3:15.3f}{vr:15.3f}{dd:+14.3f}")
        out.append('-' * w)
        out.append(f"{'합계':<12s}{r['v3d']:15.3f}{r['v_rep']:15.3f}"
                   f"{r['delta']:+14.3f}")
        if r['ok']:
            out.append(f"➔ O.K   차이는 유공을 {CIRCLE_SEG} 각형으로 근사한 분"
                       f"({r['gap']:+.3f} m³) 뿐입니다.")
        else:
            out.append(f"➔ N.G ★ 원 근사 예상량 {r['gap']:+.3f} m³ 을 넘는 차이가"
                       f" 있습니다.")
            out.append("   [헌치 코너 보정] 을 껐거나 채움 상단이 헌치 상단보다")
            out.append("   낮으면 계산서 쪽이 근사이므로 값이 갈립니다.")
            out.append("   계산서 경고를 확인하고, 3D 체적을 물량에 쓰지 마십시오.")
        return "\n".join(out)

    def save(self, path):
        if not self.meshes:
            self.build()
        d = DxfWriter()
        for m in self.meshes:
            d.polyface(m.layer, m)
        d.save(path)
        return sorted({m.layer for m in self.meshes})



# ==========================================================================
#  계산부 고정 - Streamlit 단일 파일에서는 이 장치가 반드시 있어야 한다
#
#  Streamlit 은 화면을 다시 그릴 때마다 이 파일을 처음부터 새 이름공간에서
#  다시 실행한다. 그러면 위 §1 의 class 문도 다시 돌아 Box · Spec ·
#  CaissonEngine 이 "이름은 같지만 다른" 새 클래스가 된다. 지난 회차에
#  st.session_state 에 담아 둔 엔진은 옛 클래스의 인스턴스이므로
#  isinstance(p.prim, Box) 가 거짓이 되고, 3D 모델에서 저판·측벽 같은 부재가
#  아무 경고 없이 빠져 버린다. (실제로 부재 124 개 → 118 개, 체적 대조 N.G)
#
#  그래서 맨 처음 회차에 만든 계산부를 sys.modules 에 통째로 넣어 두고,
#  다음 회차부터는 그것으로 되돌려 클래스의 정체성을 붙잡아 둔다.
#  ※ 파일을 app.py / caisson_core.py 둘로 나누면 import 가 이 일을 대신해
#     주므로 이 장치가 필요 없다. 한 파일로 합친 대가로 붙인 것이다.
# ==========================================================================
import sys as _sys
import types as _types

_CORE_NS = '_caisson_core_fixed'

if _CORE_NS in _sys.modules:                    # 두 번째 회차부터
    globals().update({_k: _v
                      for _k, _v in vars(_sys.modules[_CORE_NS]).items()
                      if not _k.startswith('__')})
else:                                           # 맨 처음 회차
    _core_mod = _types.ModuleType(_CORE_NS)
    _core_mod.__dict__.update({_k: _v for _k, _v in globals().items()
                               if not _k.startswith('__')})
    _sys.modules[_CORE_NS] = _core_mod


# ==========================================================================
# ==========================================================================
#  5. 화면 (Streamlit)
#
#     여기서부터가 원본 §5 UI(CaissonApp) 를 갈아 끼운 부분이다.
#     위쪽 §0~§4 는 원본 계산부 그대로이므로 아래 코드가 바뀌어도
#     계산 결과에는 영향을 주지 않는다.
# ==========================================================================
# ==========================================================================


# 한글 글꼴이 없는 서버에서 쏟아지는 글리프 경고가 로그를 덮지 않게 막는다.
warnings.filterwarnings('ignore', message='Glyph .* missing from font')

st.set_page_config(page_title="케이슨 체적 · 무게중심 산정 Ver.1.0",
                   page_icon="🧱", layout="wide",
                   initial_sidebar_state="expanded")

SS = st.session_state

# ==========================================================================
# 0. 상수 (원본 §5 UI 상수와 동일)
# ==========================================================================
FILE_EXT = ".cais"

#: [저장하기] 로 만들 수 있는 것들. (키, 이름, 확장자, 계산 결과가 필요한가, 설명)
SAVE_ITEMS = (
    ('spec',  '입력제원',       FILE_EXT, False, '다시 불러올 수 있는 제원 파일'),
    ('txt',   '계산서(텍스트)', '.txt',   True,  '화면의 계산서 그대로'),
    ('xlsx',  '계산서(엑셀)',   '.xlsx',  True,  '부재별·부재군·셀·검산 4 시트'),
    ('dxf2d', '도면',           '.dxf',   True,  '평면도·정면도·측면도 (CAD)'),
    ('dxf3d', '3D모델',         '.dxf',   True,  '3차원 형상 메시 (CAD)'),
    ('png',   '도해',           '.png',   True,  '3D·평면·정면·측면 이미지 4 장'),
)
SAVE_ITEM = {k: (nm, ext, need) for k, nm, ext, need, _d in SAVE_ITEMS}
FIG_SUFFIX = (('3d', '3D'), ('plan', '평면도'),
              ('front', '정면도'), ('side', '측면도'))

#: 표시 그룹 → 변수 키. 원본과 같은 이름이라 .cais 파일이 그대로 호환된다.
FIG_KEYS = {SHOW_BODY: 'fig_body', SHOW_HAUNCH: 'fig_hn', SHOW_FILL: 'fig_fill',
            SHOW_COPING: 'fig_coping', SHOW_FOOT: 'fig_foot',
            SHOW_HOLE: 'fig_hole'}

#: 계산서용 고정폭 글꼴(브라우저). 한글이 ASCII 의 정확히 2 배인 것만 골랐다.
MONO_STACK = ("'D2Coding','D2Coding ligature','NanumGothicCoding',"
              "'Nanum Gothic Coding','나눔고딕코딩','Sarasa Mono K',"
              "'굴림체','GulimChe','돋움체','DotumChe','Consolas',monospace")
WEBFONT_CSS = ("https://fonts.googleapis.com/css2?"
               "family=Nanum+Gothic+Coding&display=swap")

PV = 'v_'      # 일반 변수         (원본 self.vars)
PR = 'r_'      # 동적 표의 한 칸   (원본 lp_rows / tp_rows / hole_rows / key_rows)
PC = 'c_'      # 셀 표의 한 칸     (원본 cell_vars / cell_top_vars / cell_cover_vars)
PW = 'w_'      # 유공 대상 벽 체크 (원본 wall_vars)

# ==========================================================================
# 1. 변수 정의 - 원본의 _var() 등록을 한곳에 모았다
#    (키, 사람이 읽는 이름, 종류 f실수 i정수 s문자 b참거짓, 기본값)
# ==========================================================================
_D = Spec()

FIELD_DEFS = [
    # --- 1. 기본제원 ---
    ('project', '프로젝트명', 's', _D.project),
    ('name', '케이슨명', 's', _D.name),
    ('ctype', '케이슨 형식', 's', _D.ctype),
    ('B', '폭 B (전후 방향)', 'f', _D.B),
    ('L', '길이 L (연장 방향)', 'f', _D.L),
    ('H', '높이 H', 'f', _D.H),
    ('tb', '저판 두께', 'f', _D.tb),
    ('tf', '전면벽 두께 tf', 'f', _D.tf),
    ('tr', '후면벽 두께 tr', 'f', _D.tr),
    ('ts_l', '좌측벽 두께 ts_l', 'f', _D.ts_l),
    ('ts_r', '우측벽 두께 ts_r', 'f', _D.ts_r),
    # --- 2. 격벽 · 셀 ---
    ('lp_mode', '종격벽 배치 방식', 's', _D.lp_mode),
    ('lp_n', '종격벽 매수 (등간격)', 'i', _D.lp_n),
    ('lp_t', '종격벽 두께 (등간격)', 'f', _D.lp_t),
    ('lp_h', '종격벽 높이 (0 = 전高)', 'f', _D.lp_h),
    ('tp_mode', '횡격벽 배치 방식', 's', _D.tp_mode),
    ('tp_n', '횡격벽 매수 (등간격)', 'i', _D.tp_n),
    ('tp_t', '횡격벽 두께 (등간격)', 'f', _D.tp_t),
    ('tp_h', '횡격벽 높이 (0 = 전高)', 'f', _D.tp_h),
    ('fill_top', '기본 채움 상단 z', 'f', _D.fill_top),
    ('cover_t', '기본 덮개 두께', 'f', _D.cover_t),
    # --- 3. 유공 ---
    ('hole_on', '유공 적용', 'b', _D.hole_on),
    ('hole_shape', '유공 형식', 's', _D.hole_shape),
    ('hole_d', '원형 직경 D', 'f', _D.hole_d),
    ('hole_w', '사각 폭 (y)', 'f', _D.hole_w),
    ('hole_hh', '사각 높이 (z)', 'f', _D.hole_hh),
    ('hole_mode', '유공 배치 방식', 's', _D.hole_mode),
    ('row_n', '행 개수 (z)', 'i', _D.row_n),
    ('row_z0', '첫 행 중심 z', 'f', _D.row_z0),
    ('row_dz', '행 간격', 'f', _D.row_dz),
    ('col_mode', '열 배치', 's', _D.col_mode),
    ('col_per_cell', '격실당 개수 (셀중앙)', 'i', _D.col_per_cell),
    ('col_gap', '유공 간격 (셀중앙)', 'f', _D.col_gap),
    ('col_n', '열 개수 (등간격)', 'i', _D.col_n),
    ('col_y0', '첫 열 중심 y', 'f', _D.col_y0),
    ('col_dy', '열 간격', 'f', _D.col_dy),
    ('hole_cnt', '유공 개수', 'i', 0),
    # --- 4. 풋팅 · 헌치 · 전단키 ---
    ('ft_front', '전면 풋팅 폭', 'f', _D.ft_front),
    ('ft_rear', '후면 풋팅 폭', 'f', _D.ft_rear),
    ('ft_side', '측면 풋팅 폭', 'f', _D.ft_side),
    ('ft_t', '풋팅 두께', 'f', _D.ft_t),
    ('hn_on', '수평 헌치 적용', 'b', _D.hn_on),
    ('hn_a', '수평 다리 a', 'f', _D.hn_a),
    ('hn_b', '연직 다리 b', 'f', _D.hn_b),
    ('vh_on', '수직 헌치 적용', 'b', _D.vh_on),
    ('vh_c', '수직 헌치 다리 c', 'f', _D.vh_c),
    ('hn_corner', '헌치 코너 보정 적용', 'b', _D.hn_corner),
    ('key_cnt', '전단키 개수', 'i', 0),
    # --- 5. 상치 ---
    ('cp_on', '상치 적용', 'b', _D.cp_on),
    ('cp_x0', '본체 전면 x', 'f', _D.cp_x0),
    ('cp_w', '본체 폭', 'f', _D.cp_w),
    ('cp_h', '본체 높이', 'f', _D.cp_h),
    ('cp_z0', '본체 저면 z', 'f', _D.cp_z0),
    ('cp_par_on', '파라펫 적용', 'b', _D.cp_par_on),
    ('cp_par_side', '파라펫 설치 위치', 's', _D.cp_par_side),
    ('cp_par_w', '파라펫 폭', 'f', _D.cp_par_w),
    ('cp_par_h', '파라펫 높이', 'f', _D.cp_par_h),
    ('cp_sl_a', '전면 상단 경사 수평', 'f', _D.cp_sl_a),
    ('cp_sl_b', '전면 상단 경사 연직', 'f', _D.cp_sl_b),
    # --- 6. 재료 ---
    ('use_gamma', '합성 무게중심 산정', 'b', _D.use_gamma),
    ('g_conc', "케이슨 콘크리트 γ", 'f', _D.g_conc),
    ('g_fill', '속채움 γ', 'f', _D.g_fill),
    ('g_cop', '상치콘크리트 γ', 'f', _D.g_cop),
    # --- 7. 결과 ---
    ('detail', '상세 표기', 'b', False),
    ('mono_size', '글꼴 크기', 'i', 10),
    ('amb_w', '애매폭 문자 폭', 'i', 2),
    # --- 8. 도해 ---
    ('fig_body', '케이슨 본체 표시', 'b', True),
    ('fig_hn', '헌치 표시', 'b', True),
    ('fig_fill', '속채움 표시', 'b', True),
    ('fig_coping', '상치콘크리트 표시', 'b', True),
    ('fig_foot', '풋팅·전단키 표시', 'b', True),
    ('fig_hole', '유공 표시', 'b', True),
    ('fig_elev', '3D 올려본각 elev', 'f', 22.0),
    ('fig_azim', '3D 방위각 azim', 'f', -125.0),
]

LABELS = {k: lb for k, lb, _t, _d in FIELD_DEFS}
BOOL_KEYS = {k for k, _lb, t, _d in FIELD_DEFS if t == 'b'}

SELECT_OPTIONS = {
    'ctype': [TYPE_SOLID, TYPE_PERF],
    'lp_mode': [MODE_EVEN, MODE_MANUAL],
    'tp_mode': [MODE_EVEN, MODE_MANUAL],
    'hole_shape': [HOLE_CIRCLE, HOLE_RECT],
    'hole_mode': [HOLE_GRID, MODE_MANUAL],
    'col_mode': [COL_CELL, COL_EVEN],
    'cp_par_side': [SIDE_FRONT, SIDE_REAR],
    'mono_size': [str(v) for v in range(9, 15)],
    'amb_w': ['2', '1'],
}


# ==========================================================================
# 2. 값 취득 · 설정 (원본 gs / gb / gf / gi / _rf 와 같은 규칙)
# ==========================================================================
def gs(key):
    v = SS.get(PV + key, '')
    if isinstance(v, bool):
        return '1' if v else '0'
    return str(v).strip()


def gb(key):
    return gs(key) == '1'


def gf(key, default=0.0):
    s = gs(key)
    if s == '':
        return default
    try:
        return float(s)
    except ValueError:
        raise InputError(f"[{LABELS.get(key, key)}] 숫자가 아닙니다 : '{s}'")


def gi(key, default=0):
    s = gs(key)
    if s == '':
        return default
    try:
        return int(float(s))
    except ValueError:
        raise InputError(f"[{LABELS.get(key, key)}] 정수가 아닙니다 : '{s}'")


def _rf(s, label):
    """동적 표 한 칸의 실수 취득 (원본 CaissonApp._rf)."""
    s = str(s).strip()
    if s == '':
        return 0.0
    try:
        return float(s)
    except ValueError:
        raise InputError(f"[{label}] 숫자가 아닙니다 : '{s}'")


def set_var(key, val):
    """변수 하나를 세팅한다. 실수는 소수 셋째 자리까지 채워 계산서와 자릿수를 맞춘다."""
    k = PV + key
    if key in BOOL_KEYS:
        SS[k] = (val if isinstance(val, bool)
                 else str(val).strip() in ('1', 'True', 'true'))
    elif isinstance(val, float):
        SS[k] = f"{val:.3f}"
    else:
        SS[k] = str(val)


def row_key(name, i, col):
    return f"{PR}{name}_{i}_{col}"


def rget(name, i, col, default=''):
    return str(SS.get(row_key(name, i, col), default)).strip()


# ==========================================================================
# 3. 기본 예제 · 동기화 (원본 reset_defaults / refresh_cells / refresh_wall_list)
# ==========================================================================
def reset_defaults():
    """기본 예제 (유공 격실형 케이슨) 를 채운다. 원본 reset_defaults 와 동일."""
    d = Spec()
    for key, _lb, _t, dv in FIELD_DEFS:
        set_var(key, dv)
    SS['_last_ctype'] = d.ctype
    # 이전 입력을 물려받지 않도록 동적 표·셀·벽을 통째로 비운다
    for k in [k for k in list(SS.keys())
              if str(k).startswith((PR, PC, PW))]:
        del SS[k]
    SS['n_cells'] = 0
    SS['wall_names'] = []
    SS['cell_err'] = ''
    sync_walls(force_default=True)
    set_walls(list(d.hole_walls))
    sync_cells()
    set_all_cells(None if d.ctype == TYPE_PERF else CELL_FILL)
    sync_cells()
    SS['status'] = ('info', "기본 예제 제원 (유공 격실형 케이슨) 을 불러왔습니다.")
    SS['msgs'] = []


def _skeleton_engine():
    """격벽까지만 세운 엔진. 셀 격자·대상 벽체 목록을 뽑는 데 쓴다.

    원본 refresh_cells / refresh_wall_list 가 하던 것과 완전히 같다.
    """
    eng = CaissonEngine(build_spec(skip_holes=True, skip_cells=True))
    eng._validate_basic()
    eng.lps = eng._build_partitions(
        gs('lp_mode'), gi('lp_n'), gf('lp_t'), gf('lp_h'),
        lp_list(), eng.x_in0, eng.x_in1, '종격벽')
    eng.tps = eng._build_partitions(
        gs('tp_mode'), gi('tp_n'), gf('tp_t'), gf('tp_h'),
        tp_list(), eng.y_in0, eng.y_in1, '횡격벽')
    return eng


def sync_walls(force_default=False):
    """유공 대상 벽체 목록을 현재 격벽 수에 맞춘다 (원본 refresh_wall_list)."""
    try:
        eng = _skeleton_engine()
        nl, nt = len(eng.lps), len(eng.tps)
    except Exception:
        nl = nt = 0
    names = (list(WALL_OUTER)
             + [f"종격벽 #{i + 1}" for i in range(nl)]
             + [f"횡격벽 #{j + 1}" for j in range(nt)])
    for nm in names:
        if PW + nm not in SS:
            SS[PW + nm] = False
    SS['wall_names'] = names
    SS['wall_counts'] = (nl, nt)
    if force_default and not any(SS.get(PW + nm) for nm in names):
        SS[PW + WALL_FRONT] = True
    return names


def set_walls(names):
    for nm in SS.get('wall_names', []):
        SS[PW + nm] = (nm in names)


def refresh_wall_list_btn():
    sync_walls(force_default=True)
    nl, nt = SS.get('wall_counts', (0, 0))
    SS['status'] = ('info',
                    f"대상 벽체 목록 갱신 - 외벽 4 + 종격벽 {nl} + 횡격벽 {nt} 매")


def sync_cells():
    """현재 격벽 입력으로 셀 격자를 만들고 셀별 입력값을 정리한다.

    원본 refresh_cells 와 같다. 다만 입력이 잠깐 불완전한 동안(숫자를 지우는 중
    등) 셀 설정이 날아가지 않도록, 실패하면 이전 상태를 그대로 둔다.
    """
    n_old = int(SS.get('n_cells', 0))
    old_types = [SS.get(f"{PC}type_{i}", '') for i in range(n_old)]
    old_tops = [SS.get(f"{PC}top_{i}", '') for i in range(n_old)]
    old_covs = [SS.get(f"{PC}cov_{i}", '') for i in range(n_old)]
    try:
        eng = _skeleton_engine()
        eng.sp.cell_types = old_types
        eng.sp.cell_tops = parse_tops(old_tops)
        eng.sp.cell_covers = parse_tops(old_covs, '덮개 두께')
        eng.sp.cover_t = gf('cover_t')
        eng._make_cells()
    except InputError as ex:
        SS['cell_err'] = str(ex)
        return None
    except Exception as ex:                 # 예상 밖 오류도 표에만 알린다
        SS['cell_err'] = f"{ex}"
        return None
    SS['cell_err'] = ''
    for i, c in enumerate(eng.cells):
        SS[f"{PC}type_{i}"] = c['ctype']
        # 공셀은 정의상 채우지 않는다 (원본 sync_cell_tops)
        empty = (c['ctype'] == CELL_EMPTY)
        SS[f"{PC}top_{i}"] = '0.000' if empty else f"{c['fill_top']:.3f}"
        SS[f"{PC}cov_{i}"] = '0.000' if empty else f"{c['cover']:.3f}"
    SS['n_cells'] = len(eng.cells)
    SS['cells_view'] = [dict(name=c['name'], x0=c['x0'], x1=c['x1'],
                             y0=c['y0'], y1=c['y1']) for c in eng.cells]
    return eng


def refresh_cells_btn():
    """원본 [셀 격자 갱신] 버튼. 자동 동기화가 있어 확인용으로만 남긴다."""
    sync_cells()
    SS['status'] = ('info', f"셀 격자 갱신 - {int(SS.get('n_cells', 0))} 개")


def on_cell_type(i):
    """셀 구분을 바꿨을 때 그 행의 기본값을 맞춰 준다 (원본 _on_cell_type)."""
    ct = SS.get(f"{PC}type_{i}")
    cur = str(SS.get(f"{PC}top_{i}", '')).strip()
    blank = (cur == '' or cur in ('0', '0.0', '0.00', '0.000'))
    if ct == CELL_EMPTY:
        SS[f"{PC}top_{i}"] = '0.000'
    elif ct == CELL_FILL and blank:
        SS[f"{PC}top_{i}"] = gs('fill_top') or '0.000'
    elif ct == CELL_CHAMBER and cur == '':
        SS[f"{PC}top_{i}"] = '0.000'
    cc = str(SS.get(f"{PC}cov_{i}", '')).strip()
    cblank = (cc == '' or cc in ('0', '0.0', '0.00', '0.000'))
    if ct == CELL_EMPTY:
        SS[f"{PC}cov_{i}"] = '0.000'
    elif ct == CELL_CHAMBER and cblank:
        SS[f"{PC}cov_{i}"] = gs('cover_t') or '0.000'


def set_all_cells(value):
    """value 가 None 이면 '최전열만 유수실' 로 설정한다 (원본 _set_all_cells)."""
    if int(SS.get('n_cells', 0)) == 0:
        sync_cells()
    n = int(SS.get('n_cells', 0))
    try:
        nrow = len(tp_positions()) + 1
    except Exception:
        nrow = 1
    for i in range(n):
        if value is None:
            SS[f"{PC}type_{i}"] = CELL_CHAMBER if i < nrow else CELL_FILL
        else:
            SS[f"{PC}type_{i}"] = value
        on_cell_type(i)
    SS['status'] = ('ok', "셀 구분을 일괄 설정했습니다.")


def apply_fill_top_all():
    """[기본 채움 상단 z] 를 공셀이 아닌 모든 셀에 써 넣는다."""
    val = gs('fill_top') or '0.000'
    cnt = 0
    for i in range(int(SS.get('n_cells', 0))):
        if SS.get(f"{PC}type_{i}") == CELL_EMPTY:
            continue
        SS[f"{PC}top_{i}"] = val
        cnt += 1
    SS['status'] = ('ok', f"채움 상단 z = {val} m 를 {cnt} 셀에 적용했습니다.")


def apply_cover_all():
    """[기본 덮개 두께] 를 모든 유수실에 써 넣는다."""
    val = gs('cover_t') or '0.000'
    cnt = 0
    for i in range(int(SS.get('n_cells', 0))):
        if SS.get(f"{PC}type_{i}") != CELL_CHAMBER:
            continue
        SS[f"{PC}cov_{i}"] = val
        cnt += 1
    SS['status'] = ('ok', f"덮개 두께 {val} m 를 유수실 {cnt} 셀에 적용했습니다.")


def on_ctype_change():
    """케이슨 형식 ↔ 유공 연동 (원본 _update_gates 앞머리)."""
    solid = (gs('ctype') == TYPE_SOLID)
    if solid:
        SS[PV + 'hole_on'] = False
    elif SS.get('_last_ctype') == TYPE_SOLID:
        SS[PV + 'hole_on'] = True
    SS['_last_ctype'] = gs('ctype')


def set_fig_show(on):
    for k in FIG_KEYS.values():
        SS[PV + k] = bool(on)


# ==========================================================================
# 4. UI → Spec (원본 _lp_list / _tp_list / _parse_tops / _spec 이식)
# ==========================================================================
def lp_list():
    n = gi('lp_n') if gs('lp_mode') == MODE_MANUAL else 0
    return [PartitionSpec(_rf(rget('lp', i, 'pos'), f'종격벽 #{i + 1} 중심'),
                          _rf(rget('lp', i, 't'), f'종격벽 #{i + 1} 두께'),
                          _rf(rget('lp', i, 'h'), f'종격벽 #{i + 1} 높이'))
            for i in range(max(0, n))]


def tp_list():
    n = gi('tp_n') if gs('tp_mode') == MODE_MANUAL else 0
    return [PartitionSpec(_rf(rget('tp', i, 'pos'), f'횡격벽 #{i + 1} 중심'),
                          _rf(rget('tp', i, 't'), f'횡격벽 #{i + 1} 두께'),
                          _rf(rget('tp', i, 'h'), f'횡격벽 #{i + 1} 높이'))
            for i in range(max(0, n))]


def parse_tops(strs, label='채움 상단 z'):
    """셀별 값 문자열 목록 → [float | None]. 빈칸은 None (기본값 적용)."""
    out = []
    for i, s in enumerate(strs):
        s = str(s).strip()
        if s == '':
            out.append(None)
            continue
        try:
            out.append(float(s))
        except ValueError:
            raise InputError(f"[셀 #{i + 1} {label}] 숫자가 아닙니다 : '{s}'")
    return out


def tp_positions():
    if gs('tp_mode') == MODE_MANUAL:
        return tp_list()
    return [None] * gi('tp_n')


def build_spec(skip_holes=False, skip_cells=False):
    """원본 CaissonApp._spec 과 완전히 같다."""
    sp = Spec()
    sp.project = gs('project')
    sp.name = gs('name')
    sp.ctype = gs('ctype')
    for k in ('B', 'L', 'H', 'tb', 'tf', 'tr', 'ts_l', 'ts_r', 'lp_t', 'lp_h',
              'tp_t', 'tp_h', 'fill_top', 'cover_t',
              'hole_d', 'hole_w', 'hole_hh',
              'row_z0', 'row_dz', 'col_y0', 'col_dy', 'col_gap',
              'ft_front', 'ft_rear', 'ft_side', 'ft_t', 'hn_a', 'hn_b', 'vh_c',
              'cp_x0', 'cp_w', 'cp_h', 'cp_z0', 'cp_par_w', 'cp_par_h',
              'cp_sl_a', 'cp_sl_b', 'g_conc', 'g_fill', 'g_cop'):
        setattr(sp, k, gf(k))
    for k in ('lp_n', 'tp_n', 'row_n', 'col_n', 'col_per_cell'):
        setattr(sp, k, gi(k))
    for k in ('lp_mode', 'tp_mode', 'hole_shape', 'hole_mode',
              'col_mode', 'cp_par_side'):
        setattr(sp, k, gs(k))
    sp.hole_walls = [nm for nm in SS.get('wall_names', []) if SS.get(PW + nm)]
    for k in ('hole_on', 'hn_on', 'vh_on', 'hn_corner', 'cp_on', 'cp_par_on',
              'use_gamma'):
        setattr(sp, k, gb(k))
    sp.lp_list = lp_list()
    sp.tp_list = tp_list()
    nh = gi('hole_cnt') if gs('hole_mode') == MODE_MANUAL else 0
    sp.hole_list = [HoleSpec(_rf(rget('hole', i, 'y'), f'유공 #{i + 1} y'),
                             _rf(rget('hole', i, 'z'), f'유공 #{i + 1} z'))
                    for i in range(max(0, nh))]
    nk = gi('key_cnt')
    sp.keys = [KeySpec(_rf(rget('key', i, 'x'), f'전단키 #{i + 1} x'),
                       _rf(rget('key', i, 'w'), f'전단키 #{i + 1} 폭'),
                       _rf(rget('key', i, 'd'), f'전단키 #{i + 1} 깊이'))
               for i in range(max(0, nk))]
    n = int(SS.get('n_cells', 0))
    sp.cell_types = [] if skip_cells else [SS.get(f"{PC}type_{i}", '')
                                           for i in range(n)]
    sp.cell_tops = [] if skip_cells else parse_tops(
        [SS.get(f"{PC}top_{i}", '') for i in range(n)])
    sp.cell_covers = [] if skip_cells else parse_tops(
        [SS.get(f"{PC}cov_{i}", '') for i in range(n)], '덮개 두께')
    # 무공 형식이면 유공을 절대 뚫지 않는다 (UI 연동과 이중으로 보장)
    if skip_holes or sp.ctype == TYPE_SOLID:
        sp.hole_on = False
    return sp


# ==========================================================================
# 5. 실행 (원본 run / run_verify)
# ==========================================================================
def run(verify=False):
    SS['msgs'] = []
    try:
        sp = build_spec()
        eng = CaissonEngine(sp).build()
    except InputError as ex:
        SS['status'] = ('error', "입력 오류")
        SS['msgs'] = [('error', f"**입력 오류**　　{ex}")]
        return
    except Exception as ex:
        SS['status'] = ('error', "계산 오류")
        SS['msgs'] = [('error', f"**계산 오류**　　{ex}\n\n```\n"
                                f"{traceback.format_exc()}\n```")]
        return

    if verify:
        with st.spinner("몬테카를로 검증 중 … (표본 2,000,000 점)"):
            eng.mc = eng.monte_carlo(groups=(GRP_CONC,), n=2_000_000)

    SS['eng'] = eng
    SS['calc_token'] = SS.get('calc_token', 0) + 1
    SS['files'] = {}                # 계산이 바뀌면 만들어 둔 파일은 버린다
    SS['last3d_text'] = ''

    ck = eng.check
    Vc = eng.group_props(GRP_CONC)[0]
    cb = eng.combined()
    gtxt = (f" · G({cb['gx']:.3f}, {cb['gy']:.3f}, {cb['gz']:.3f})" if cb else '')
    msg = (f"콘크리트 {Vc:.3f} / 속채움 {eng.group_props(GRP_FILL)[0]:.3f}"
           f" / 상치 {eng.group_props(GRP_COPING)[0]:.3f} m³{gtxt}"
           f" · Δ={ck['delta']:+.1e}")
    if not ck['ok']:
        SS['status'] = ('error', "★ 검산 실패 - " + msg)
        SS['msgs'].append(
            ('error', f"**검산 실패** — 외곽체적 대조 잔차가 {ck['delta']:+.9f} m³ "
                      f"입니다. 부재 분해에 중복 또는 누락이 있습니다. "
                      f"결과를 쓰지 마십시오. 계산서 §6 을 확인하십시오."))
    else:
        SS['status'] = ('ok', "계산 완료 - " + msg)
    if eng.warns:
        SS['msgs'].append(('warn', "**경고**\n\n"
                           + "\n\n".join(f"- {w}" for w in eng.warns)))


# ==========================================================================
# 6. 계산서 · 도해
# ==========================================================================
def report_text():
    """현재 설정으로 계산서 본문을 만든다 (원본 _refresh_report)."""
    set_ambiguous_width(gi('amb_w', 2))
    return ReportBuilder(SS['eng'], detail=gb('detail')).build()


def fig_show():
    """도해 탭 체크박스 → 표시 그룹 집합 (원본 _fig_show)."""
    return {g for g, k in FIG_KEYS.items() if gb(k)}


def build_figs():
    """4 개 도해를 만든다. 같은 조건이면 다시 그리지 않는다 (원본 _refresh_figs)."""
    show = fig_show()
    try:
        elev, azim = gf('fig_elev', 22.0), gf('fig_azim', -125.0)
    except InputError:
        elev, azim = 22.0, -125.0
    sig = (SS.get('calc_token', 0), tuple(sorted(show)), elev, azim)
    if SS.get('fig_sig') == sig and SS.get('figs'):
        return SS['figs']
    pl = Plotter(SS['eng'], show=show)
    figs = {}
    for key, fn, is3d in (('3d', pl.draw_3d, True),
                          ('plan', pl.draw_plan, False),
                          ('front', pl.draw_front, False),
                          ('side', pl.draw_side, False)):
        fig = plt.Figure(figsize=(9, 6), dpi=100)
        ax = (fig.add_subplot(111, projection='3d') if is3d
              else fig.add_subplot(111))
        try:
            fn(ax)
            if is3d:
                ax.view_init(elev=elev, azim=azim)   # 원본 툴바 회전을 대신한다
            fig.tight_layout()
        except Exception:
            ax.clear()
            ax.text(0.5, 0.5, "도해를 그릴 수 없습니다.\n" + traceback.format_exc(),
                    ha='center', va='center', fontsize=7)
        figs[key] = fig
    SS['figs'], SS['fig_sig'] = figs, sig
    return figs


# ==========================================================================
# 7. 파일 입출력 (원본 _spec_data / apply_data / _write_*)
# ==========================================================================
def spec_data():
    """제원 파일(.cais)에 담을 입력값 일체. 원본과 같은 형식이다."""
    n = int(SS.get('n_cells', 0))
    nl = gi('lp_n') if gs('lp_mode') == MODE_MANUAL else 0
    nt = gi('tp_n') if gs('tp_mode') == MODE_MANUAL else 0
    nh = gi('hole_cnt') if gs('hole_mode') == MODE_MANUAL else 0
    nk = gi('key_cnt')
    return dict(
        vars={k: gs(k) for k, _lb, _t, _d in FIELD_DEFS},
        lp=[{c: rget('lp', i, c) for c in ('pos', 't', 'h')} for i in range(nl)],
        tp=[{c: rget('tp', i, c) for c in ('pos', 't', 'h')} for i in range(nt)],
        holes=[{c: rget('hole', i, c) for c in ('y', 'z')} for i in range(nh)],
        keys=[{c: rget('key', i, c) for c in ('x', 'w', 'd')} for i in range(nk)],
        cells=[str(SS.get(f"{PC}type_{i}", '')) for i in range(n)],
        cell_tops=[str(SS.get(f"{PC}top_{i}", '')) for i in range(n)],
        cell_covers=[str(SS.get(f"{PC}cov_{i}", '')) for i in range(n)],
        hole_walls=[nm for nm in SS.get('wall_names', []) if SS.get(PW + nm)],
    )


def apply_data(data):
    """저장 데이터를 화면에 반영한다 (원본 apply_data 와 같은 순서)."""
    v = dict(data.get('vars', {}))
    # --- 구버전 호환 : 측벽 두께 ts 하나 → 좌·우로 나눠 옮긴다 ---
    if 'ts' in v and 'ts_l' not in v:
        v['ts_l'] = v['ts_r'] = v['ts']
    for k, val in v.items():
        if k in LABELS:
            set_var(k, val)
    for k, opts in SELECT_OPTIONS.items():          # 깨진 파일 보호
        if gs(k) not in opts:
            set_var(k, opts[0])
    # 동적 표 - 개수 키를 먼저 넣었으므로 이제 행 값을 채운다
    for name, key in (('lp', 'lp'), ('tp', 'tp'),
                      ('holes', 'hole'), ('keys', 'key')):
        for i, d in enumerate(data.get(name, [])):
            for c, val in d.items():
                SS[row_key(key, i, c)] = str(val)
    SS['_last_ctype'] = gs('ctype')
    # 대상 벽체 목록은 격벽 수에 따라 달라지므로 목록을 먼저 갱신한다
    sync_walls()
    walls = data.get('hole_walls')
    if walls is None:                   # 구버전 파일은 hole_wall 문자열 하나뿐이다
        walls = [v.get('hole_wall', WALL_FRONT)]
    set_walls(walls)
    # 셀 - 구분을 먼저 넣는다 (채움 상단 기본값이 구분에 따라 달라지기 때문)
    cells = data.get('cells', [])
    SS['n_cells'] = len(cells)
    for i, val in enumerate(cells):
        SS[f"{PC}type_{i}"] = str(val)
        # 구버전 파일 - 셀별 채움/덮개가 없으므로 구분에 따른 기본값으로 되돌린다
        SS[f"{PC}top_{i}"] = ((gs('fill_top') or '0.000')
                              if str(val) == CELL_FILL else '0.000')
        SS[f"{PC}cov_{i}"] = '0.000'
    for i, val in enumerate(data.get('cell_tops') or []):
        SS[f"{PC}top_{i}"] = str(val)
    for i, val in enumerate(data.get('cell_covers') or []):
        SS[f"{PC}cov_{i}"] = str(val)
    sync_cells()
    SS['status'] = ('ok', "불러오기 완료")
    SS['msgs'] = []


def safe_name(s):
    """파일 이름으로 쓸 수 없는 글자를 바꾼다 (원본 _safe_name)."""
    s = str(s).strip()
    for ch in '\\/:*?"<>|':
        s = s.replace(ch, '_')
    return s or '케이슨'


def save_targets(key, name):
    """항목 하나가 만들 파일 이름 목록 (원본 _save_targets)."""
    it = SAVE_ITEM[key]
    if key == 'png':
        return [f"{name}_{it[0]}_{sfx}{it[1]}" for _k, sfx in FIG_SUFFIX]
    return [f"{name}_{it[0]}{it[1]}"]


def write_spec_bytes():
    blob = json.dumps(spec_data(), ensure_ascii=False, indent=4).encode('utf-8')
    return [blob], "입력값 일체 (다시 불러올 수 있음)"


def write_txt_bytes():
    txt = report_text()
    return [txt.encode('utf-8')], f"{len(txt.splitlines()):,} 줄"


def write_xlsx_bytes():
    """부재별·부재군·셀·검산 4 시트 (원본 _write_report_xlsx 와 동일)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise InputError("openpyxl 이 설치되어 있지 않습니다.  pip install openpyxl")
    e = SS['eng']
    wb = Workbook()
    ws = wb.active
    ws.title = "부재별"
    hdr = ['부재군', '부재명', '부호', '산출식', '체적(m³)',
           'Xg', 'Yg', 'Zg', 'V·Xg', 'V·Yg', 'V·Zg', '비고']
    ws.append(hdr)
    for p in e.parts:
        cx, cy, cz = p.c
        ws.append([GRP_LABEL[p.group], p.name,
                   '+' if p.sign > 0 else '-', p.prim.desc(),
                   p.v, cx, cy, cz, p.v * cx, p.v * cy, p.v * cz, p.note])

    ws2 = wb.create_sheet("부재군")
    ws2.append(['부재군', '체적(m³)', 'Xg', 'Yg', 'Zg', 'e=Xg-B/2'])
    for g in (GRP_CONC, GRP_FILL, GRP_COPING):
        V, cx, cy, cz = e.group_props(g)
        ws2.append([GRP_LABEL[g], V, cx, cy, cz, cx - e.sp.B / 2])
    cb = e.combined()
    if cb:
        ws2.append([])
        ws2.append(['합성 무게중심', '', cb['gx'], cb['gy'], cb['gz'],
                    cb['gx'] - e.sp.B / 2])
        ws2.append(['총 중량 (kN)', cb['W']])

    ws3 = wb.create_sheet("셀")
    ws3.append(['셀', '구분', 'x0', 'x1', 'y0', 'y1', '순폭', '순장',
                '채움상단z', '채움높이', '덮개두께', '덮개 V(m³)',
                '수평헌치 V(m³)', '수직헌치 V(m³)',
                '속채움 V(m³)', 'Zg'])
    for c in e.cells:
        fp = [p for p in e.parts if p.group == GRP_FILL
              and p.roll == f"속채움 {c['name']}"]
        V, _, _, cz = mass_props(fp) if fp else (0, 0, 0, 0)
        ft = c['fill_top'] if fp else 0.0
        vh = sum(p.v for p in e.parts if p.group == GRP_CONC
                 and p.roll == f"헌치 {c['name']}")
        vv = sum(p.v for p in e.parts if p.group == GRP_CONC
                 and p.roll == f"수직헌치 {c['name']}")
        cv = c['cover']
        ws3.append([c['name'], c['ctype'], c['x0'], c['x1'],
                    c['y0'], c['y1'], c['x1'] - c['x0'],
                    c['y1'] - c['y0'], ft,
                    (ft - e.sp.tb) if fp else 0.0,
                    cv, cv * (c['x1'] - c['x0']) * (c['y1'] - c['y0']),
                    vh, vv, V, cz])

    ws4 = wb.create_sheet("검산")
    ck = e.check
    for k, v in (('외곽체적 B·L·H', ck['v_outer']),
                 ('외곽 내부 콘크리트', ck['v_conc_in']),
                 ('공극 ① 부재 분해', ck['void_parts']),
                 ('공극 ② 셀 격자', ck['void_cells']),
                 ('잔차 Δ', ck['delta']),
                 ('판정', 'O.K' if ck['ok'] else 'N.G'),
                 ('좌우대칭 오차', ck['sym_err'])):
        ws4.append([k, v])

    for w in (ws, ws2, ws3, ws4):
        for cell in w[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
    buf = io.BytesIO()
    wb.save(buf)
    return [buf.getvalue()], "부재별 · 부재군 · 셀 · 검산 4 시트"


def _dxf_to_bytes(writer_call):
    """DXF 는 cp949 로 파일에 쓰므로 임시 파일을 거쳐 바이트로 읽는다."""
    with tempfile.TemporaryDirectory() as td:
        path = os.path.join(td, 'out.dxf')
        info = writer_call(path)
        with open(path, 'rb') as f:
            return f.read(), info


def write_dxf2d_bytes():
    data, layers = _dxf_to_bytes(lambda p: DxfExporter(SS['eng']).save(p))
    return [data], f"평면·정면·측면 3 도면, 레이어 {len(layers)} 개 (단위 m)"


def write_dxf3d_bytes():
    ex = Dxf3DExporter(SS['eng']).build()
    data, layers = _dxf_to_bytes(ex.save)
    SS['last3d_text'] = ex.reconcile_text()
    nf = sum(len(m.faces) for m in ex.meshes)
    rc = ex.reconcile()
    return [data], (f"부재 {len(ex.meshes)} 개 · 면 {nf:,} 개 · "
                    f"레이어 {len(layers)} 개  "
                    f"[체적 대조 {'O.K' if rc['ok'] else '★ N.G'}]")


def write_png_bytes():
    figs = build_figs()
    out = []
    for key, _sfx in FIG_SUFFIX:
        buf = io.BytesIO()
        figs[key].savefig(buf, dpi=200, bbox_inches='tight', format='png')
        out.append(buf.getvalue())
    return out, f"{len(out)} 장 (300 dpi 상당)"


WRITERS = {'spec': write_spec_bytes, 'txt': write_txt_bytes,
           'xlsx': write_xlsx_bytes, 'dxf2d': write_dxf2d_bytes,
           'dxf3d': write_dxf3d_bytes, 'png': write_png_bytes}


def set_all_saves(on):
    """[모두 선택] / [모두 해제]. 계산이 필요한 항목은 결과가 있어야 켜진다."""
    ready = SS.get('eng') is not None
    for key, _lb, _e, need, _d in SAVE_ITEMS:
        SS[f"save_{key}"] = bool(on) and (ready or not need)


def make_files(keys, name):
    """고른 항목을 실제로 만든다 (원본 _run_save). 결과를 세션에 담는다."""
    files, okmsg, errmsg = [], [], []
    for key, label, _e, _n, _d in SAVE_ITEMS:
        if key not in keys:
            continue
        names = save_targets(key, name)
        try:
            blobs, note = WRITERS[key]()
            files.extend(zip(names, blobs))
            okmsg.append(f"· {label} : {note}")
        except Exception as ex:
            errmsg.append(f"· {label} : {ex}")
            if not isinstance(ex, InputError):
                errmsg.append(f"　　{traceback.format_exc().splitlines()[-1]}")
    SS['files'] = dict(items=files, ok=okmsg, err=errmsg, name=name)


def zip_bytes(files):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as z:
        for nm, blob in files:
            z.writestr(nm, blob)
    return buf.getvalue()


# ==========================================================================
# 8. 비활성 판정 (원본 _update_gates 를 그대로 옮긴 것)
# ==========================================================================
def compute_gates():
    """지금 선택 상태에서 뜻이 없는 입력칸을 모두 찾아 둔다."""
    g = {}
    solid = (gs('ctype') == TYPE_SOLID)
    if solid:                       # 무공이면 유공을 항상 끈다
        SS[PV + 'hole_on'] = False

    # --- 격벽 : 등간격 전용 값 ---
    for pre in ('lp', 'tp'):
        even = (gs(f'{pre}_mode') == MODE_EVEN)
        g[f'{pre}_t'] = g[f'{pre}_h'] = even

    # --- 유공 ---
    on = gb('hole_on') and not solid
    g['hole_on'] = not solid
    g['hole_shape'] = g['hole_mode'] = on
    g['wall_box'] = on
    circle = (gs('hole_shape') == HOLE_CIRCLE)
    g['hole_d'] = on and circle
    g['hole_w'] = g['hole_hh'] = on and not circle
    grid = (gs('hole_mode') == HOLE_GRID)
    for k in ('row_n', 'row_z0', 'row_dz', 'col_mode'):
        g[k] = on and grid
    cell_col = (on and grid and gs('col_mode') == COL_CELL)
    g['col_per_cell'] = cell_col
    try:
        many = gi('col_per_cell', 1) > 1
    except InputError:
        many = True
    g['col_gap'] = cell_col and many    # 간격은 격실당 2 개 이상일 때만 뜻이 있다
    even_col = (on and grid and gs('col_mode') == COL_EVEN)
    g['col_n'] = g['col_y0'] = g['col_dy'] = even_col
    g['hole_cnt'] = g['hole_box'] = on and not grid
    g['hole_note'] = ("※ [1. 기본제원] 의 케이슨 형식이 '무공' 입니다. "
                      "유공을 쓰려면 형식을 '유공' 으로 바꾸십시오." if solid else
                      ("※ [유공 적용] 이 꺼져 있습니다. 무공 케이슨으로 산정합니다."
                       if not on else ''))

    # --- 풋팅 : 폭이 모두 0 이면 두께는 쓰이지 않는다 ---
    try:
        has_ft = max(gf('ft_front'), gf('ft_rear'), gf('ft_side')) > EPS
    except InputError:
        has_ft = True                   # 아직 숫자가 아니면 건드리지 않는다
    g['ft_t'] = has_ft

    # --- 헌치 ---
    hn, vh = gb('hn_on'), gb('vh_on')
    g['hn_a'] = g['hn_b'] = hn
    g['vh_c'] = vh
    g['hn_corner'] = hn                 # 코너 보정은 수평 헌치가 있어야 뜻이 있다

    # --- 상치 ---
    cp = gb('cp_on')
    for k in ('cp_x0', 'cp_w', 'cp_h', 'cp_z0', 'cp_par_on', 'cp_sl_a'):
        g[k] = cp
    for k in ('cp_par_side', 'cp_par_w', 'cp_par_h'):
        g[k] = cp and gb('cp_par_on')
    try:
        has_sl = gf('cp_sl_a') > EPS
    except InputError:
        has_sl = True
    g['cp_sl_b'] = cp and has_sl

    # --- 재료 ---
    for k in ('g_conc', 'g_fill', 'g_cop'):
        g[k] = gb('use_gamma')
    return g


GATES = {}


def _off(key):
    """disabled 인자용. 등록되지 않은 키는 항상 활성이다."""
    return not GATES.get(key, True)


# ==========================================================================
# 9. 입력 위젯 헬퍼 (원본 _row / _combo / _check / _tbl 자리)
# ==========================================================================
def w_text(col, key, unit='', note='', disabled=None):
    col.text_input(LABELS[key] + (f"  [{unit}]" if unit else ''),
                   key=PV + key, help=note or None,
                   disabled=_off(key) if disabled is None else disabled)


def w_sel(col, key, note='', disabled=None, on_change=None, label=None):
    opts = SELECT_OPTIONS[key]
    if gs(key) not in opts:
        SS[PV + key] = opts[0]
    col.selectbox(label or LABELS[key], opts, key=PV + key, help=note or None,
                  disabled=_off(key) if disabled is None else disabled,
                  on_change=on_change)


def w_chk(col, key, label=None, note='', disabled=None, on_change=None):
    col.checkbox(label or LABELS[key], key=PV + key, help=note or None,
                 disabled=_off(key) if disabled is None else disabled,
                 on_change=on_change)


def dyn_table(container, name, cols, count, defaults, disabled=False):
    """동적 입력표 (원본 _tbl). cols = [(키, 머리글), ...]"""
    if count <= 0:
        container.caption("(없음)")
        return
    widths = [0.5] + [1.4] * len(cols)
    head = container.columns(widths)
    head[0].markdown("**#**")
    for c, (k, t) in enumerate(cols):
        head[c + 1].markdown(f"**{t}**")
    for i in range(count):
        cs = container.columns(widths)
        cs[0].markdown(f"<div style='padding-top:8px'>#{i + 1}</div>",
                       unsafe_allow_html=True)
        for c, (k, t) in enumerate(cols):
            rk = row_key(name, i, k)
            if rk not in SS:
                SS[rk] = str(defaults.get(k, ''))
            cs[c + 1].text_input(f"{t} #{i + 1}", key=rk, disabled=disabled,
                                 label_visibility='collapsed')


def box(title):
    """원본 ttk.LabelFrame 자리."""
    try:
        c = st.container(border=True)
    except TypeError:                   # 옛 Streamlit
        c = st.container()
    c.markdown(f"**{title}**")
    return c


def cell_text(col, s):
    col.markdown(f"<div style='padding-top:8px'>{s}</div>",
                 unsafe_allow_html=True)


# ==========================================================================
# 10. 초기화
# ==========================================================================
if 'inited' not in SS:
    SS['inited'] = True
    SS['eng'] = None
    SS['calc_token'] = 0
    SS['status'] = ('info', "제원을 입력하고 [▶ 계산 실행] 을 누르십시오.")
    SS['msgs'] = []
    SS['files'] = {}
    SS['last3d_text'] = ''
    reset_defaults()

# ---------------------------------------------------------------- 사이드바
with st.sidebar:
    st.markdown("### 케이슨 체적 · 무게중심 산정")
    st.caption("Ver.1.0 · 부호 있는 기본도형 분해 (해석적 정해)")

    if st.button("▶ 계산 실행", use_container_width=True, type="primary"):
        run()
    if st.button("형상 검증 (몬테카를로)", use_container_width=True,
                 help="바운딩박스에 난수 200 만 점을 뿌려 해석해와 대조한다. "
                      "몇 초 걸린다."):
        run(verify=True)
    st.button("기본 예제로 되돌리기", on_click=reset_defaults,
              use_container_width=True,
              help="입력 전체를 기본 예제(유공 격실형)로 되돌립니다.")

    up = st.file_uploader("불러오기 (.cais)", type=['cais', 'json'],
                          key='uploader')
    if up is not None and SS.get('loaded_id') != (up.name, up.size):
        SS['loaded_id'] = (up.name, up.size)
        try:
            apply_data(json.loads(up.getvalue().decode('utf-8')))
            st.rerun()
        except Exception as ex:
            SS['status'] = ('error', f"불러오기 실패 : {ex}")

    st.divider()
    _lvl, _txt = SS.get('status', ('info', ''))
    {'ok': st.success, 'error': st.error,
     'warn': st.warning}.get(_lvl, st.info)(_txt)

    if SS.get('eng') is not None:
        _e = SS['eng']
        st.metric("케이슨 콘크리트 (m³)", f"{_e.group_props(GRP_CONC)[0]:,.3f}")
        st.metric("속채움 (m³)", f"{_e.group_props(GRP_FILL)[0]:,.3f}")
        st.metric("상치콘크리트 (m³)", f"{_e.group_props(GRP_COPING)[0]:,.3f}")
        _cb = _e.combined()
        if _cb:
            st.metric("합성 Xg / Zg (m)", f"{_cb['gx']:.3f} / {_cb['gz']:.3f}",
                      help=f"Yg = {_cb['gy']:.3f} m ／ "
                           f"e = Xg − B/2 = {_cb['gx'] - _e.sp.B / 2:+.3f} m ／ "
                           f"총 중량 {_cb['W']:,.1f} kN")
    if KOREAN_FONT is None:
        st.caption("※ 서버에 한글 글꼴이 없어 도해 글자가 깨질 수 있습니다. "
                   "(리눅스 : apt install fonts-nanum)")

# ------------------------------------------------------- 동기화 · 비활성 판정
sync_walls()
sync_cells()
GATES = compute_gates()

for _lvl, _m in SS.get('msgs', []):
    {'error': st.error, 'warn': st.warning}.get(_lvl, st.info)(_m)

TABS = st.tabs([" 1. 기본제원 ", " 2. 격벽 · 셀 ", " 3. 유공 ",
                " 4. 풋팅 · 헌치 · 전단키 ", " 5. 상치 ", " 6. 재료 ",
                " 7. 결과 ", " 8. 도해 ", " 9. 저장 · 불러오기 "])

# ======================================================== 1. 기본제원
with TABS[0]:
    g = box(" 프로젝트 ")
    c = g.columns(3)
    w_text(c[0], 'project')
    w_text(c[1], 'name')
    w_sel(c[2], 'ctype', '유공을 고르면 최전열 셀이 유수실 기본값이 된다',
          on_change=on_ctype_change)
    g.button("기본 예제로 되돌리기", on_click=reset_defaults, key='btn_reset_1',
             help="입력 전체를 기본 예제(유공 격실형)로 되돌립니다.")

    g = box(" 외곽 치수 ")
    c = g.columns(3)
    w_text(c[0], 'B', 'm', 'x 방향. 전면 외면 = 0')
    w_text(c[1], 'L', 'm', 'y 방향. 좌측 외면 = 0')
    w_text(c[2], 'H', 'm', 'z 방향. 저판 저면 = 0 → 정단')

    g = box(" 부재 두께 ")
    c = g.columns(5)
    w_text(c[0], 'tb', 'm')
    w_text(c[1], 'tf', 'm', '해측')
    w_text(c[2], 'tr', 'm', '항내측')
    w_text(c[3], 'ts_l', 'm', 'y = 0 쪽')
    w_text(c[4], 'ts_r', 'm', 'y = L 쪽')

    st.caption(
        "좌표계　x : 폭 방향 (전면 외면 = 0 → 후면 외면 = B)　·　"
        "y : 길이 방향 (좌측 단부 외면 = 0 → 우측 외면 = L)　·　"
        "z : 연직 방향 (저판 저면 = 0 → 정단 = H). 전단키는 z < 0, 상치는 z > H  "
        "／ 측벽은 전·후면벽 사이 구간만 차지시켜 모서리 중복을 없앤다.")

# ======================================================== 2. 격벽 · 셀
with TABS[1]:
    L1, L2 = st.columns(2)
    with L1:
        g = box(" 종격벽 (연장 방향으로 뻗음 - x 위치로 지정) ")
        c = g.columns(2)
        w_sel(c[0], 'lp_mode')
        w_text(c[1], 'lp_n', '매')
        c = g.columns(2)
        w_text(c[0], 'lp_t', 'm')
        w_text(c[1], 'lp_h', 'm', '감격벽이면 입력')
        if gs('lp_mode') == MODE_MANUAL:
            g.caption("※ 직접입력 모드 - 매수만큼 행이 생긴다.")
            dyn_table(g, 'lp', [('pos', '중심 좌표'), ('t', '두께'),
                                ('h', '높이(0=전高)')],
                      gi('lp_n'),
                      dict(pos='0.000', t=gs('lp_t') or '0.400', h='0.000'))
    with L2:
        g = box(" 횡격벽 (폭 방향으로 뻗음 - y 위치로 지정) ")
        c = g.columns(2)
        w_sel(c[0], 'tp_mode')
        w_text(c[1], 'tp_n', '매')
        c = g.columns(2)
        w_text(c[0], 'tp_t', 'm')
        w_text(c[1], 'tp_h', 'm', '감격벽이면 입력')
        if gs('tp_mode') == MODE_MANUAL:
            g.caption("※ 직접입력 모드 - 매수만큼 행이 생긴다.")
            dyn_table(g, 'tp', [('pos', '중심 좌표'), ('t', '두께'),
                                ('h', '높이(0=전高)')],
                      gi('tp_n'),
                      dict(pos='0.000', t=gs('tp_t') or '0.400', h='0.000'))

    st.caption("※ 등간격은 순내부 구간을 (매수+1) 등분한다.　"
               "종격벽 × 횡격벽 교차부는 전 조합을 자동 공제한다.")

    g = box(" 속채움 (셀별 지정이 없을 때의 기본값) ")
    c = g.columns(4)
    w_text(c[0], 'fill_top', 'm', '저판 저면 기준. 정단이면 H')
    w_text(c[1], 'cover_t', 'm', '유수실 상부 콘크리트. 0 이면 없음')
    c[2].markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
    c[2].button("채움높이 일괄적용", key='btn_fill_all',
                use_container_width=True, on_click=apply_fill_top_all)
    c[3].markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
    c[3].button("덮개 일괄적용", key='btn_cover_all',
                use_container_width=True, on_click=apply_cover_all)
    g.caption(
        "※ 채움 높이·덮개 두께는 셀(격실)마다 따로 지정한다. 위 값은 셀별 지정이 "
        "없을 때의 기본값이며, [채움높이 일괄적용]·[덮개 일괄적용] 으로 아래 셀 표에 "
        "한꺼번에 써 넣을 수 있다.　덮개는 정단(H) 아래로 얹히며 헌치는 덮개 "
        "밑면에서 멈춘다.")

    g = box(" 셀 구분 · 셀별 속채움 ")
    bc = g.columns([1, 1, 1, 3])
    bc[0].button("셀 격자 갱신", key='btn_cell_ref', use_container_width=True,
                 on_click=refresh_cells_btn)
    bc[1].button("모두 속채움", key='btn_cell_fill', use_container_width=True,
                 on_click=set_all_cells, args=(CELL_FILL,))
    bc[2].button("최전열 유수실", key='btn_cell_front', use_container_width=True,
                 on_click=set_all_cells, args=(None,))
    g.caption(
        "※ 격벽을 바꾸면 셀 격자는 자동으로 다시 만들어진다 (원본의 [셀 격자 갱신] "
        "이 필요 없다). 채움 상단 z 는 저판 저면 기준 표고이며 0 이면 채우지 않는다. "
        "유수실도 값을 주면 그 높이까지 채운다. 공셀은 입력할 수 없다. 덮개 두께는 "
        "유수실 상부 콘크리트이며 정단에서 그만큼 아래로 얹힌다 (채움 상단은 덮개 "
        "밑면 아래여야 한다).")

    if SS.get('cell_err'):
        g.error(SS['cell_err'])
    _n = int(SS.get('n_cells', 0))
    _view = SS.get('cells_view', [])
    if _n and len(_view) >= _n:
        _wd = [0.7, 1.1, 1.1, 1.1, 1.5, 1.5, 1.4]
        hd = g.columns(_wd)
        for _k, _t in enumerate(('셀', '구분', '채움 상단 z (m)', '덮개 두께 (m)',
                                 'x 범위', 'y 범위', '순치수')):
            hd[_k].markdown(f"**{_t}**")
        for i in range(_n):
            cv = _view[i]
            cs = g.columns(_wd)
            cell_text(cs[0], cv['name'])
            tkey = f"{PC}type_{i}"
            if SS.get(tkey) not in CELL_TYPES:
                SS[tkey] = CELL_FILL
            cs[1].selectbox("구분", CELL_TYPES, key=tkey,
                            label_visibility='collapsed',
                            on_change=on_cell_type, args=(i,))
            _empty = (SS[tkey] == CELL_EMPTY)
            cs[2].text_input("채움 상단", key=f"{PC}top_{i}",
                             label_visibility='collapsed', disabled=_empty)
            cs[3].text_input("덮개", key=f"{PC}cov_{i}",
                             label_visibility='collapsed', disabled=_empty)
            cell_text(cs[4], f"{cv['x0']:.3f} ~ {cv['x1']:.3f}")
            cell_text(cs[5], f"{cv['y0']:.3f} ~ {cv['y1']:.3f}")
            cell_text(cs[6], f"{cv['x1'] - cv['x0']:.3f} × "
                             f"{cv['y1'] - cv['y0']:.3f}")
        g.caption(f"셀 {_n} 개")

# ======================================================== 3. 유공
with TABS[2]:
    g = box(" 유공 설정 ")
    c = g.columns(3)
    w_chk(c[0], 'hole_on', note='끄면 무공 케이슨으로 산정한다')
    w_sel(c[1], 'hole_shape')
    w_sel(c[2], 'hole_mode')
    c = g.columns(3)
    w_text(c[0], 'hole_d', 'm')
    w_text(c[1], 'hole_w', 'm')
    w_text(c[2], 'hole_hh', 'm')
    if GATES.get('hole_note'):
        g.error(GATES['hole_note'])

    g = box(" 대상 벽체 (여러 벽에 동시에 뚫을 수 있다) ")
    bc = g.columns(4)
    bc[0].button("대상 벽체 목록 갱신", key='btn_wall_ref',
                 use_container_width=True, on_click=refresh_wall_list_btn,
                 disabled=_off('wall_box'))
    bc[1].button("전면벽만", key='btn_wall_front', use_container_width=True,
                 on_click=set_walls, args=([WALL_FRONT],),
                 disabled=_off('wall_box'))
    bc[2].button("외벽 전부", key='btn_wall_outer', use_container_width=True,
                 on_click=set_walls, args=(list(WALL_OUTER),),
                 disabled=_off('wall_box'))
    bc[3].button("모두 끄기", key='btn_wall_none', use_container_width=True,
                 on_click=set_walls, args=([],), disabled=_off('wall_box'))
    g.caption("※ 격벽을 바꾸면 목록은 자동으로 따라간다. "
              "전면·후면·종격벽은 x 방향, 측벽·횡격벽은 y 방향으로 뚫린다.")
    _names = SS.get('wall_names', [])
    if _names:
        wc = g.columns(6)
        for _k, _nm in enumerate(_names):
            wc[_k % 6].checkbox(f"{_nm} ({wall_axis(_nm)})", key=PW + _nm,
                                disabled=_off('wall_box'))

    g = box(" 격자 배치 ")
    c = g.columns(3)
    w_text(c[0], 'row_n', '개')
    w_text(c[1], 'row_z0', 'm')
    w_text(c[2], 'row_dz', 'm')
    c = g.columns(3)
    w_sel(c[0], 'col_mode', '셀중앙 = 횡격벽으로 나뉜 각 칸에 자동 정렬')
    w_text(c[1], 'col_per_cell', '개', '1 = 칸 중앙 1개')
    w_text(c[2], 'col_gap', 'm', '중심간 거리. 0 = 칸 순폭을 개수만큼 등분')
    c = g.columns(3)
    w_text(c[0], 'col_n', '개')
    w_text(c[1], 'col_y0', 'm')
    w_text(c[2], 'col_dy', 'm')

    g = box(" 직접입력 (중심 좌표) ")
    c = g.columns([1, 3])
    w_text(c[0], 'hole_cnt', '개')
    dyn_table(g, 'hole', [('y', '중심 y'), ('z', '중심 z')],
              gi('hole_cnt') if gs('hole_mode') == MODE_MANUAL else 0,
              dict(y='0.000', z='0.000'), disabled=_off('hole_box'))

# ======================================================== 4. 풋팅·헌치·전단키
with TABS[3]:
    g = box(" 풋팅 (저판 돌출부) ")
    c = g.columns(4)
    w_text(c[0], 'ft_front', 'm', '0 이면 없음')
    w_text(c[1], 'ft_rear', 'm')
    w_text(c[2], 'ft_side', 'm', '좌·우 동일')
    w_text(c[3], 'ft_t', 'm', '저판 두께 이하')
    g.caption("※ 링(외곽 − 본체) 방식으로 산정하므로 모서리 중복이 없다.")

    g = box(" 헌치 (저판-벽체 접합부 · 우각부) ")
    c = g.columns(3)
    w_chk(c[0], 'hn_on', '수평 헌치 적용 (저판-벽체)', '모든 셀의 4면에 생성한다')
    w_text(c[1], 'hn_a', 'm')
    w_text(c[2], 'hn_b', 'm')
    c = g.columns(3)
    w_chk(c[0], 'vh_on', '수직 헌치 적용 (우각부)',
          '격벽-격벽 · 격벽-외벽 연직 모서리. 셀 4코너')
    w_text(c[1], 'vh_c', 'm', '45°. 수평 헌치 상단부터 벽체 상단까지')
    w_chk(c[2], 'hn_corner', '헌치 코너 보정 적용',
          '수평·수직 헌치의 코너 누락분을 정확한 도심과 함께 되살린다')

    g = box(" 전단키 (저판 저면 돌출) ")
    c = g.columns([1, 3])
    w_text(c[0], 'key_cnt', '개')
    dyn_table(g, 'key', [('x', '전면에서 x'), ('w', '폭'), ('d', '깊이')],
              gi('key_cnt'), dict(x='0.000', w='0.500', d='0.500'))

# ======================================================== 5. 상치
with TABS[4]:
    g = box(" 상치콘크리트 ")
    w_chk(g, 'cp_on')
    c = g.columns(4)
    w_text(c[0], 'cp_x0', 'm', '0 = 케이슨 전면과 일치')
    w_text(c[1], 'cp_w', 'm')
    w_text(c[2], 'cp_h', 'm')
    w_text(c[3], 'cp_z0', 'm', '0 이면 케이슨 정단 H')

    g = box(" 파라펫 (흉벽) ")
    w_chk(g, 'cp_par_on')
    c = g.columns(3)
    w_sel(c[0], 'cp_par_side')
    w_text(c[1], 'cp_par_w', 'm')
    w_text(c[2], 'cp_par_h', 'm')

    g = box(" 전면 상단 경사 (모따기) ")
    c = g.columns(2)
    w_text(c[0], 'cp_sl_a', 'm', '0 이면 없음')
    w_text(c[1], 'cp_sl_b', 'm')
    g.caption("※ 파라펫이 전면에 있으면 파라펫 상단을, 없으면 본체 상단을 깎는다.")

# ======================================================== 6. 재료
with TABS[5]:
    g = box(" 단위중량  [KDS 64 10 10 재료표] ")
    w_chk(g, 'use_gamma', note='끄면 부재군별 체적·도심만 출력한다')
    c = g.columns(3)
    w_text(c[0], 'g_conc', 'kN/m³', f"철근Con'c {G_RC:.2f}")
    w_text(c[1], 'g_fill', 'kN/m³', f"사석 {G_STONE:.2f}")
    w_text(c[2], 'g_cop', 'kN/m³', f"무근Con'c {G_PC:.2f}")
    st.caption("케이슨 콘크리트 · 속채움 · 상치는 재료가 다르므로 체적만으로는 "
               "하나의 무게중심을 낼 수 없다. 단위중량은 이 세 부재군을 합성하기 "
               "위한 가중치로만 쓴다.")
    st.latex(r"X_g=\frac{\sum (V\cdot\gamma\cdot x)}{\sum (V\cdot\gamma)}")
    st.caption("중량 · 부력 · 수중중량 · 흘수 · 경심고 GM 은 이 프로그램의 범위 밖이다.")

# ======================================================== 7. 결과
with TABS[6]:
    if SS.get('eng') is None:
        st.info("아직 계산하지 않았습니다. 왼쪽의 [▶ 계산 실행] 을 누르십시오.")
    else:
        c = st.columns([2.4, 1, 1.2, 2])
        w_chk(c[0], 'detail', '부재를 모두 개별 표기 (요약 해제)')
        w_sel(c[1], 'mono_size')
        w_sel(c[2], 'amb_w',
              "· × ³ Σ γ π § → ■ 같은 글자를 몇 칸으로 보고 표를 맞출지 정한다. "
              "한글 고정폭 글꼴이면 2, 서양 글꼴이면 1 이다. "
              "표가 어긋나 보이면 바꿔 보십시오.")
        _txt = report_text()
        c[3].download_button("계산서 내려받기 (.txt)", data=_txt.encode('utf-8'),
                             file_name=f"{safe_name(gs('name'))}_계산서.txt",
                             mime='text/plain', use_container_width=True)

        _size = max(6, min(20, gi('mono_size', 10)))
        components.html(f"""
        <link href="{WEBFONT_CSS}" rel="stylesheet">
        <style>
          html, body {{ margin:0; padding:0; background:#FBFCFC; }}
          pre.rep {{ font-family:{MONO_STACK}; font-size:{_size + 3}px;
                     line-height:1.45; color:#1B2631; background:#FBFCFC;
                     margin:0; padding:10px 14px; white-space:pre; }}
        </style>
        <pre class="rep">{html.escape(_txt)}</pre>
        """, height=760, scrolling=True)

        with st.expander("표 정렬 자체 점검 - 계산서 폭(118 칸)을 넘는 줄 찾기"):
            _bad = ReportBuilder(SS['eng'], detail=gb('detail')).layout_errors()
            if _bad:
                st.warning("\n\n".join(f"- {b}" for b in _bad))
            else:
                st.success("모든 줄이 계산서 폭(118 칸) 안에 들어옵니다.")

# ======================================================== 8. 도해
with TABS[7]:
    if SS.get('eng') is None:
        st.info("아직 계산하지 않았습니다. 왼쪽의 [▶ 계산 실행] 을 누르십시오.")
    else:
        g = box(" 표시 ")
        c = g.columns(8)
        for _k, (_grp, _key) in enumerate(FIG_KEYS.items()):
            c[_k].checkbox(SHOW_LABEL[_grp], key=PV + _key)
        c[6].button("모두 켜기", key='btn_fig_on', use_container_width=True,
                    on_click=set_fig_show, args=(True,))
        c[7].button("모두 끄기", key='btn_fig_off', use_container_width=True,
                    on_click=set_fig_show, args=(False,))
        c = g.columns([1, 1, 6])
        c[0].text_input("3D 올려본각 elev", key=PV + 'fig_elev',
                        help="원본 기본값 22")
        c[1].text_input("3D 방위각 azim", key=PV + 'fig_azim',
                        help="원본 기본값 −125. 값을 바꾸면 3D 를 돌려 볼 수 있다.")
        st.caption(
            "※ 축척은 항상 전체 형상 기준으로 고정되므로 부재군을 꺼도 뷰끼리 대조할 "
            "수 있다. 3D 가 느리면 [헌치] 를 끄십시오.　[9. 저장] 에서 [도면] 은 "
            "평면도·정면도·측면도를 CAD 도면(레이어 분리)으로, [3D모델] 은 3차원 "
            "형상을 CAD 메시로, [도해] 는 이 화면 4 장을 이미지로 내보낸다.")

        _figs = build_figs()
        _view = st.radio("보기", ["3D 형상", "평면도", "정면도", "측면도"],
                         horizontal=True, key='fig_view',
                         label_visibility='collapsed')
        st.pyplot(_figs[{'3D 형상': '3d', '평면도': 'plan',
                         '정면도': 'front', '측면도': 'side'}[_view]])

# ======================================================== 9. 저장 · 불러오기
with TABS[8]:
    _ready = SS.get('eng') is not None
    g = box(" ① 저장할 항목 ")
    if not _ready:
        g.warning("※ 계산 결과가 없습니다. [▶ 계산 실행] 을 먼저 누르면 "
                  "계산서·도면·3D 도 저장할 수 있습니다.")
    _picks = []
    for _key, _label, _ext, _need, _desc in SAVE_ITEMS:
        _k = f"save_{_key}"
        if _k not in SS:
            SS[_k] = (_key == 'spec' or _ready)
        if _need and not _ready:
            SS[_k] = False
        c = g.columns([1.6, 4])
        c[0].checkbox(f"{_label}  ({_ext})", key=_k,
                      disabled=(_need and not _ready))
        c[1].markdown(f"<div style='padding-top:6px;color:#7F8C8D'>{_desc}</div>",
                      unsafe_allow_html=True)
        if SS[_k]:
            _picks.append(_key)
    bc = g.columns([1, 1, 4])
    bc[0].button("모두 선택", key='btn_save_all', use_container_width=True,
                 on_click=set_all_saves, args=(True,))
    bc[1].button("모두 해제", key='btn_save_none', use_container_width=True,
                 on_click=set_all_saves, args=(False,))

    g = box(" ② 이름 ")
    if 'save_name' not in SS:
        SS['save_name'] = ''
    g.text_input("파일 이름 (앞머리)", key='save_name',
                 placeholder=safe_name(gs('name')),
                 help="'이름_항목.확장자' 로 만들어집니다. 비워 두면 케이슨명을 씁니다.")
    _nm = safe_name(SS.get('save_name') or gs('name'))

    g = box(" ③ 만들어질 파일 ")
    _plan = [p for k in _picks for p in save_targets(k, _nm)]
    g.code("\n".join(_plan) if _plan else "고른 항목이 없습니다.", language=None)

    if st.button("파일 만들기", type='primary', disabled=not _picks,
                 key='btn_make'):
        with st.spinner("만드는 중 … (3D 모델은 몇 초 걸립니다)"):
            make_files(_picks, _nm)

    _fl = SS.get('files') or {}
    if _fl.get('items') or _fl.get('err'):
        g = box(" ④ 내려받기 ")
        for _m in _fl.get('ok', []):
            g.success(_m)
        for _m in _fl.get('err', []):
            g.error(_m)
        _items = _fl.get('items', [])
        if _items:
            g.download_button("전체 ZIP 으로 내려받기", data=zip_bytes(_items),
                              file_name=f"{_fl.get('name', _nm)}.zip",
                              mime='application/zip', key='dl_zip',
                              use_container_width=True, type='primary')
            dc = g.columns(2)
            for _i, (_fname, _blob) in enumerate(_items):
                dc[_i % 2].download_button(
                    _fname, data=_blob, file_name=_fname, key=f'dl_{_i}',
                    mime='application/octet-stream', use_container_width=True)
        if SS.get('last3d_text'):
            with g.expander("3D 모델 ↔ 계산서 체적 대조"):
                st.code(SS['last3d_text'], language=None)

    g = box(" 불러오기 ")
    g.markdown("왼쪽 사이드바의 **[불러오기 (.cais)]** 로 제원 파일을 올리십시오. "
               "원본 Tkinter 판이 저장한 `.cais` 파일을 그대로 읽습니다.")
